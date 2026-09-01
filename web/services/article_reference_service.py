from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from web.portal_db import connect


def _valid_url(value):
    parsed = urlparse(str(value or "").strip())
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def import_reference_workbook(path):
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    imported = 0
    try:
        sheet = next((item for item in workbook.worksheets if "article" in item.title.casefold()
                      and "url" in item.title.casefold()), workbook.active)
        current_type = "General Article"
        with connect() as db:
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if not values:
                    continue
                urls = [value for value in values if _valid_url(value)]
                if not urls:
                    heading = values[0]
                    if "sample url" not in heading.casefold():
                        current_type = heading
                    continue
                for url in urls:
                    cursor = db.execute(
                        """INSERT OR IGNORE INTO article_references(
                           label,url,module,article_type,user_state,source_file)
                           VALUES(?,?,?,?,?,?)""",
                        (current_type, url.strip(), "Article Page", current_type, "ANY", path.name),
                    )
                    imported += cursor.rowcount
    finally:
        workbook.close()
    return imported


def list_references(query="", article_type="", user_state="", include_archived=False):
    clauses = [] if include_archived else ["active=1"]
    parameters = []
    if query:
        clauses.append("(label LIKE ? OR url LIKE ? OR notes LIKE ?)")
        parameters.extend([f"%{query}%"] * 3)
    if article_type:
        clauses.append("article_type=?")
        parameters.append(article_type)
    if user_state:
        clauses.append("user_state=?")
        parameters.append(user_state)
    where = " WHERE " + " AND ".join(clauses) if clauses else ""
    with connect() as db:
        rows = db.execute(
            f"SELECT * FROM article_references{where} ORDER BY article_type,label,id",
            parameters,
        ).fetchall()
        types = [row[0] for row in db.execute(
            "SELECT DISTINCT article_type FROM article_references WHERE active=1 ORDER BY article_type"
        ).fetchall()]
    return [dict(row) for row in rows], types


def save_reference(reference_id, label, url, article_type, user_state, notes=""):
    label, url = str(label or "").strip(), str(url or "").strip()
    article_type = str(article_type or "General Article").strip()
    state = str(user_state or "ANY").strip().upper()
    if not label or not _valid_url(url):
        raise ValueError("A label and valid HTTP/HTTPS article URL are required.")
    if state not in {"ANY", "ANONYMOUS", "SUBSCRIBER", "REGISTERED_USER", "EXPIRED_USER"}:
        raise ValueError("Select a valid user state.")
    with connect() as db:
        if reference_id:
            db.execute(
                """UPDATE article_references SET label=?,url=?,article_type=?,user_state=?,notes=?,
                   active=1,updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (label, url, article_type, state, str(notes or "").strip(), reference_id),
            )
            return reference_id
        return db.execute(
            """INSERT INTO article_references(label,url,module,article_type,user_state,notes)
               VALUES(?,?,'Article Page',?,?,?)""",
            (label, url, article_type, state, str(notes or "").strip()),
        ).lastrowid


def get_reference(reference_id):
    with connect() as db:
        row = db.execute("SELECT * FROM article_references WHERE id=?", (reference_id,)).fetchone()
    return dict(row) if row else None


def archive_reference(reference_id):
    with connect() as db:
        db.execute("UPDATE article_references SET active=0,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                   (reference_id,))

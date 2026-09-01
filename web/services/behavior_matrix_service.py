"""Import the approved functionality matrix into generation-time bot memory."""

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from web.portal_db import connect


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_MATRIX = ROOT / "Uploads" / "Source" / "THG App_Functionality_Matrix.xlsx"
STATE_COLUMNS = {
    "Anonymous": "ANONYMOUS",
    "Non-Subscriber": "REGISTERED_USER",
    "Subscriber": "SUBSCRIBER",
}


def _slug(value):
    return re.sub(r"[^A-Z0-9]+", "_", str(value).upper()).strip("_")


def _guidance(result, feature, expected):
    normalized = str(result or "").casefold()
    if "not appear" in normalized or normalized.startswith("not "):
        assertion = "Use assertNotVisible for the validated selector or text."
    elif "appear" in normalized:
        assertion = "Use assertVisible for the validated selector or text."
    else:
        assertion = "Use a strict runtime assertion matching the state-specific matrix result."
    conditional = ""
    if any(token in str(expected).casefold() for token in ("metering", "if an article", "more than 1000")):
        conditional = " Use controlled article test data; do not convert conditional eligibility into a universal assertion."
    return f"{assertion} Assert the {feature} outcome for this user state.{conditional}"


def import_behavior_matrix(path=DEFAULT_MATRIX):
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["TH App"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    columns = {header: index for index, header in enumerate(headers)}
    current_module = ""
    imported = []
    with connect() as db:
        for row_number, row in enumerate(rows[1:], start=2):
            module = str(row[columns["Module"]] or "").strip()
            if module:
                current_module = module
            feature = str(row[columns["Feature , Functionality"]] or "").strip()
            expected = str(row[columns["Expected Behaviour"]] or "").strip()
            if not feature or not expected:
                continue
            trigger_terms = list(dict.fromkeys(
                token for token in re.findall(r"[A-Za-z0-9_-]+", f"{current_module} {feature}")
                if len(token) > 2
            ))
            for column, state in STATE_COLUMNS.items():
                result = str(row[columns[column]] or "").strip()
                if not result:
                    continue
                rule_id = f"FM_R{row_number:03d}_{_slug(state)}"
                behavior = (
                    f"Functionality Matrix row {row_number}: {result}. "
                    f"Authoritative behavior: {expected}"
                )
                db.execute(
                    """INSERT INTO app_behavior_rules(
                       rule_id,user_state,intent,trigger_terms,expected_behavior,
                       yaml_guidance,status,updated_at)
                       VALUES(?,?,?,?,?,?, 'approved',CURRENT_TIMESTAMP)
                       ON CONFLICT(rule_id) DO UPDATE SET
                         user_state=excluded.user_state,intent=excluded.intent,
                         trigger_terms=excluded.trigger_terms,
                         expected_behavior=excluded.expected_behavior,
                         yaml_guidance=excluded.yaml_guidance,status='approved',
                         updated_at=CURRENT_TIMESTAMP""",
                    (rule_id, state, f"{current_module}: {feature}",
                     json.dumps(trigger_terms), behavior,
                     _guidance(result, feature, expected)),
                )
                imported.append(rule_id)
    return {"source": str(path), "rules": len(imported), "rows": len(imported) // 3}


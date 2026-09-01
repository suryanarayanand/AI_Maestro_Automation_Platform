import json
import re
import uuid
from pathlib import Path

from openpyxl import load_workbook

from web.portal_db import connect


ROOT = Path(__file__).resolve().parents[2]
PUBLISHED = ROOT / "Scenarios" / "PublishedFlows"
PUBLISHED_SUITE = ROOT / "Suites" / "published_flows.json"


def infer_user_state(text):
    lowered = str(text or "").casefold()
    anonymous = any(word in lowered for word in ("anonymous", "logged out", "non-subscriber"))
    subscriber = any(word in lowered for word in ("subscriber", "subscribed", "premium user"))
    if anonymous and subscriber:
        return "MIXED"
    if subscriber:
        return "SUBSCRIBER"
    if anonymous:
        return "ANONYMOUS"
    return "ANONYMOUS"


def infer_tags(text, user_state):
    lowered = str(text or "").casefold()
    tags = {user_state.casefold(), "functional"}
    for tag, terms in {
        "login": ("login", "sign in"),
        "article": ("article",),
        "premium": ("premium", "paywall"),
        "trending": ("trending",),
        "hamburger": ("hamburger",),
        "search": ("search",),
        "account": ("account", "user menu"),
        "screenshot": ("screenshot",),
    }.items():
        if any(term in lowered for term in terms):
            tags.add(tag)
    return sorted(tags)


def import_catalog(path):
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    imported = 0
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().casefold() for value in next(rows)]
        lookup = {name: index for index, name in enumerate(headers)}
        if not {"s.no", "description"}.issubset(lookup):
            raise ValueError("Atomic import expects S.No and Description columns.")
        component_index = lookup.get("component/module")
        scenario_index = lookup.get("test scenario")
        with connect() as db:
            for source_row, values in enumerate(rows, start=2):
                serial = values[lookup["s.no"]]
                description = str(values[lookup["description"]] or "").strip()
                if serial is None or not description:
                    continue
                module = str(values[component_index] or "").strip() if component_index is not None else ""
                scenario = str(values[scenario_index] or "").strip() if scenario_index is not None else ""
                scenario = scenario or module or f"Scenario {serial}"
                case_id = f"THCAT_{int(serial):04d}" if str(serial).replace(".0", "").isdigit() else f"THCAT_R{source_row}"
                evidence = f"{module} {scenario} {description}"
                state = infer_user_state(evidence)
                source_text = f"User State: {state.title()}. Action: {description}"
                tags = infer_tags(evidence, state)
                cursor = db.execute(
                    """INSERT OR IGNORE INTO atomic_flow_steps(
                       case_id,scenario,step_number,source_text,user_state,module,tags,
                       source_file,source_row) VALUES(?,?,?,?,?,?,?,?,?)""",
                    (case_id, scenario, 1, source_text, state, module,
                     json.dumps(tags), path.name, source_row),
                )
                imported += cursor.rowcount
    finally:
        workbook.close()
    return imported


def create_manual_step(scenario, action, user_state, module="", tags_text=""):
    scenario = str(scenario or "").strip()
    action = str(action or "").strip()
    state = str(user_state or "").strip().upper()
    if not scenario or not action:
        raise ValueError("Scenario and automation requirement are required.")
    if state not in {"ANONYMOUS", "SUBSCRIBER", "MIXED"}:
        raise ValueError("Select a valid user state.")
    case_id = f"MANUAL_{uuid.uuid4().hex[:10].upper()}"
    tags = sorted({
        *infer_tags(f"{scenario} {action}", state),
        *(tag.strip().casefold() for tag in str(tags_text or "").split(",") if tag.strip()),
    })
    with connect() as db:
        cursor = db.execute(
            """INSERT INTO atomic_flow_steps(
               case_id,scenario,step_number,source_text,user_state,module,tags,source_file,source_row)
               VALUES(?,?,?,?,?,?,?,?,?)""",
            (case_id, scenario, 1, f"User State: {state.title()}. Action: {action}",
             state, str(module or "").strip(), json.dumps(tags), "manual-intake", None),
        )
        return cursor.lastrowid


def proposal_readiness(step, yaml_text=None):
    from web.services.yaml_editor_service import validate_maestro_yaml

    text = str(yaml_text if yaml_text is not None else step.get("proposal_yaml", ""))
    checks = []
    def add(name, passed, detail):
        checks.append({"name": name, "passed": bool(passed), "detail": detail})

    try:
        validate_maestro_yaml(text)
    except ValueError as exc:
        add("Maestro structure", False, str(exc))
    else:
        add("Maestro structure", True, "Header, separator, and commands are present.")
    command_count = len(re.findall(r"(?m)^\s*-\s+[A-Za-z][A-Za-z0-9]*", text))
    add("Complete flow", command_count >= 4, f"{command_count} executable commands found; at least 4 required.")

    repository = json.loads((ROOT / "LocatorRepository" / "validated_locator_repository.json").read_text(encoding="utf-8"))
    valid_ids = {item.get("locator", {}).get("value") for item in repository
                 if item.get("locator", {}).get("type") == "id"}
    used_ids = set(re.findall(r'(?m)^\s+id:\s*["\']?([^"\'\r\n]+)', text))
    missing_ids = sorted(used_ids - valid_ids)
    add("Validated locators", not missing_ids,
        "All ID selectors are validated." if not missing_ids else "Unvalidated: " + ", ".join(missing_ids))

    flow_paths = re.findall(r'(?m)^\s*-\s+runFlow:\s*["\']?([^"\'\r\n]+\.yaml)', text)
    base = ROOT / "Scenarios" / "PublishedFlows"
    missing_flows = [path for path in flow_paths if not (base / Path(path.replace("/", "\\"))).resolve().is_file()]
    add("Reusable flow paths", not missing_flows,
        "All referenced flows exist." if not missing_flows else "Missing: " + ", ".join(missing_flows))

    state = str(step.get("user_state", "")).upper()
    if state == "SUBSCRIBER":
        entitlement = bool(re.search(r'assertNotVisible:\s*\n\s+text:\s*"SUBSCRIBE"', text)) and bool(
            re.search(r'assertNotVisible:\s*\n\s+text:\s*"ADVERTISEMENT"', text))
        detail = "Subscriber flow asserts that SUBSCRIBE and ADVERTISEMENT are absent."
    elif state == "ANONYMOUS":
        entitlement = bool(re.search(r'assertVisible:\s*\n\s+text:\s*"SUBSCRIBE"', text)) and bool(
            re.search(r'assertVisible:\s*\n\s+text:\s*"ADVERTISEMENT"', text))
        detail = "Anonymous flow asserts that SUBSCRIBE and ADVERTISEMENT are visible."
    else:
        entitlement, detail = False, "Choose ANONYMOUS or SUBSCRIBER before publishing."
    add("User-state contract", entitlement, detail)
    add("Source provenance", "# Generated from:" in text, "The proposal retains its source requirement.")
    return {"ready": bool(checks) and all(check["passed"] for check in checks), "checks": checks}


def proposal_guidance(step):
    """Explain how repository knowledge and past runs affect this proposal."""
    from web.services.adaptive_test_agent import AdaptiveTestAgent

    evidence = AdaptiveTestAgent().retrieve(
        [step.get("scenario", ""), step.get("source_text", "")],
        limit=20, user_state=step.get("user_state", ""),
    )
    passed_flows = [item for item in evidence if item.get("evidence_type") == "passed_repository_flow"]
    all_failures = [item for item in evidence if item.get("evidence_type") == "negative_learning"]
    failures = [item for item in all_failures if item.get("payload", {}).get("component") != "device_transport"]
    environment_failures = [item for item in all_failures if item.get("payload", {}).get("component") == "device_transport"]
    validated = [item for item in evidence if item.get("validated")]
    readiness = proposal_readiness(step) if step.get("proposal_yaml") else {"ready": False, "checks": []}
    suggestions = []
    if passed_flows:
        suggestions.append(
            "Reuse the command ordering from passed flow(s): "
            + ", ".join(item["path"] for item in passed_flows[:3]) + "."
        )
    else:
        suggestions.append(
            "No semantically matching passed YAML was found; execute this proposal before treating it as reusable."
        )
    for item in failures[:3]:
        payload = item.get("payload", {})
        reason = payload.get("root_cause") or payload.get("error") or payload.get("failed_step")
        if reason:
            suggestions.append(f"Avoid past failure from {item.get('case_id') or 'a related case'}: {reason}")
    if environment_failures:
        suggestions.append(
            f"{len(environment_failures)} related run failure(s) were device/environment issues; "
            "they do not reduce YAML design confidence."
        )
    for check in readiness.get("checks", []):
        if not check["passed"]:
            suggestions.append(f"Fix {check['name']}: {check['detail']}")
    confidence = min(100, (45 if readiness["ready"] else 20)
                     + min(25, len(passed_flows) * 8)
                     + min(20, len(validated) * 2)
                     - min(15, len(failures) * 3))
    return {
        "confidence": max(0, confidence), "suggestions": suggestions,
        "passed_flows": passed_flows[:5], "failures": failures[:5],
        "evidence_count": len(evidence), "validated_count": len(validated),
    }


def list_steps(query="", state="", status="", limit=200):
    clauses, parameters = [], []
    if query:
        clauses.append("(case_id LIKE ? OR scenario LIKE ? OR source_text LIKE ?)")
        parameters.extend([f"%{query}%"] * 3)
    if state:
        clauses.append("user_state=?")
        parameters.append(state.upper())
    if status:
        clauses.append("status=?")
        parameters.append(status)
    where = " WHERE " + " AND ".join(clauses) if clauses else ""
    with connect() as db:
        rows = db.execute(
            f"SELECT * FROM atomic_flow_steps{where} ORDER BY id DESC LIMIT ?",
            (*parameters, limit),
        ).fetchall()
    return [{**dict(row), "tags_list": json.loads(row["tags"] or "[]")} for row in rows]


def get_step(step_id):
    with connect() as db:
        row = db.execute("SELECT * FROM atomic_flow_steps WHERE id=?", (step_id,)).fetchone()
    if not row:
        return None
    step = dict(row)
    step["tags_list"] = json.loads(step["tags"] or "[]")
    return step


def update_step_context(step_id, user_state, tags_text):
    state = str(user_state or "").strip().upper()
    if state not in {"ANONYMOUS", "SUBSCRIBER", "MIXED"}:
        raise ValueError("Select a valid user state.")
    step = get_step(step_id)
    if not step:
        raise ValueError("Atomic step not found.")
    tags = sorted({tag.strip().casefold() for tag in str(tags_text or "").split(",") if tag.strip()})
    if state.casefold() not in tags:
        tags.append(state.casefold())
        tags.sort()
    source_text = re.sub(
        r"^User State:\s*[^.]+\.", f"User State: {state.title()}.",
        step["source_text"], count=1, flags=re.I,
    )
    with connect() as db:
        db.execute(
            """UPDATE atomic_flow_steps SET user_state=?,tags=?,source_text=?,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (state, json.dumps(tags), source_text, step_id),
        )


def generate_proposal(step_id):
    from web.services.atomic_flow_composer import compose_atomic_flow

    step = get_step(step_id)
    if not step:
        raise ValueError("Atomic step not found.")
    try:
        yaml_text = compose_atomic_flow(step)
        error = None
        status = "proposed"
    except Exception as exc:
        yaml_text = ""
        error, status = str(exc), "needs_input"
    with connect() as db:
        db.execute(
            """UPDATE atomic_flow_steps SET proposal_yaml=?,status=?,error=?,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (yaml_text, status, error, step_id),
        )
    return get_step(step_id)


def save_proposal(step_id, yaml_text):
    from web.services.yaml_editor_service import validate_maestro_yaml
    validate_maestro_yaml(yaml_text)
    with connect() as db:
        db.execute(
            """UPDATE atomic_flow_steps SET proposal_yaml=?,status='edited',error=NULL,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (yaml_text, step_id),
        )


def publish_proposal(step_id, yaml_text, reviewer):
    step = get_step(step_id)
    readiness = proposal_readiness(step, yaml_text)
    if not readiness["ready"]:
        failed = "; ".join(check["detail"] for check in readiness["checks"] if not check["passed"])
        raise ValueError("Proposal is not publish-ready: " + failed)
    save_proposal(step_id, yaml_text)
    step = get_step(step_id)
    safe_id = re.sub(r"[^A-Za-z0-9_-]", "_", f"{step['case_id']}_STEP_{step['step_number']:02d}")
    relative = f"PublishedFlows/{safe_id}.yaml"
    destination = ROOT / "Scenarios" / relative
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(yaml_text, encoding="utf-8")
    with connect() as db:
        db.execute(
            """INSERT INTO published_flows(
               atomic_step_id,flow_id,title,yaml_path,tags,user_state,published_by)
               VALUES(?,?,?,?,?,?,?)
               ON CONFLICT(atomic_step_id) DO UPDATE SET
                 title=excluded.title,yaml_path=excluded.yaml_path,tags=excluded.tags,
                 user_state=excluded.user_state,published_by=excluded.published_by,
                 published_at=CURRENT_TIMESTAMP""",
            (step_id, safe_id, step["scenario"], relative, step["tags"],
             step["user_state"], reviewer),
        )
        db.execute(
            "UPDATE atomic_flow_steps SET status='published',updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (step_id,),
        )
        flows = db.execute("SELECT * FROM published_flows ORDER BY id").fetchall()
    tests = [
        {"id": row["flow_id"], "module": "Published Flows", "priority": "P2",
         "name": row["title"], "yaml": row["yaml_path"]}
        for row in flows
    ]
    PUBLISHED_SUITE.write_text(
        json.dumps({"suite": "Published Atomic Flows", "tests": tests}, indent=2) + "\n",
        encoding="utf-8",
    )
    return relative

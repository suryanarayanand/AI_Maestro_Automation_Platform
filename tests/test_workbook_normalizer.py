import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from generation.workbook_normalizer import WorkbookNormalizer


class WorkbookNormalizerTests(unittest.TestCase):
    def _workbook(self, directory, headers, rows, name="input.xlsx"):
        path = Path(directory) / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path

    def test_splits_multiline_canonical_steps_and_uses_source_name(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self._workbook(directory, ["test_case_id", "source_name", "step"], [
                ["TH_1", "Launch", "Launch app\nVerify Home\nCapture screenshot Home"],
            ])
            result = WorkbookNormalizer().normalize(source)
            self.assertEqual((result.case_count, result.step_count, result.source_format), (1, 3, "canonical"))
            workbook = load_workbook(result.canonical_path, read_only=True, data_only=True)
            self.assertEqual(workbook.active[2][1].value, "Launch")
            workbook.close()

    def test_accepts_legacy_headers_and_continuation_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self._workbook(directory, ["Scenario No", "Test Scenario", "Test Steps"], [
                ["SC_1", "Home", "Open Home"],
                [None, None, "Verify Home"],
            ])
            result = WorkbookNormalizer().normalize(source)
            self.assertEqual((result.case_count, result.step_count, result.source_format), (1, 2, "legacy"))

    def test_finds_headers_after_title_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self._workbook(directory, ["Mobile regression scenarios"], [
                ["Owner", "QA Team"],
                ["TC ID", "Test Case Name", "Action"],
                ["LOGIN-1", "Valid login", "Open the login page\nEnter valid credentials"],
            ])
            result = WorkbookNormalizer().normalize(source)
            self.assertEqual((result.case_count, result.step_count), (1, 2))

    def test_generates_ids_for_named_scenarios(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self._workbook(directory, ["Scenario Description", "Step Description"], [
                ["Valid login", "Open login"],
                [None, "Enter valid credentials"],
                ["Invalid login", "Enter invalid credentials"],
            ])
            result = WorkbookNormalizer().normalize(source)
            self.assertEqual((result.case_count, result.step_count), (2, 3))
            workbook = load_workbook(result.canonical_path, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
            workbook.close()
            self.assertEqual([row[0] for row in rows], ["TC_001", "TC_001", "TC_002"])

    def test_rejects_requirement_only_workbook_with_clear_routing(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "requirements.xlsx"
            workbook = Workbook()
            workbook.active.title = "Cases"
            workbook.active.append(["test_case_id", "name"])
            workbook.create_sheet("Validation Points").append(["test_case_id", "validation_point"])
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "AI Scenario Expander"):
                WorkbookNormalizer().normalize(path)


if __name__ == "__main__":
    unittest.main()

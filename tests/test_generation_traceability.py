import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from web.services.generation_service import (_generation_items, _ordered_yaml, _requirement_input,
                                             traceability_coverage)
from convert_to_supported_excel import convert
from yaml_generator import YAMLGenerator
from excel_reader import ExcelReader


class GenerationTraceabilityTests(unittest.TestCase):
    def test_traceability_coverage_requires_real_rows(self):
        self.assertEqual(traceability_coverage([]), 0.0)
        self.assertEqual(traceability_coverage([
            {"status": "covered"}, {"status": "covered"}, {"status": "incomplete"},
        ]), 2 / 3)
        self.assertEqual(traceability_coverage('[{"status":"covered"},{"status":"incomplete"}]'), .5)

    def test_advertisement_element_column_is_converted_to_text_locator(self):
        positive = {"expected_result": "ADVERTISEMENT is visible", "element_id": "ADVERTISEMENT"}
        negative = {"expected_result": "ADVERTISEMENT is not visible", "element_id": "ADVERTISEMENT"}
        self.assertEqual(_requirement_input(positive), "ASSERT_VISIBLE_TEXT(ADVERTISEMENT)")
        self.assertEqual(_requirement_input(negative), "ASSERT_NOT_VISIBLE_TEXT(ADVERTISEMENT)")

    def test_negative_expected_result_becomes_assertion_obligation(self):
        case = {
            "id": "TC_NEG", "name": "Subscriber header", "steps": ["Launch app"],
            "requirements": [{
                "step_number": 1, "step": "Verify header",
                "expected_result": "Subscribe button must not be visible",
                "source_sheet": "Cases", "source_row": 2,
            }],
        }
        yaml, _, confidence, assumptions, trace = _ordered_yaml(YAMLGenerator(), case, False)
        self.assertIn('assertNotVisible:', yaml)
        self.assertIn('text: "Subscribe"', yaml)
        expected = [item for item in trace if item["source_type"] == "expected_result"]
        self.assertEqual(expected[0]["status"], "covered")
        self.assertEqual(confidence, 1.0)
        self.assertEqual(assumptions, [])
        self.assertEqual(
            expected[0]["selector_grounding"][0]["source"], "stable_text_fallback"
        )

    def test_unsupported_expected_result_is_explicitly_incomplete(self):
        case = {
            "id": "TC_GAP", "name": "Gap", "steps": ["Launch app"],
            "requirements": [{"step_number": 1, "step": "Login",
                              "expected_result": "Subscriber login is successful"}],
        }
        _, _, confidence, assumptions, trace = _ordered_yaml(YAMLGenerator(), case, False)
        expected = [item for item in trace if item["source_type"] == "expected_result"]
        self.assertEqual(expected[0]["status"], "incomplete")
        self.assertLess(confidence, 1.0)
        self.assertTrue(any("not converted" in item for item in assumptions))

    def test_converter_preserves_architecture_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "supported.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Test Case ID", "Test Step", "Expected Result", "User State",
                          "Element ID", "Automatable", "Priority"])
            sheet.append(["TC_1", "Verify home", "Home page is displayed", "SUBSCRIBER",
                          "screen_home", "Yes", "P1"])
            workbook.save(source)
            convert(source, output)
            converted = load_workbook(output, read_only=True, data_only=True)
            headers = [cell.value for cell in converted.active[1]]
            values = dict(zip(headers, [cell.value for cell in converted.active[2]]))
            converted.close()
            self.assertEqual(values["user_state"], "SUBSCRIBER")
            self.assertEqual(values["element_id"], "screen_home")
            self.assertEqual(values["automatable"], "Yes")
            self.assertEqual(values["priority"], "P1")

    def test_canonical_expected_result_header_is_preserved(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "supported.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["test_case_id", "step", "expected_result"])
            sheet.append(["TC_1", "ASSERT_VISIBLE(screen_home)", "Home is visible"])
            workbook.save(source)
            convert(source, output)
            converted = load_workbook(output, read_only=True, data_only=True)
            headers = [cell.value for cell in converted.active[1]]
            values = dict(zip(headers, [cell.value for cell in converted.active[2]]))
            converted.close()
            self.assertEqual(values["expected_result"], "Home is visible")

    def test_yaml_command_reference_does_not_duplicate_canonical_step(self):
        case = {
            "id": "SIMPLE_002", "name": "Menu",
            "steps": ["TAP(nav_menu)"],
            "requirements": [{
                "step": "TAP(nav_menu)", "expected_result": "",
                "automation_intent": "", "yaml_command": "tapOn: {id: nav_menu}",
            }],
        }
        items = _generation_items(case)
        self.assertEqual([item["text"] for item in items], ["TAP(nav_menu)"])

    def test_yaml_mapped_expected_result_does_not_create_false_assertion(self):
        case = {
            "id": "SIMPLE_002", "name": "Menu",
            "steps": ["TAP(nav_menu)"],
            "requirements": [{
                "step": "TAP(nav_menu)",
                "expected_result": "Hamburger menu button is tapped.",
                "element_id": "nav_menu", "yaml_command": "tapOn: {id: nav_menu}",
            }],
        }
        items = _generation_items(case)
        self.assertEqual([item["text"] for item in items], ["TAP(nav_menu)"])

    def test_descriptive_automation_intent_is_not_an_extra_requirement(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "canonical.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["test_case_id", "step", "automation_intent", "yaml_command"])
            sheet.append(["SIMPLE_002", "TAP(nav_menu)",
                          "One explicit action becomes one command.",
                          "tapOn: {id: nav_menu}"])
            workbook.save(path)
            case = ExcelReader().group_cases(path)[0]
            self.assertEqual(case["requirements"], [])

    def test_indexed_text_tap_is_preserved(self):
        generated = YAMLGenerator().generate_test(
            ["TAP_TEXT_INDEX(Sport,1)"], case_id="INDEXED_TEXT"
        )
        self.assertEqual(generated["steps"][0]["command"], "tapOn")
        self.assertEqual(generated["steps"][0]["parameters"]["text"], "Sport")
        self.assertEqual(generated["steps"][0]["parameters"]["index"], 1)

    def test_multiline_expected_results_are_separate_obligations(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "canonical.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["test_case_id", "step", "expected_result", "automation_intent"])
            sheet.append(["TC_1", "Login", "Login successful\nHome page is displayed\nSubscribe button must not be visible",
                          "ASSERT_NOT_VISIBLE(cta_subscribe)"])
            workbook.save(path)
            case = ExcelReader().group_cases(path)[0]
            self.assertEqual([r["expected_result"] for r in case["requirements"]], [
                "Login successful", "Home page is displayed", "Subscribe button must not be visible",
            ])
            self.assertEqual(case["requirements"][2]["automation_intent"],
                             "ASSERT_NOT_VISIBLE(cta_subscribe)")

    def test_unresolved_ai_verify_step_cannot_inject_commands(self):
        case = {"id": "TC_VERIFY", "name": "Article", "steps": [
            "Verify the article title, image, author/date, and article content."
        ], "requirements": []}
        with self.assertRaisesRegex(ValueError, "No Excel steps"):
            _ordered_yaml(YAMLGenerator(), case, False)

    def test_unrelated_assertion_cannot_cover_requirement(self):
        case = {"id": "TC_HEADER", "name": "Header", "steps": ["Launch app"], "requirements": [{
            "expected_result": "The Hindu logo, Subscribe button, and user icon are visible",
            "automation_intent": "",
        }]}
        _, _, confidence, assumptions, trace = _ordered_yaml(YAMLGenerator(), case, False)
        self.assertEqual(trace[-1]["status"], "incomplete")
        self.assertLess(confidence, 1.0)

    def test_click_requirement_needs_an_interaction_command(self):
        case = {"id": "TC_MENU", "name": "Menu", "steps": [
            "Click the hamburger menu and verify it opens"
        ], "requirements": []}
        with self.assertRaisesRegex(ValueError, "No Excel steps"):
            _ordered_yaml(YAMLGenerator(), case, False)

    def test_visual_alignment_requires_explicit_automation_evidence(self):
        case = {"id": "TC_LAYOUT", "name": "Layout", "steps": ["Launch app"], "requirements": [{
            "expected_result": "Primary navigation tabs are visible and spaced correctly",
            "automation_intent": "",
        }]}
        _, _, _, _, trace = _ordered_yaml(YAMLGenerator(), case, False)
        self.assertEqual(trace[-1]["status"], "incomplete")


if __name__ == "__main__":
    unittest.main()

import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from AI.ai_scenario_expander import (
    AIScenarioExpander,
    AutomationRepositoryCatalog,
    ExpansionBatchProcessor,
    NormalizedScenarioReader,
)


class FakeResponses:
    def create(self, **kwargs):
        case_id = "TH_0002" if "TH_0002" in kwargs["input"] else "TH_0001"
        result = {
            "test_case_id": case_id,
            "scenario": "Verify app launch after fresh installation",
            "test_steps": ["Launch the application", "Verify onboarding is visible", "Capture a screenshot"],
            "expected_results": ["The onboarding screen is displayed"],
            "automation_coverage": ["Application launch", "Onboarding verification"],
            "tags": ["Smoke", "OOBE"],
            "module": "Home",
            "common_flows": ["OOBE_terms_and_condition.yaml", "invented.yaml"],
            "locator_repositories": ["home", "invented_screen"],
            "confidence": 0.85,
            "unresolved_assumptions": [],
        }
        return SimpleNamespace(output_text=json.dumps(result))


class FakeCatalog:
    common_flows = ["OOBE_terms_and_condition.yaml"]
    locator_names = ["Home", "Games"]
    locator_repositories = ["home"]
    modules = ["Home"]


class ScenarioExpanderTests(unittest.TestCase):
    def test_expansion_is_structured_and_repository_grounded(self):
        client = SimpleNamespace(responses=FakeResponses())
        expander = AIScenarioExpander(client=client, model="test-model", catalog=FakeCatalog())
        result = expander.expand({
            "test_case_id": "TH_0002", "name": "App Launch after fresh install",
            "module": "Unassigned", "validation_points": ["Verify onboarding is shown"],
        })
        self.assertEqual(result["common_flows"], ["OOBE_terms_and_condition.yaml"])
        self.assertEqual(result["locator_repositories"], ["home"])
        self.assertEqual(len(result["unresolved_assumptions"]), 2)

    def test_reader_and_resumable_batch_output(self):
        with tempfile.TemporaryDirectory() as directory:
            directory = Path(directory)
            source = directory / "source.xlsx"
            workbook = Workbook()
            cases = workbook.active
            cases.title = "Cases"
            cases.append(["test_case_id", "name", "module"])
            cases.append(["TH_0002", "App Launch", "Unassigned"])
            points = workbook.create_sheet("Validation Points")
            points.append(["test_case_id", "validation_point"])
            points.append(["TH_0002", "Verify onboarding"])
            workbook.save(source)

            client = SimpleNamespace(responses=FakeResponses())
            processor = ExpansionBatchProcessor(
                expander=AIScenarioExpander(client=client, model="test-model", catalog=FakeCatalog()),
                reader=NormalizedScenarioReader(),
            )
            output, checkpoint = directory / "output.xlsx", directory / "progress.jsonl"
            first = processor.process(source, output, checkpoint)
            second = processor.process(source, output, checkpoint)
            self.assertEqual(len(first), 1)
            self.assertEqual(len(second), 1)
            self.assertEqual(len(checkpoint.read_text(encoding="utf-8").splitlines()), 1)
            generated = load_workbook(output, read_only=True)
            self.assertEqual(generated["Expanded Scenarios"].max_row, 2)
            generated.close()

    def test_real_catalog_contains_only_existing_assets(self):
        catalog = AutomationRepositoryCatalog()
        self.assertIn("OOBE_terms_and_condition.yaml", catalog.common_flows)
        self.assertTrue(catalog.locator_repositories)


if __name__ == "__main__":
    unittest.main()

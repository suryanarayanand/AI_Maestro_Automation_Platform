"""Detect uploaded test-design formats and produce canonical YAML-generator input."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


CANONICAL_HEADERS = ("test_case_id", "name", "step")
ID_ALIASES = (
    "test_case_id", "test case id", "testcase id", "test case no", "test case #",
    "tc id", "scenario id", "scenario no", "scenario_no", "sc id", "case id", "id",
)
NAME_ALIASES = (
    "name", "source_name", "test case", "test case name", "test case description",
    "test scenario", "scenario", "scenario name", "scenario description", "title",
)
STEP_ALIASES = (
    "step", "test step", "test steps", "test_steps", "steps", "step description",
    "action", "actions", "procedure", "automation step", "automation steps",
)


@dataclass(frozen=True)
class NormalizationResult:
    source_path: Path
    canonical_path: Path
    source_format: str
    case_count: int
    step_count: int


def _header(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _find_header(headers, aliases):
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    return None


def _split_steps(value):
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return []
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    return [re.sub(r"^\s*\d+[.)]\s*", "", line).strip() for line in lines]


class WorkbookNormalizer:
    """Normalize canonical, expanded, and legacy step-based Excel workbooks."""

    def normalize(self, source_path, output_path=None):
        source_path = Path(source_path)
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet, mapping, source_format = self._select_sheet(workbook)
            rows = sheet.iter_rows(min_row=mapping["header_row"] + 1, values_only=True)
            canonical_rows = []
            previous_id = ""
            previous_name = ""
            generated_ids = {}
            for source_row, values in enumerate(rows, start=mapping["header_row"] + 1):
                name = self._value(values, mapping["name"]) if mapping["name"] is not None else ""
                case_id = self._value(values, mapping["id"])
                if mapping["id"] is None and name:
                    key = name.casefold()
                    if key not in generated_ids:
                        generated_ids[key] = f"TC_{len(generated_ids) + 1:03d}"
                    case_id = generated_ids[key]
                if case_id:
                    previous_id = case_id
                    previous_name = name or case_id
                else:
                    case_id = previous_id
                    name = name or previous_name
                if not case_id:
                    continue
                for step in _split_steps(self._value(values, mapping["step"])):
                    canonical_rows.append((case_id, name or case_id, step, sheet.title, source_row))
            if not canonical_rows:
                raise ValueError("No test steps were found in the uploaded workbook.")
        finally:
            workbook.close()

        output_path = Path(output_path or source_path.with_name(f"{source_path.stem}_normalized.xlsx"))
        self._write(output_path, canonical_rows)
        return NormalizationResult(
            source_path=source_path,
            canonical_path=output_path,
            source_format=source_format,
            case_count=len({row[0] for row in canonical_rows}),
            step_count=len(canonical_rows),
        )

    def _select_sheet(self, workbook):
        failures = []
        for sheet in workbook.worksheets:
            for header_row, values in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True), start=1
            ):
                headers = [_header(value) for value in values]
                id_index = _find_header(headers, ID_ALIASES)
                step_index = _find_header(headers, STEP_ALIASES)
                name_index = _find_header(headers, NAME_ALIASES)
                # A named scenario can safely receive a generated canonical ID.
                if step_index is not None and (id_index is not None or name_index is not None):
                    canonical = headers[id_index] == "test_case_id" if id_index is not None else False
                    canonical = canonical and headers[step_index] == "step" and header_row == 1
                    expanded = headers[step_index] in {"test_steps", "automation steps"}
                    source_format = (
                        "canonical" if canonical else "scenario-with-generated-ids"
                        if id_index is None else "expanded" if expanded else "legacy"
                    )
                    return sheet, {
                        "id": id_index, "name": name_index, "step": step_index,
                        "header_row": header_row,
                    }, source_format
            failures.append(sheet.title)
        if "Cases" in workbook.sheetnames and "Validation Points" in workbook.sheetnames:
            raise ValueError(
                "This is a business-requirement workbook without executable test steps. "
                "Run AI Scenario Expander and review its steps before YAML generation."
            )
        raise ValueError(
            "Unsupported Excel format. Could not find a test-case table in the first 25 rows. "
            "Expected a step/action column plus either a case ID or scenario/name column."
        )

    @staticmethod
    def _value(values, index):
        if index is None or index >= len(values) or values[index] is None:
            return ""
        return str(values[index]).strip()

    @staticmethod
    def _write(path, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Test Cases"
        sheet.append([*CANONICAL_HEADERS, "source_sheet", "source_row"])
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="316FEA")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 42
        sheet.column_dimensions["C"].width = 100
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(".tmp.xlsx")
        workbook.save(temporary)
        os.replace(temporary, path)

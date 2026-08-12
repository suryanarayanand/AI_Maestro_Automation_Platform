import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from large_workbook_reader import LargeWorkbookReader


def normalize(source, output):
    groups = LargeWorkbookReader().read_groups(source)
    workbook = Workbook()
    cases = workbook.active
    cases.title = "Cases"
    headers = [
        "test_case_id", "name", "module", "validation_point_count",
        "automation_status", "source_serial",
    ]
    cases.append(headers)
    for group in groups:
        cases.append([
            group["id"], group["name"], group["module"],
            len(group["validation_points"]), "needs_step_design",
            group["source_serial"],
        ])

    points = workbook.create_sheet("Validation Points")
    point_headers = [
        "test_case_id", "name", "module", "source_serial", "validation_point"
    ]
    points.append(point_headers)
    for group in groups:
        for point in group["validation_points"]:
            points.append([
                group["id"], group["name"], group["module"],
                point["source_serial"], point["description"],
            ])

    for sheet in (cases, points):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="316FEA")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return groups


def main():
    parser = argparse.ArgumentParser(description="Normalize the TH App scenario workbook")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    groups = normalize(args.source, args.output)
    print(f"Cases: {len(groups)}")
    print(f"Validation points: {sum(len(item['validation_points']) for item in groups)}")
    print(args.output)


if __name__ == "__main__":
    main()

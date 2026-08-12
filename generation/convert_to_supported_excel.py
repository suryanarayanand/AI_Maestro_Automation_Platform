"""Convert test-case workbooks into the Excel format used by AI Maestro.

Usage:
    python generation/convert_to_supported_excel.py "input.xlsx"
    python generation/convert_to_supported_excel.py "input.xlsx" -o "output.xlsx"
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from workbook_normalizer import (
    ID_ALIASES,
    NAME_ALIASES,
    STEP_ALIASES,
    _find_header,
    _header,
    _split_steps,
)

EXPECTED_ALIASES = ("expected result", "expected results", "expected outcome")
INTENT_ALIASES = ("automation intent", "automation_intent", "automation command")


def _value(values, index):
    if index is None or index >= len(values) or values[index] is None:
        return ""
    return str(values[index]).strip()


def _find_table(sheet):
    """Return the header row and column mapping for a recognizable table."""
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True),
        start=1,
    ):
        headers = [_header(value) for value in values]
        mapping = {
            "id": _find_header(headers, ID_ALIASES),
            "name": _find_header(headers, NAME_ALIASES),
            "step": _find_header(headers, STEP_ALIASES),
            "expected": _find_header(headers, EXPECTED_ALIASES),
            "intent": _find_header(headers, INTENT_ALIASES),
        }
        if mapping["step"] is not None and (
            mapping["id"] is not None or mapping["name"] is not None
        ):
            return row_number, mapping
    return None


def _read_sheet(sheet, header_row, mapping):
    rows = []
    previous_id = previous_name = ""
    generated_ids = {}

    for source_row, values in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        case_id = _value(values, mapping["id"])
        name = _value(values, mapping["name"])

        if mapping["id"] is None and name:
            key = name.casefold()
            if key not in generated_ids:
                generated_ids[key] = f"TC_{len(generated_ids) + 1:03d}"
            case_id = generated_ids[key]

        if case_id:
            previous_id = case_id
            previous_name = name or case_id
        else:
            case_id = previous_id
            name = name or previous_name

        if not case_id:
            continue
        for step in _split_steps(_value(values, mapping["step"])):
            rows.append((
                case_id, name or case_id, step,
                _value(values, mapping["expected"]),
                _value(values, mapping["intent"]),
                sheet.title, source_row,
            ))
    return rows


def convert(source, output=None):
    """Convert every recognizable worksheet and return conversion statistics."""
    source = Path(source).resolve()
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise ValueError(f"Input must be an existing .xlsx file: {source}")

    output = Path(output).resolve() if output else source.with_name(
        f"{source.stem}_Supported_Format.xlsx"
    )
    if output == source:
        raise ValueError("Output must be different from the input file.")

    source_workbook = load_workbook(source, read_only=True, data_only=True)
    converted_rows = []
    converted_sheets = []
    try:
        for sheet in source_workbook.worksheets:
            table = _find_table(sheet)
            if not table:
                continue
            sheet_rows = _read_sheet(sheet, *table)
            if sheet_rows:
                converted_rows.extend(sheet_rows)
                converted_sheets.append(sheet.title)
    finally:
        source_workbook.close()

    if not converted_rows:
        raise ValueError(
            "No supported test-case table was found. A table needs a step/action "
            "column and either a case-ID or scenario/name column in its first 25 rows."
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"
    sheet.append([
        "test_case_id", "name", "step", "expected_result", "automation_intent",
        "source_sheet", "source_row",
    ])
    for row in converted_rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="316FEA")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 100

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    os.replace(temporary, output)

    return {
        "output": output,
        "sheets": len(converted_sheets),
        "cases": len({(row[5], row[0]) for row in converted_rows}),
        "steps": len(converted_rows),
    }


def main():
    parser = argparse.ArgumentParser(
        description="Convert test-case Excel sheets into AI Maestro's supported format."
    )
    parser.add_argument("input", help="Source .xlsx workbook")
    parser.add_argument("-o", "--output", help="Optional output .xlsx path")
    args = parser.parse_args()

    result = convert(args.input, args.output)
    print(f"Created: {result['output']}")
    print(
        f"Converted {result['cases']} cases and {result['steps']} steps "
        f"from {result['sheets']} worksheet(s)."
    )


if __name__ == "__main__":
    main()

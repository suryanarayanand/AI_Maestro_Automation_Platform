from pathlib import Path
from openpyxl import load_workbook


class ExcelReader:
    REQUIRED_COLUMNS = {"test_case_id", "step"}

    def read(self, path):
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value).strip().lower() if value is not None else "" for value in next(rows)]
            missing = self.REQUIRED_COLUMNS.difference(headers)
            if missing:
                raise ValueError(f"Missing Excel columns: {', '.join(sorted(missing))}")

            records = []
            for values in rows:
                record = dict(zip(headers, values))
                if record.get("test_case_id") and record.get("step"):
                    records.append(record)
            return records
        finally:
            workbook.close()

    def group_cases(self, path):
        cases = {}
        for row in self.read(path):
            case_id = str(row["test_case_id"]).strip()
            case = cases.setdefault(case_id, {
                "id": case_id, "name": row.get("name") or case_id, "steps": [],
                "expected_results": [], "automation_intents": [],
            })
            case["steps"].append(str(row["step"]).strip())
            expected = str(row.get("expected_result") or "").strip()
            intent = str(row.get("automation_intent") or "").strip()
            for value, key in ((expected, "expected_results"), (intent, "automation_intents")):
                if value and value not in case[key]:
                    case[key].append(value)
        return list(cases.values())

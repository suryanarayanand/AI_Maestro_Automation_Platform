from pathlib import Path

from openpyxl import load_workbook


class LargeWorkbookReader:
    """Read the TH App requirement workbook as grouped scenario blocks."""

    REQUIRED_COLUMNS = {
        "s.no", "component/module", "test scenario", "description"
    }

    @staticmethod
    def _normalize(value):
        return str(value).strip().lower() if value is not None else ""

    def read_groups(self, path):
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [self._normalize(value) for value in next(rows)]
        missing = self.REQUIRED_COLUMNS.difference(headers)
        if missing:
            raise ValueError(
                "Missing workbook columns: " + ", ".join(sorted(missing))
            )
        index = {name: headers.index(name) for name in self.REQUIRED_COLUMNS}

        groups = []
        current = None
        current_module = "Unassigned"

        for values in rows:
            serial = values[index["s.no"]]
            module = values[index["component/module"]]
            scenario = values[index["test scenario"]]
            description = values[index["description"]]

            if module is not None and str(module).strip():
                current_module = str(module).strip()

            if scenario is not None and str(scenario).strip():
                serial_text = str(serial).strip() if serial is not None else str(len(groups) + 1)
                numeric = "".join(character for character in serial_text if character.isdigit())
                case_id = f"TH_{int(numeric):04d}" if numeric else f"TH_{len(groups) + 1:04d}"
                current = {
                    "id": case_id,
                    "source_serial": serial,
                    "module": current_module,
                    "name": str(scenario).strip(),
                    "validation_points": [],
                }
                groups.append(current)

            if description is not None and str(description).strip():
                if current is None:
                    raise ValueError(
                        f"Description at source row {serial!r} has no Test Scenario"
                    )
                current["validation_points"].append({
                    "source_serial": serial,
                    "description": str(description).strip(),
                })

        workbook.close()
        return groups

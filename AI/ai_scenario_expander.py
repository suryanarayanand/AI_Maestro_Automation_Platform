"""Expand normalized business requirements into reviewable automation designs.

This module deliberately stops before Maestro YAML generation. Its output is a
master automation workbook that a reviewer can approve before the existing
YAML generator consumes it.
"""

from __future__ import annotations

import argparse
import json
import os
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "Uploads" / "TH_App_Normalized_1397_Cases.xlsx"
DEFAULT_OUTPUT = ROOT / "Uploads" / "TH_App_Expanded_Automation_Cases.xlsx"
DEFAULT_CHECKPOINT = ROOT / "Uploads" / "TH_App_Expanded_Automation_Cases.jsonl"


EXPANSION_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "test_case_id": {"type": "string"},
        "scenario": {"type": "string"},
        "test_steps": {"type": "array", "items": {"type": "string"}, "minItems": 1},
        "expected_results": {"type": "array", "items": {"type": "string"}, "minItems": 1},
        "automation_coverage": {"type": "array", "items": {"type": "string"}},
        "tags": {"type": "array", "items": {"type": "string"}},
        "module": {"type": "string"},
        "common_flows": {"type": "array", "items": {"type": "string"}},
        "locator_repositories": {"type": "array", "items": {"type": "string"}},
        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
        "unresolved_assumptions": {"type": "array", "items": {"type": "string"}},
    },
    "required": [
        "test_case_id", "scenario", "test_steps", "expected_results",
        "automation_coverage", "tags", "module", "common_flows",
        "locator_repositories", "confidence", "unresolved_assumptions",
    ],
}


class NormalizedScenarioReader:
    """Read the two-sheet normalized requirements workbook."""

    CASE_COLUMNS = {"test_case_id", "name", "module"}
    POINT_COLUMNS = {"test_case_id", "validation_point"}

    @staticmethod
    def _rows(sheet):
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        for values in rows:
            yield {headers[index]: value for index, value in enumerate(values) if index < len(headers)}

    def read(self, path: Path | str) -> list[dict[str, Any]]:
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
        try:
            if "Cases" not in workbook.sheetnames or "Validation Points" not in workbook.sheetnames:
                raise ValueError("Normalized workbook requires Cases and Validation Points sheets.")
            cases_sheet = workbook["Cases"]
            points_sheet = workbook["Validation Points"]
            case_headers = {str(cell.value).strip() for cell in cases_sheet[1] if cell.value}
            point_headers = {str(cell.value).strip() for cell in points_sheet[1] if cell.value}
            if missing := self.CASE_COLUMNS - case_headers:
                raise ValueError("Cases sheet is missing: " + ", ".join(sorted(missing)))
            if missing := self.POINT_COLUMNS - point_headers:
                raise ValueError("Validation Points sheet is missing: " + ", ".join(sorted(missing)))

            points: dict[str, list[str]] = defaultdict(list)
            for row in self._rows(points_sheet):
                case_id = str(row.get("test_case_id") or "").strip()
                point = str(row.get("validation_point") or "").strip()
                if case_id and point:
                    points[case_id].append(point)

            cases = []
            for row in self._rows(cases_sheet):
                case_id = str(row.get("test_case_id") or "").strip()
                if not case_id:
                    continue
                cases.append({
                    "test_case_id": case_id,
                    "name": str(row.get("name") or "").strip(),
                    "module": str(row.get("module") or "Unassigned").strip(),
                    "validation_points": points.get(case_id, []),
                })
            return cases
        finally:
            workbook.close()


class AutomationRepositoryCatalog:
    """Expose only repository assets that really exist in this project."""

    def __init__(self, root: Path = ROOT):
        self.root = Path(root)
        self.common_flows = sorted(path.name for path in (self.root / "Common").glob("*.yaml"))
        self.locator_names = self._locator_names()
        self.locator_repositories = self._locator_screens()
        self.modules = self._navigation_pages()

    def _read_json(self, name, fallback):
        path = self.root / "LocatorRepository" / name
        if not path.is_file():
            return fallback
        return json.loads(path.read_text(encoding="utf-8"))

    def _locator_screens(self):
        data = self._read_json("validated_locator_repository.json", [])
        return sorted({str(item.get("screen", "")).strip() for item in data if item.get("screen")}, key=str.casefold)

    def _locator_names(self):
        data = self._read_json("validated_locator_repository.json", [])
        return sorted(
            {str(item.get("name", "")).strip() for item in data if item.get("name")},
            key=str.casefold,
        )

    def _navigation_pages(self):
        data = self._read_json("navigation_repository.json", [])
        pages = {str(item.get("page", "")).strip() for item in data if item.get("page")}
        pages.update(screen.replace("_", " ").title() for screen in self.locator_repositories)
        return sorted(filter(None, pages), key=str.casefold)


class AIScenarioExpander:
    def __init__(self, client=None, model=None, catalog=None):
        self.client = client or OpenAI(
            api_key=os.getenv("OPENAI_API_KEY"),
            timeout=float(os.getenv("OPENAI_SCENARIO_TIMEOUT", "90")),
            max_retries=int(os.getenv("OPENAI_SCENARIO_RETRIES", "1")),
        )
        self.model = model or os.getenv("OPENAI_SCENARIO_MODEL", "gpt-5.6-sol")
        self.catalog = catalog or AutomationRepositoryCatalog()

    def _prompt(self, case):
        points = "\n".join(f"- {point}" for point in case["validation_points"])
        return f"""You are a Senior Mobile QA Automation Engineer designing reviewable Maestro automation scenarios.

Convert the business requirement into a structured automation design. Do not generate Maestro YAML and do not invent locators, screens, flows, credentials, test data, or backend capabilities.

Business requirement:
ID: {case['test_case_id']}
Title: {case['name']}
Current module: {case['module']}
Validation points:
{points or '- No validation point supplied'}

Available common flows (return exact filenames only):
{json.dumps(self.catalog.common_flows)}

Available locator repository screens (return exact names only):
{json.dumps(self.catalog.locator_repositories)}

Available validated locator names (use exact names only in Open, Tap, and Verify steps):
{json.dumps(getattr(self.catalog, 'locator_names', []))}

Known modules/pages:
{json.dumps(self.catalog.modules)}

Rules:
- Produce reusable, observable mobile test steps in execution order.
- Include preconditions as steps only when automation can perform them.
- Preserve the source steps one-for-one and in the same order wherever possible.
- Do not add a verification unless a supplied validation point explicitly asks to
  verify, assert, check, or confirm something visible.
- Do not add a screenshot unless a supplied validation point explicitly requests one.
- Navigation and tap steps must not be followed by an invented screen-loaded assertion.
- Preserve the supplied test case ID exactly.
- Use only listed common flows and locator repository screens.
- If required information or infrastructure is missing, record it in unresolved_assumptions.
- Confidence must reflect how executable the design is with the supplied information.
- Every test_steps item must use one of these executable sentence forms:
  Run <existing-flow.yaml>
  Launch application
  Relaunch application
  Navigate to <known page or section>
  Open <known control, game, page, or hamburger menu>
  Tap <known control>
  Verify <known page, screen, or visible control>
  Scroll
  Navigate back
  Capture screenshot <short evidence description>
- Keep each step to one action. Do not combine actions with "and" or conditional clauses.
- Do not include phrases such as "using the locator repository", "if supported",
  "when available", iteration matrices, external setup, comparisons, or backend actions
  in test_steps. Put those requirements in unresolved_assumptions instead.
- Refer to a common flow only as "Run <exact filename>".
- Use repository-visible names exactly and do not invent a Play Now locator or destination.
- Do not add a UI verification or interaction to test_steps unless its target appears
  in the available validated locator names or is one of the listed locator screens.
- Never add setup, navigation, assertions, waits, or evidence solely because they are
  considered automation best practice; every action must trace to supplied input.
- Move ungrounded logo, description, subtitle, highlighted-state, and destination
  checks into unresolved_assumptions instead of test_steps.
"""

    def expand(self, case):
        response = self.client.responses.create(
            model=self.model,
            input=self._prompt(case),
            text={
                "format": {
                    "type": "json_schema",
                    "name": "automation_scenario",
                    "strict": True,
                    "schema": EXPANSION_SCHEMA,
                }
            },
        )
        result = json.loads(response.output_text)
        return self._validate_and_ground(case, result)

    def _validate_and_ground(self, case, result):
        if result.get("test_case_id") != case["test_case_id"]:
            raise ValueError("AI response changed the test case ID.")
        for field in ("scenario", "test_steps", "expected_results"):
            if not result.get(field):
                raise ValueError(f"AI response has no {field}.")

        assumptions = list(result.get("unresolved_assumptions", []))
        allowed_flows = {item.casefold(): item for item in self.catalog.common_flows}
        allowed_repositories = {item.casefold(): item for item in self.catalog.locator_repositories}

        grounded_flows = []
        for value in result.get("common_flows", []):
            match = allowed_flows.get(str(value).casefold())
            if match:
                grounded_flows.append(match)
            else:
                assumptions.append(f"Suggested common flow does not exist: {value}")
        grounded_repositories = []
        for value in result.get("locator_repositories", []):
            match = allowed_repositories.get(str(value).casefold())
            if match:
                grounded_repositories.append(match)
            else:
                assumptions.append(f"Suggested locator repository screen does not exist: {value}")

        result["common_flows"] = list(dict.fromkeys(grounded_flows))
        result["locator_repositories"] = list(dict.fromkeys(grounded_repositories))
        result["unresolved_assumptions"] = list(dict.fromkeys(str(item).strip() for item in assumptions if str(item).strip()))
        result["source_name"] = case["name"]
        result["source_module"] = case["module"]
        result["validation_points"] = list(case["validation_points"])
        return result


class ExpansionBatchProcessor:
    def __init__(self, expander=None, reader=None):
        self.expander = expander or AIScenarioExpander()
        self.reader = reader or NormalizedScenarioReader()

    @staticmethod
    def _load_checkpoint(path):
        completed = {}
        if not path.is_file():
            return completed
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            record = json.loads(line)
            if record.get("test_case_id"):
                completed[record["test_case_id"]] = record
        return completed

    @staticmethod
    def _append_checkpoint(path, record):
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")
            handle.flush()
            os.fsync(handle.fileno())

    def process(self, source=DEFAULT_SOURCE, output=DEFAULT_OUTPUT, checkpoint=DEFAULT_CHECKPOINT,
                case_ids: Iterable[str] | None = None, limit: int | None = None, resume=True):
        source, output, checkpoint = Path(source), Path(output), Path(checkpoint)
        cases = self.reader.read(source)
        requested = {item.casefold() for item in case_ids or []}
        if requested:
            cases = [case for case in cases if case["test_case_id"].casefold() in requested]
        completed = self._load_checkpoint(checkpoint) if resume else {}
        pending = [case for case in cases if case["test_case_id"] not in completed]
        if limit is not None:
            pending = pending[:limit]

        for case in pending:
            try:
                record = self.expander.expand(case)
                record["expansion_status"] = "expanded"
                record["expansion_error"] = ""
            except Exception as exc:
                record = {
                    "test_case_id": case["test_case_id"], "source_name": case["name"],
                    "source_module": case["module"], "validation_points": case["validation_points"],
                    "expansion_status": "error", "expansion_error": str(exc),
                }
            self._append_checkpoint(checkpoint, record)
            completed[case["test_case_id"]] = record
            self.write_workbook(output, completed.values())
        return completed

    @staticmethod
    def write_workbook(path, records):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Expanded Scenarios"
        headers = [
            "test_case_id", "source_name", "scenario", "test_steps", "expected_results",
            "validation_points", "automation_coverage", "tags", "module", "common_flows",
            "locator_repositories", "confidence", "unresolved_assumptions",
            "expansion_status", "expansion_error", "review_status",
        ]
        sheet.append(headers)
        for record in sorted(records, key=lambda item: item.get("test_case_id", "")):
            def lines(name):
                value = record.get(name, [])
                return "\n".join(str(item) for item in value) if isinstance(value, list) else str(value or "")
            sheet.append([
                record.get("test_case_id", ""), record.get("source_name", ""), record.get("scenario", ""),
                lines("test_steps"), lines("expected_results"), lines("validation_points"),
                lines("automation_coverage"), ", ".join(record.get("tags", [])), record.get("module", ""),
                lines("common_flows"), lines("locator_repositories"), record.get("confidence", ""),
                lines("unresolved_assumptions"), record.get("expansion_status", ""),
                record.get("expansion_error", ""), "pending",
            ])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="316FEA")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.column_dimensions["A"].width = 14
        for column in "BCDEFGHIJKLMNO":
            sheet.column_dimensions[column].width = 32
        sheet.column_dimensions["P"].width = 14
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(".tmp.xlsx")
        workbook.save(temporary)
        os.replace(temporary, path)


def main():
    parser = argparse.ArgumentParser(description="Expand business requirements into automation scenarios")
    parser.add_argument("source", nargs="?", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--checkpoint", type=Path, default=DEFAULT_CHECKPOINT)
    parser.add_argument("--case-id", action="append", dest="case_ids")
    parser.add_argument("--limit", type=int)
    parser.add_argument("--no-resume", action="store_true")
    args = parser.parse_args()
    records = ExpansionBatchProcessor().process(
        args.source, args.output, args.checkpoint, args.case_ids, args.limit, not args.no_resume
    )
    expanded = sum(item.get("expansion_status") == "expanded" for item in records.values())
    errors = sum(item.get("expansion_status") == "error" for item in records.values())
    print(f"Expanded: {expanded}; errors: {errors}; output: {args.output}")


if __name__ == "__main__":
    main()

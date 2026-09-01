from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
READY = ROOT / "Uploads" / "Ready"
HEADERS = ["test_case_id","name","module","user_state","precondition","step_number","step","expected_result","priority","automatable","master_sheet_reference","existing_yaml_reference"]
PRE = "The Hindu app is installed; network is available; user is signed out; dismiss the welcome popup when shown. Handle a true interstitial by capturing it, waiting for Close, and closing only the ad."

HAM = [
("ANON_HAM_001","Hamburger open, selected state, and close","Blocker",[("Launch anonymously and tap the Hamburger icon.","The Hamburger menu opens and screen_hamburger is visible."),("Capture the menu, close it, and verify Home.","The drawer closes without closing the app and Home is restored.")]),
("ANON_HAM_002","Hamburger complete menu presentation","Major",[("Open Hamburger and scroll from top to bottom with screenshots.","All available section groups render without clipping or overlap."),("Verify the expected section labels are present.","The menu inventory is available for section navigation.")]),
("ANON_HAM_003","Primary news section navigation","Critical",[("Open India, World, Sport, News, and Business one at a time, returning to Hamburger after each.","Every selected primary section opens its matching page with article cards."),("Capture each destination title.","Primary-section navigation evidence is saved.")]),
("ANON_HAM_004","Specialist section navigation","Critical",[("Open Data, Health, Opinion, Science, Technology, and Society one at a time.","Every specialist section opens its matching page with content."),("Return to Hamburger after each destination.","The drawer remains reusable and navigation state is stable.")]),
("ANON_HAM_005","Culture and lifestyle section navigation","Critical",[("Open Entertainment, Lifestyle, Movies, Food, Children, Books, and Education one at a time.","Every culture/lifestyle section reaches the correct destination."),("Capture representative destinations.","Unique section coverage is recorded without repeating article checks.")]),
("ANON_HAM_006","Cities and States nested navigation","Critical",[("Open Hamburger and expand or select Cities and a visible city.","The selected city page opens with its title and content."),("Repeat for States and a visible state.","The selected state page opens and nested navigation works." )]),
("ANON_HAM_007","Videos section media presentation","Major",[("Open Videos from Hamburger.","The Videos destination opens with media cards."),("Open one media card and navigate Back.","The media destination responds and Videos is restored." )]),
("ANON_HAM_008","Representative section article and anonymous paywall","Critical",[("Open one representative section and its first article.","The article detail page opens."),("Scroll until a paywall or article-end content is reached.","A paywall shows its offer and Login, or unrestricted article-end content is documented." )]),
("ANON_HAM_009","Hamburger advertisement and interstitial handling","Critical",[("Scroll a representative section and open an article while observing advertisements.","Sticky ads do not fail the flow; problematic or blank ads are captured."),("If an interstitial appears, wait for Close and close only the ad.","The article or section is restored without closing the app." )]),
("ANON_HAM_010","Hamburger navigation recovery","Major",[("Open a section, open an article, then navigate Back to the section and Hamburger.","Each Back action restores the expected previous level."),("Close Hamburger and verify Home remains usable.","Navigation recovery completes without stale overlays or an app exit." )]),
]

ACCOUNT = [
("ANON_ACCOUNT_001","Anonymous Account page presentation","Blocker",[("Launch anonymously and tap Account.","The anonymous Account page opens."),("Verify Login, Create account, Subscribe, and application settings areas and capture the page.","Anonymous account actions and settings are visible." )]),
("ANON_ACCOUNT_002","Account Login destination","Critical",[("From Account, tap Login.","Login to your account opens with supported sign-in controls."),("Capture Login and navigate Back.","Account is restored without authentication." )]),
("ANON_ACCOUNT_003","Create Account destination","Critical",[("From Account, tap Create account.","The registration page opens."),("Capture it and navigate Back.","Account is restored and the user remains anonymous." )]),
("ANON_ACCOUNT_004","Account subscription-plan destination","Critical",[("From Account, tap Subscribe.","The subscription offer opens."),("Verify Yearly, Monthly, and Already a subscriber Login, then return.","Purchase and existing-subscriber routes are visible without purchasing." )]),
("ANON_ACCOUNT_005","Notification Inbox behavior","Major",[("From Account, open Notification Inbox.","The Inbox page or anonymous restriction opens."),("Capture the observed state and return.","The behavior is documented and Account is restored." )]),
("ANON_ACCOUNT_006","History anonymous behavior","Major",[("From Account, open History.","History shows its anonymous restriction or valid empty state."),("Capture it and return.","History does not expose another user's activity." )]),
("ANON_ACCOUNT_007","Bookmarks anonymous behavior","Critical",[("From Account, open Bookmarks.","Bookmarks shows a Login restriction or valid anonymous empty state."),("Capture it and return.","Anonymous access does not expose saved content." )]),
("ANON_ACCOUNT_008","Text Size controls and restoration","Major",[("Open Text Size and exercise decrease and increase controls.","Text-size controls respond and remain readable."),("Restore the original size and return.","The test leaves the display setting unchanged." )]),
("ANON_ACCOUNT_009","Appearance mode and restoration","Major",[("Open Appearance and switch from the current mode to the alternate mode.","The portal-supported Light or Dark appearance is applied without unreadable content."),("Capture it, restore the original mode, and return.","The test leaves the appearance setting unchanged." )]),
("ANON_ACCOUNT_010","Account full-page scrolling and layout","Major",[("Scroll Account from top to bottom with top, middle, and lower screenshots.","All account and application-setting rows remain readable and tappable." )]),
("ANON_ACCOUNT_011","Account back-navigation recovery","Major",[("Open Login, Notification Inbox, and one application setting, returning after each.","Every Back action restores Account rather than closing the app." )]),
("ANON_ACCOUNT_012","Anonymous Account access-boundary summary","Blocker",[("Verify Login and Create account routes without authenticating.","The session remains anonymous."),("Verify Subscribe reaches plans and History/Bookmarks do not expose private content.","Authentication, entitlement, and private-data boundaries remain enforced." )]),
]

def build(path, title, cases, module, yaml_ref):
    wb=Workbook(); ws=wb.active; ws.title=title; ws.append(HEADERS)
    for cid,name,priority,steps in cases:
        for n,(step,expected) in enumerate(steps,1): ws.append([cid,name,module,"ANONYMOUS",PRE,n,step,expected,priority,"Yes","Relevant scenarios in TH App Testing Scenarios_AutomationCopy.xlsx",yaml_ref])
    style(ws); wb.save(path)

def style(ws):
    for c in ws[1]: c.fill=PatternFill("solid",fgColor="273C75"); c.font=Font(color="FFFFFF",bold=True)
    for i,w in enumerate([22,50,20,16,82,13,90,90,14,14,55,80],1): ws.column_dimensions[ws.cell(1,i).column_letter].width=w
    for row in ws.iter_rows():
        for c in row: c.alignment=Alignment(vertical="top",wrap_text=True)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions

def main():
    READY.mkdir(parents=True,exist_ok=True)
    ham=READY/"Anonymous_Hamburger_Approved_Test_Cases.xlsx"; acc=READY/"Anonymous_Account_Settings_Approved_Test_Cases.xlsx"
    build(ham,"Anonymous Hamburger",HAM,"Hamburger","HamburgerOptions/HAM_OPT_*.yaml; CPLX_HAM_*.yaml; SC_48_Games_page_verification.yaml")
    build(acc,"Anonymous Account",ACCOUNT,"Account Settings","SC_33_Anonymous_Account_Settings.yaml; SC_49_Check_account_settings_for_anonymousAccount.yaml")
    games=READY/"Anonymous_Games_Approved_Test_Cases.xlsx"; combined=Workbook(); combined.remove(combined.active)
    for source in (games,ham,acc):
        sw=load_workbook(source,data_only=True); src=sw.active; dst=combined.create_sheet(src.title)
        for row in src.iter_rows(values_only=True): dst.append(list(row))
        style(dst); sw.close()
    out=READY/"Anonymous_Games_Hamburger_Account_Settings_Approved_Test_Cases.xlsx"; combined.save(out)
    print(ham); print(acc); print(out)

if __name__=="__main__": main()

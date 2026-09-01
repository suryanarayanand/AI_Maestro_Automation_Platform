from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT = Path("Uploads/Ready/Anonymous_Videos_Quick_Section_Approved_Test_Cases.xlsx")
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "runtime_assertion",
]
PRECONDITION = (
    "The Hindu app is installed; network is available; the user is signed out; "
    "video entitlements and advertisements may vary between runs."
)

cases = [
    ("ANON_VIDEO_001", "Videos quick section launch and selected tab", "Critical", [
        ("Launch the app in a fresh anonymous state and skip the welcome popup when it appears.", "Anonymous Home loads and the welcome popup is dismissed.", "assertVisible screen_home"),
        ("Tap Videos in the Home quick-section navigation.", "The Videos quick section opens.", "tapOn Videos; assertVisible Videos"),
        ("Verify Videos is highlighted and capture the top of the page.", "Videos is visibly selected and screenshot evidence is saved.", "assert selected Videos tab; screenshot"),
    ]),
    ("ANON_VIDEO_002", "Videos quick section refresh", "High", [
        ("Open the Videos quick section.", "The live Videos feed is visible.", "assertVisible Videos"),
        ("Pull down once from the content area to refresh.", "Refresh completes without leaving Videos.", "swipe DOWN; wait for animation"),
        ("Verify the Videos label and video cards remain visible after refresh.", "The refreshed Videos feed is usable.", "assertVisible Videos and video content"),
    ]),
    ("ANON_VIDEO_003", "Video hero and listing presentation", "High", [
        ("Open the Videos quick section and inspect the hero and first visible video cards.", "The hero/card thumbnails, play indicators and headlines render without overlap.", "assertVisible video thumbnail/play indicator and headline"),
        ("Verify visible cards show available metadata such as section label, date/time or author.", "Available video metadata is readable and aligned.", "assertVisible VIDEOS/Video and metadata when supplied"),
        ("Capture the hero and listing presentation.", "Visual evidence is saved.", "screenshot"),
    ]),
    ("ANON_VIDEO_004", "Videos scrolling and advertisement handling", "High", [
        ("Scroll through at least three Videos page viewports with a wait between each gesture.", "The feed continues loading without frozen or overlapping cards.", "repeat swipe UP 3 with waits"),
        ("When a normal inline or sticky advertisement is visible, capture it without treating it as a failure.", "A rendered advertisement is recorded as expected content.", "conditional advertisement screenshot"),
        ("When a blank, failed or non-sticky overlapping advertisement is visible, capture it for review.", "The abnormal advertisement is classified NEEDS_REVIEW for a product bug; it is not an automation failure.", "conditional blank-ad screenshot and NEEDS_REVIEW"),
    ]),
    ("ANON_VIDEO_005", "Video article header and reading time", "Critical", [
        ("Open a current video card using its visible card container rather than a fixed headline.", "The selected video article opens.", "tap dynamic video card; assert article screen"),
        ("Verify the top article label identifies the content as VIDEO or VIDEOS.", "The video section/type label is visible at the top of the article.", "assertVisible regex VIDEO|VIDEOS"),
        ("Verify a reading-time value such as 1 min read or 2 min read is displayed when supplied by the article.", "A valid numeric minute-read label is visible; if the product omits it, capture NEEDS_REVIEW.", "assertVisible regex [0-9]+ min(s)? read or screenshot NEEDS_REVIEW"),
    ]),
    ("ANON_VIDEO_006", "Anonymous video paywall overlay and playback restriction", "Blocker", [
        ("Open video articles until a paywall-blocked video is found, using a bounded maximum of three current cards.", "A paywall-blocked video is identified without relying on a fixed title.", "bounded dynamic selection; detect paywall"),
        ("Verify the paywall popup or overlay is displayed over the video/content area.", "The paywall is visibly associated with and blocks the video.", "assertVisible paywall/Keep reading/offer/Subscribe"),
        ("Attempt to access the blocked video without purchasing.", "Playback remains restricted for the anonymous user.", "assert paywall remains visible; screenshot"),
    ]),
    ("ANON_VIDEO_007", "Video paywall Subscribe action", "Blocker", [
        ("From a paywall-blocked video, verify the Subscribe call-to-action is visible.", "Subscribe is available on the paywall.", "assertVisible Subscribe|SUBSCRIBE"),
        ("Tap Subscribe.", "The subscription offer/plan page opens.", "tapOn Subscribe"),
        ("Verify monthly/yearly plan or offer information and return to the video.", "A valid plan page is shown and Back returns safely.", "assertVisible Monthly|Yearly|Choose a plan|offer; back"),
    ]),
    ("ANON_VIDEO_008", "Video paywall subscriber-login link", "Critical", [
        ("From a paywall-blocked video, scroll within the paywall when necessary until Already a subscriber? Login is visible.", "The existing-subscriber login link becomes visible.", "scroll if needed; assertVisible Already a subscriber/Login"),
        ("Tap the Login link.", "Login to your account opens.", "tapOn Login"),
        ("Verify the login page and return to the video article.", "The login destination is correct and Back is safe.", "assertVisible Login to your account; back"),
    ]),
    ("ANON_VIDEO_009", "Accessible video playback controls", "Critical", [
        ("Open a current video that is not paywall-blocked; if none is available after three cards, record the case as SKIPPED with evidence.", "An accessible video is found or the dynamic unavailability is documented without an automation failure.", "bounded selection; playable or SKIPPED screenshot"),
        ("Tap Play, wait, and verify playback changes to Pause or the progress time advances.", "The accessible video starts playing.", "tapOn Play; assert Pause/progress"),
        ("Exercise pause, seek/forward and fullscreen, then exit fullscreen.", "Supported playback controls respond and the article remains open.", "tap playback controls; assert article after exit"),
    ]),
    ("ANON_VIDEO_010", "Video Reading Options functionality", "High", [
        ("Open a video article and tap Reading Options or the Text Size icon in the floating bar.", "The reading-options panel opens.", "tapOn Reading Options|Text size"),
        ("Select an available text-size option and verify the control responds.", "The selected reading size/state changes without closing the article.", "tap available size; assert panel/article"),
        ("Close the reading-options panel.", "The panel closes and the video article remains visible.", "tapOn Close; assert article"),
    ]),
    ("ANON_VIDEO_011", "Video Bookmark anonymous login gate", "Critical", [
        ("Open a video article and tap Bookmark in the floating bar.", "Anonymous bookmarking invokes authentication.", "tapOn Bookmark"),
        ("Verify Login to your account or the sign-in prompt is displayed and capture it.", "The anonymous login restriction is visible.", "assertVisible Login to your account|Sign in; screenshot"),
        ("Return to the video article.", "Back returns to the same article context.", "back; assert article"),
    ]),
    ("ANON_VIDEO_012", "Video Share functionality", "High", [
        ("Open a video article and tap Share in the floating bar.", "The system share sheet opens.", "tapOn Share; assert share sheet"),
        ("Capture the share sheet without selecting an external application.", "Share destinations are visible and evidence is saved.", "screenshot"),
        ("Close the share sheet using Back.", "The same video article is restored.", "back; assert article"),
    ]),
    ("ANON_VIDEO_013", "Video Comment anonymous login gate", "Critical", [
        ("Open a video article and tap Comment or Post a comment when the action is available.", "The comment action opens its sheet or authentication requirement.", "tapOn Comment|Post a comment"),
        ("Verify the comment surface contains Sign in/Login for the anonymous user and capture it.", "Anonymous commenting is correctly protected.", "assertVisible Sign in|Login; screenshot"),
        ("Close the comment surface and return to the article.", "The video article remains usable.", "close sheet/back; assert article"),
    ]),
    ("ANON_VIDEO_014", "Video floating Subscribe navigation", "Critical", [
        ("Open a video article and verify Subscribe is present in the bottom floating bar.", "The floating Subscribe action is visible.", "assertVisible Subscribe|SUBSCRIBE"),
        ("Tap the floating Subscribe action.", "The subscription offer/plan page opens.", "tapOn Subscribe"),
        ("Verify plan information and return to the article.", "Subscription navigation works and Back restores the article.", "assertVisible Monthly|Yearly|Choose a plan|offer; back"),
    ]),
    ("ANON_VIDEO_015", "Video interstitial advertisement recovery", "High", [
        ("While navigating or opening video content, detect an interstitial advertisement when it appears.", "The flow continues normally when no interstitial appears; a displayed interstitial enters the ad branch.", "conditional interstitial branch"),
        ("If an interstitial or video ad appears, capture it and wait at least 5 seconds plus until its Close/Skip control becomes available.", "Evidence is saved and the automation does not prematurely close the app.", "screenshot; wait 5000; extendedWaitUntil Close|Skip"),
        ("Close only the advertisement and verify the Videos page or video article is restored.", "The ad closes and application content remains open.", "tap ad Close/Skip; assert Videos/article"),
    ]),
    ("ANON_VIDEO_016", "Video article back navigation and retained position", "Medium", [
        ("Open a video card from a scrolled position in the Videos feed.", "The video article opens from the selected list position.", "scroll; tap dynamic card; assert article"),
        ("Use the article Back control.", "The Videos quick section is restored.", "back; assertVisible Videos"),
        ("Verify the Videos tab remains selected and the feed returns near its previous position.", "Section selection and reasonable list context are retained.", "assert selected Videos; screenshot"),
    ]),
]

workbook = Workbook()
sheet = workbook.active
sheet.title = "Anonymous Videos Quick"
sheet.append(HEADERS)
for case_id, name, priority, steps in cases:
    for number, (step, expected, assertion) in enumerate(steps, 1):
        sheet.append([
            case_id, name, "Videos Quick Access", "ANONYMOUS", PRECONDITION, number,
            step, expected, priority, "Yes", assertion,
        ])

header_fill = PatternFill("solid", fgColor="9C0006")
for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [20, 48, 18, 16, 74, 14, 88, 88, 12, 14, 65]
for index, width in enumerate(widths, 1):
    sheet.column_dimensions[chr(64 + index)].width = width
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions
sheet.row_dimensions[1].height = 28
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
workbook.save(OUTPUT)
print(f"Created {OUTPUT} with {len(cases)} cases and {sheet.max_row - 1} executable steps")

from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
SRC=ROOT/'Uploads'/'Ready'/'Anonymous_Hamburger_Approved_Test_Cases.xlsx'
OUT=ROOT/'Uploads'/'Ready'/'Subscriber_Hamburger_Approved_Test_Cases.xlsx'
wb=load_workbook(SRC);ws=wb.active;ws.title='Subscriber Hamburger'
headers={c.value:i for i,c in enumerate(ws[1],1)}
for row in ws.iter_rows(min_row=2):
 cid=row[headers['test_case_id']-1]; cid.value=str(cid.value).replace('ANON_HAM_','SUB_HAM_')
 row[headers['user_state']-1].value='SUBSCRIBER'
 row[headers['precondition']-1].value='The app, network and Maestro device are available; valid subscriber credentials are provided by protected portal configuration. Login before evidence capture; wait before screenshots.'
 row[headers['existing_yaml_reference']-1].value='CPLX_HAM_*.yaml; HamburgerOptions/HAM_OPT_*.yaml; Common/OPEN_SUBSCRIBER_HOME.yaml; anonymous Hamburger YAML'
 step=str(row[headers['step']-1].value); expected=str(row[headers['expected_result']-1].value)
 if cid.value=='SUB_HAM_008':
  step=step.replace('anonymous paywall','subscriber entitlement').replace('paywall or article-end content','article body and article-end content')
  expected=expected.replace('A paywall shows its offer and Login, or unrestricted article-end content is documented.','The subscriber reads the article without Subscribe, Login, or paywall obstruction.')
 if cid.value=='SUB_HAM_009':
  step=step.replace('observing advertisements','checking the subscriber no-advertisement contract').replace('If an interstitial appears, wait for Close and close only the ad.','Verify inline, sticky and interstitial advertisements are absent.')
  expected=expected.replace('Sticky ads do not fail the flow; problematic or blank ads are captured.','No advertisement surface appears for the subscriber.').replace('The article or section is restored without closing the app.','Subscriber navigation remains available without ad dismissal.')
 row[headers['step']-1].value=step;row[headers['expected_result']-1].value=expected
wb.save(OUT);print(OUT)

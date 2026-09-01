from pathlib import Path

from openpyxl import load_workbook

FILES = (
    Path("Uploads/Ready/Anonymous_Premium_Approved_Test_Cases.xlsx"),
    Path("Uploads/Ready/Anonymous_EBooks_Approved_Test_Cases.xlsx"),
)

for path in FILES:
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"\n{path}")
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in rows[0]]
        data = rows[1:]
        index = {header: position for position, header in enumerate(headers)}
        ids = [str(row[index["test_case_id"]] or "").strip() for row in data]
        unique = list(dict.fromkeys(ids))
        duplicates = sorted({case_id for case_id in unique if ids.count(case_id) > 1})
        case_names = {}
        for row in data:
            case_names.setdefault(str(row[index["test_case_id"]]), set()).add(str(row[index["name"]]))
        print("sheet", ws.title, "rows", len(data), "cases", len(unique))
        print("id_range", unique[0], unique[-1])
        print("modules", sorted({str(row[index["module"]]) for row in data}))
        print("states", sorted({str(row[index["user_state"]]) for row in data}))
        print("blank_steps", sum(not row[index["step"]] for row in data))
        print("blank_expected", sum(not row[index["expected_result"]] for row in data))
        print("conflicting_names", sum(len(names) > 1 for names in case_names.values()))
        print("repeated_ids_are_step_rows", len(duplicates))
        for optional in ("runtime_assertion", "yaml_reference", "master_sheet_reference", "existing_yaml_reference"):
            if optional in index:
                print(optional + "_blank", sum(not row[index[optional]] for row in data))
        full_text = "\n".join(" | ".join(str(value or "") for value in row) for row in data).casefold()
        print("mentions_free_ebook", "free ebook" in full_text or "free e-book" in full_text)
        print("mentions_non_all_stories_listing_ads", any(
            section in full_text and phrase in full_text
            for section, phrase in (("specials", "specials advertisement"), ("packages", "packages advertisement"), ("webinar", "webinar advertisement"))
        ))

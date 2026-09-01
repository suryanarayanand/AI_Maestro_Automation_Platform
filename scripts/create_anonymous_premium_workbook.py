from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Anonymous_Premium_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
]


CASES = [
    ("ANON_PREM_001", "Anonymous Premium launch and selected navigation", [
        ("Launch the app in a fresh anonymous state, dismiss the welcome popup when shown, and tap Premium.", "The Premium root screen opens."),
        ("Verify the Premium bottom-navigation icon is selected and capture the page.", "Premium is visibly selected and evidence is saved."),
        ("Verify Briefing, Specials, Packages, Webinar, and All Stories tabs.", "All five Premium sections are available."),
    ]),
    ("ANON_PREM_002", "Premium pull to refresh", [
        ("Open Premium as an anonymous user and pull down from the content area.", "Refresh completes without leaving Premium."),
        ("Verify the Premium screen, selected Premium navigation, five tabs, and live content after refresh.", "Premium remains selected and usable after refresh."),
    ]),
    ("ANON_PREM_003", "Briefing locked-state presentation", [
        ("Open Premium and select Briefing.", "The Briefing section is selected."),
        ("Verify Briefing is locked for the anonymous user.", "The locked Briefing state prevents access to subscriber content."),
        ("Verify Unlock Briefing and other subscription benefits, Subscribe, and Already a subscriber Login wording, then capture it.", "All required lock-state subscription and login actions are visible."),
    ]),
    ("ANON_PREM_004", "Briefing Subscribe plan destination", [
        ("Open the locked Briefing section and tap Subscribe.", "The subscription-offer page opens."),
        ("Verify Yearly, Monthly, and Already a subscriber Login.", "Both plan types and subscriber-login route are visible."),
        ("Capture the plan page and return to Briefing.", "The locked Briefing screen is restored."),
    ]),
    ("ANON_PREM_005", "Briefing subscriber-login destination", [
        ("Open the locked Briefing section and tap Already a subscriber Login.", "The account login page opens."),
        ("Verify Login to your account and capture it.", "The correct login destination is visible."),
        ("Press Back without signing in.", "The anonymous locked Briefing state is restored."),
    ]),
    ("ANON_PREM_006", "Briefing advertisement absence", [
        ("Open Briefing and inspect the complete accessible locked-state page.", "The Briefing lock state remains visible."),
        ("Verify inline Advertisement, sticky-ad container, and Taboola are absent.", "Briefing does not display advertisements."),
        ("Capture the final Briefing state.", "Advertisement-absence evidence is saved."),
    ]),
    ("ANON_PREM_007", "Specials full-page content coverage", [
        ("Open Premium and select Specials, dismissing any informational OKAY prompt.", "The Specials page loads."),
        ("Scroll Specials from top to bottom and capture top, middle, and lower-page screenshots.", "Specials content remains usable throughout the page."),
        ("Verify no inline Advertisement, sticky-ad container, or Taboola appears on the Specials listing page.", "The Specials listing is advertisement-free."),
    ]),
    ("ANON_PREM_008", "Specials promotional popup handling", [
        ("Open Specials and activate a See More or available event entry.", "The selected Specials destination opens."),
        ("If the Dear Reader event-information popup appears, capture and close it safely.", "The informational popup is documented and dismissed without closing the app."),
        ("Verify the Specials destination remains usable.", "The user stays in the Premium Specials journey."),
    ]),
    ("ANON_PREM_009", "Specials article anonymous paywall", [
        ("Open a representative article from Specials.", "The article-detail screen opens."),
        ("Handle any interstitial advertisement by capturing it, waiting in five-second intervals, and closing only the ad.", "The article is restored after interstitial handling."),
        ("Scroll the article and verify anonymous reading is blocked by the paywall.", "The complete Premium article cannot be read anonymously."),
        ("Capture the paywall, subscription offer, and subscriber-login action.", "Paywall evidence is saved."),
    ]),
    ("ANON_PREM_010", "Packages full-page content coverage", [
        ("Open Premium and select Packages, dismissing any informational OKAY prompt.", "The Packages page loads."),
        ("Scroll Packages from top to bottom and capture top, middle, and lower-page screenshots.", "Packages content remains usable throughout the page."),
        ("Verify no inline Advertisement, sticky-ad container, or Taboola appears on the Packages listing page.", "The Packages listing is advertisement-free."),
    ]),
    ("ANON_PREM_011", "Packages article anonymous paywall", [
        ("Open a representative article from Packages.", "The article-detail screen opens."),
        ("Handle any interstitial advertisement using delayed-close behavior.", "Only the ad closes and the article is restored."),
        ("Scroll until the Premium paywall blocks further reading and capture it.", "Anonymous full-article access is restricted."),
    ]),
    ("ANON_PREM_012", "Webinar full-page content coverage", [
        ("Open Premium and select Webinar, dismissing any informational OKAY prompt.", "The Webinar page loads."),
        ("Scroll Webinar from top to bottom and capture top, middle, and lower-page screenshots.", "Webinar content remains usable throughout the page."),
        ("Verify no inline Advertisement, sticky-ad container, or Taboola appears on the Webinar listing page.", "The Webinar listing is advertisement-free."),
    ]),
    ("ANON_PREM_013", "Webinar entry and anonymous restriction", [
        ("Open a representative Webinar entry.", "The selected Webinar destination opens."),
        ("Handle an event-information popup or interstitial advertisement when present.", "The interruption is captured and safely dismissed."),
        ("Verify the applicable anonymous subscription or login restriction and capture it.", "Restricted Webinar access provides a subscription or authentication route."),
    ]),
    ("ANON_PREM_014", "All Stories full-page content coverage", [
        ("Open Premium and select All Stories.", "The All Stories feed loads."),
        ("Scroll All Stories from top to bottom and capture representative screenshots.", "All Stories remains usable throughout the feed."),
        ("Check and capture advertisements encountered during scrolling.", "Advertisement presence is documented because All Stories is the Premium section expected to contain ads."),
    ]),
    ("ANON_PREM_015", "All Stories rendered advertisement landing page", [
        ("Scroll All Stories until a visibly rendered and clickable advertisement creative appears.", "A usable creative is found, or the bounded search records no rendered ad."),
        ("Capture and tap only the visible advertisement creative.", "The external advertisement destination opens."),
        ("Verify the destination is not Webpage not available, capture it, and press Back.", "The destination is documented and All Stories is restored."),
    ]),
    ("ANON_PREM_016", "All Stories blank and sticky advertisement classification", [
        ("Scroll All Stories while checking inline and sticky advertisement areas.", "Advertisement states are observed."),
        ("Capture blank Advertisement containers without blind-tapping them.", "Blank-ad defect evidence is saved."),
        ("Capture non-overlapping sticky ads as acceptable and intrusive overlapping ads for defect review.", "Advertisement behavior is classified correctly."),
    ]),
    ("ANON_PREM_017", "All Stories article anonymous paywall", [
        ("Open a representative All Stories article.", "The article-detail screen opens."),
        ("Handle any interstitial advertisement using delayed-close behavior.", "The article is restored without closing the app."),
        ("Scroll until the paywall blocks full reading and capture the offer and login route.", "Anonymous users cannot read the complete Premium article."),
    ]),
    ("ANON_PREM_018", "Premium article paywall offer details", [
        ("Open and scroll a Premium article until its reader paywall appears.", "The paywall blocks further article reading."),
        ("Verify Keep reading free for 3 days or an available discount, Subscribe, and Already a subscriber Login.", "The paywall exposes an offer, subscription action, and subscriber-login route."),
        ("Capture the complete paywall state.", "Paywall offer evidence is saved."),
    ]),
    ("ANON_PREM_019", "Premium article paywall Login link", [
        ("Open a paywalled Premium article and tap the Already a subscriber Login link.", "Login to your account opens."),
        ("Capture the login page and press Back.", "The paywalled article is restored without authentication."),
    ]),
    ("ANON_PREM_020", "Premium article controls", [
        ("Open a representative Premium article.", "The article-detail screen is visible."),
        ("Verify available Text size, Bookmark, Share, Comment, Listen to article, and AI Summary actions.", "Applicable article controls are visible; optional audio and summary controls are recorded conditionally."),
        ("Capture the article header and controls.", "Control evidence is saved."),
    ]),
    ("ANON_PREM_021", "Premium article reading options", [
        ("Open a Premium article and tap Text size.", "Reading Options opens."),
        ("Verify A-minus and A-plus controls, exercise both, and capture the panel.", "Text-size controls respond."),
        ("Close Reading Options.", "The same article is restored."),
    ]),
    ("ANON_PREM_022", "Premium article Bookmark login gate", [
        ("Open a Premium article and tap Bookmark.", "Login to your account opens for the anonymous user."),
        ("Capture the login gate and press Back.", "The same article is restored without bookmarking."),
    ]),
    ("ANON_PREM_023", "Premium article Share sheet", [
        ("Open a Premium article and tap Share.", "The system share sheet opens."),
        ("Verify a share target, capture the sheet, and press Back without sharing.", "The same article is restored."),
    ]),
    ("ANON_PREM_024", "Premium article Comment restriction", [
        ("Open and scroll a Premium article while checking for the paywall and Post a comment.", "The article reaches its anonymous restriction state."),
        ("When paywall blocks the article first, capture it and mark Comment as not applicable for that article.", "A blocked Premium article is not incorrectly required to show comments."),
        ("If Post a comment is available, open it and verify the anonymous sign-in requirement.", "Commenting requires authentication."),
    ]),
    ("ANON_PREM_025", "Premium article Listen and AI Summary", [
        ("Open a Premium article and inspect for Listen to article and AI Summary.", "Optional actions are identified."),
        ("When Listen is available, tap it, capture the resulting state, and return.", "Audio behavior is documented."),
        ("When AI Summary is available, tap it and capture its anonymous subscription gate or content state, then return.", "AI Summary behavior is documented without failing an ineligible article."),
    ]),
    ("ANON_PREM_026", "Premium interstitial delayed-close recovery", [
        ("Open Premium articles through a bounded search until an interstitial appears.", "An interstitial is found or absence is recorded."),
        ("Capture it immediately and wait in five-second intervals for Close.", "Delayed video-ad behavior is documented."),
        ("Tap Close when available; if it never appears, capture the blocked state and use Back only to dismiss the ad activity.", "Only the advertisement closes and the Premium article is restored."),
    ]),
    ("ANON_PREM_027", "Premium article Taboola coverage", [
        ("Open a Premium article and scroll through the accessible content and restriction region.", "The article scroll remains stable."),
        ("When Image for Taboola Advertising Unit appears, capture it.", "Taboola evidence is saved; absence during the bounded search is a valid conditional result."),
        ("Verify any reader paywall remains correctly positioned and usable.", "Advertising does not bypass or break the Premium restriction."),
    ]),
    ("ANON_PREM_028", "Premium five-tab navigation continuity", [
        ("Open Briefing, Specials, Packages, Webinar, and All Stories sequentially.", "Each tab loads its correct Premium section."),
        ("Capture every selected section and verify the Premium screen remains active.", "Five-tab navigation evidence is saved."),
        ("Return to Briefing and verify its anonymous locked state remains intact.", "Premium navigation does not change the anonymous entitlement."),
    ]),
    ("ANON_PREM_029", "Premium subscription entry-point consistency", [
        ("Open subscription offers from Briefing Subscribe, a Premium article paywall, and another available Premium Subscribe action.", "Each Subscribe action reaches the subscription-offer page."),
        ("At every destination verify Yearly, Monthly, and Already a subscriber Login.", "Subscription options are consistent across entry points."),
        ("Return after each check without purchasing or signing in.", "The anonymous Premium session remains active."),
    ]),
    ("ANON_PREM_030", "Premium article paging and entitlement persistence", [
        ("Open a Premium article and swipe left through five article positions.", "Article paging works to the left."),
        ("Swipe right through five article positions.", "Article paging works to the right."),
        ("Handle every interstitial safely and capture paywalls encountered.", "Anonymous restrictions persist and no ad closes the app."),
    ]),
    ("ANON_PREM_031", "Briefing lock overlay visual integrity", [
        ("Open locked Briefing as an anonymous user.", "The Briefing paywall mask is visible."),
        ("Verify the translucent lock overlay covers protected content while leaving only an allowed teaser visible.", "Protected Briefing text cannot be read through the overlay."),
        ("Capture the overlay, padlock, Subscribe button, and login link without clipping or overlap.", "The complete lock presentation is documented."),
    ]),
    ("ANON_PREM_032", "Briefing protected-content gesture lock", [
        ("Open locked Briefing and attempt vertical scrolling over the protected content.", "The gesture does not expose additional protected Briefing text."),
        ("Attempt horizontal article-card movement or text selection where supported.", "The lock layer continues to prevent content leakage."),
        ("Capture the final locked state.", "Briefing remains securely locked after gesture attempts."),
    ]),
    ("ANON_PREM_033", "Briefing numbered-card entitlement persistence", [
        ("Open locked Briefing and inspect the numbered multi-card navigation.", "The available Briefing card indices are visible."),
        ("Attempt to open multiple numbered Briefing cards within a bounded sample.", "The paywall persists on every sampled card and no protected body text is exposed."),
        ("Capture each sampled locked-card state.", "Entitlement enforcement evidence is saved."),
    ]),
    ("ANON_PREM_034", "Briefing header and overlay Subscribe consistency", [
        ("Open locked Briefing and open the subscription page from the overlay Subscribe action.", "The subscription plan page opens with Yearly and Monthly."),
        ("Return and open the subscription page from the top-right header Subscribe action.", "The same subscription plan destination opens."),
        ("Compare and capture both destinations.", "Both Briefing Subscribe entry points behave consistently."),
    ]),
    ("ANON_PREM_035", "Briefing secure state after connectivity interruption", [
        ("Using a controlled network test, interrupt connectivity before opening or refreshing Briefing.", "Premium entitlement verification cannot complete normally."),
        ("Verify Briefing defaults to the locked state and never exposes protected content.", "The app fails securely when entitlement status is unavailable."),
        ("Restore connectivity and refresh Premium.", "The correct anonymous locked Briefing state remains available."),
    ]),
    ("ANON_PREM_036", "Briefing lock accessibility and dark-mode layout", [
        ("Open locked Briefing using maximum supported device font scaling.", "Paywall text, padlock, Subscribe, and Login reflow without overlap or clipping."),
        ("Enable dark mode and reopen the locked Briefing page.", "The lock overlay remains readable with correct contrast."),
        ("Capture both accessibility states and restore default display settings.", "Visual evidence is saved without leaving device settings modified."),
    ]),
    ("ANON_PREM_037", "Specials horizontal carousel behavior", [
        ("Open Specials and locate an article-card carousel.", "A Specials carousel is visible."),
        ("Swipe the carousel left and right through multiple cards.", "Horizontal scrolling is smooth and cards do not overlap."),
        ("Verify card images and text remain associated and capture the carousel.", "Carousel content renders correctly after movement."),
    ]),
    ("ANON_PREM_038", "Specials See More pagination and return state", [
        ("Open Specials, scroll to a collection, and tap See More.", "The collection page opens with related articles."),
        ("Scroll the See More page through additional or infinitely loaded content.", "Additional collection articles load correctly."),
        ("Press Back and verify the previous Specials section and scroll position.", "The prior Specials context is retained."),
    ]),
    ("ANON_PREM_039", "Packages horizontal carousel behavior", [
        ("Open Packages and locate an article-card carousel.", "A Packages carousel is visible."),
        ("Swipe left and right through multiple cards.", "Carousel movement is smooth with no overlapping cards."),
        ("Verify package images, typography, spacing, and text remain correct.", "Package cards render consistently."),
    ]),
    ("ANON_PREM_040", "Packages See More return state", [
        ("Open Packages, scroll to a package, and tap See More.", "The selected package destination opens."),
        ("Capture the destination and press Back.", "Packages is restored."),
        ("Verify the selected Packages tab and prior scroll position are retained.", "Navigation does not reset the Packages journey."),
    ]),
    ("ANON_PREM_041", "Webinar card metadata and imagery", [
        ("Open Webinar and inspect the featured banner plus representative cards.", "Webinar content is visible."),
        ("Verify webinar title, category label, publication date, event date/time, banner, and thumbnails where applicable.", "Available Webinar metadata and images render correctly."),
        ("Capture the featured and standard card layouts.", "Webinar presentation evidence is saved."),
    ]),
    ("ANON_PREM_042", "Webinar banner and standard-card routing", [
        ("Tap the featured Webinar banner.", "The Webinar detail or applicable anonymous paywall opens."),
        ("Return and tap a non-banner Webinar card.", "Its corresponding article or video destination opens rather than the wrong featured page."),
        ("Capture both destinations and return to Webinar.", "Banner and standard cards route to their intended content."),
    ]),
    ("ANON_PREM_043", "Webinar video controls and lifecycle", [
        ("Open a Webinar article that contains video and start playback when anonymous access permits.", "The video player opens and begins playback, or the expected Premium restriction is shown."),
        ("When playback is available, verify player controls and background then foreground the app.", "Controls remain usable and playback resumes according to product behavior."),
        ("Capture the player or applicable paywall state.", "The observed Webinar video outcome is documented."),
    ]),
    ("ANON_PREM_044", "Webinar detail return-state retention", [
        ("Scroll the Webinar listing to a non-top position and open an entry.", "The Webinar destination opens."),
        ("Press Back after handling any popup, paywall, or interstitial.", "The Webinar listing is restored."),
        ("Verify Webinar remains selected and the prior scroll position is retained.", "The user returns to the same Webinar context."),
    ]),
    ("ANON_PREM_045", "All Stories hero and article-card rendering", [
        ("Open All Stories and inspect the hero story and representative article cards.", "The hero and article feed are visible."),
        ("Verify hero image, headline, summary, timestamps, thumbnails, and associated story text.", "Images and metadata correspond to their story cards without distortion."),
        ("Capture the hero and representative cards.", "All Stories rendering evidence is saved."),
    ]),
    ("ANON_PREM_046", "All Stories refresh and newly loaded content", [
        ("Open All Stories and record the visible top stories and advertisement positions.", "A baseline feed state is captured."),
        ("Pull to refresh and wait for content loading to complete.", "The feed refreshes successfully."),
        ("Verify current stories, relative timestamps, and configured advertisement placement after refresh.", "Refreshed content remains correctly ordered and ads remain in valid positions."),
    ]),
    ("ANON_PREM_047", "All Stories infinite scroll and lazy loading", [
        ("Open All Stories and scroll through a large bounded number of story cards.", "Additional stories load as the user approaches the feed end."),
        ("Verify newly visible images lazy-load without leaving persistent blank or mismatched thumbnails.", "Images load for their corresponding stories."),
        ("Verify scrolling remains smooth and previously rendered cards do not move unexpectedly.", "Infinite scrolling is stable."),
    ]),
    ("ANON_PREM_048", "All Stories advertisement cadence stability", [
        ("Open All Stories and record the positions of advertisements among story cards during a bounded scroll.", "Initial advertisement placement is documented."),
        ("Load additional stories and verify existing advertisements do not shift into invalid or overlapping positions.", "Previously rendered advertisement placement remains stable."),
        ("Refresh the feed and capture the configured advertisement cadence for comparison.", "Advertisements remain separated by the expected editorial configuration."),
    ]),
    ("ANON_PREM_049", "All Stories article return-state retention", [
        ("Scroll All Stories to a non-top position and open a representative article.", "The article opens with its applicable paywall."),
        ("Handle any interstitial and press Back.", "All Stories is restored."),
        ("Verify All Stories remains selected and the prior feed position is retained.", "Article navigation does not reset the feed context."),
    ]),
    ("ANON_PREM_050", "Premium controlled loading-error and recovery states", [
        ("Using controlled test data or network interception, trigger a Premium listing load failure.", "The app shows an appropriate error or retry state without crashing or exposing protected content."),
        ("Trigger missing-image data and verify a placeholder is used instead of a broken layout.", "Cards remain usable with placeholder imagery."),
        ("Restore the service or connectivity and retry or refresh.", "Premium content recovers successfully and the selected tab remains correct."),
    ]),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous Premium"
    sheet.append(HEADERS)
    precondition = (
        "The Hindu app is installed; network is available; the user is signed out; "
        "Premium content, advertisements, optional article actions, and interstitials may vary between runs."
    )
    # Cases 31-50 duplicate visual, carousel, return-state, refresh, and loading
    # obligations already covered by the first 30 end-to-end cases.
    for case_id, name, steps in CASES[:30]:
        for number, (step, expected) in enumerate(steps, 1):
            sheet.append([
                case_id, name, "Premium", "ANONYMOUS", precondition,
                number, step, expected, "High", "Yes",
            ])
    header_fill = PatternFill("solid", fgColor="7A3E9D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [20, 50, 18, 16, 76, 14, 90, 90, 12, 14]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "Uploads" / "TH App Testing Scenarios_AutomationCopy.xlsx"
OUTPUT = ROOT / "Uploads" / "Normalized" / "TH_App_Catalog_Extracted_Maestro_Flows.xlsx"
FIRST_BATCH = ROOT / "Uploads" / "TH_App_Catalog_Flow_Batch_001.xlsx"


def classify(text):
    text = text.casefold()
    external = (
        "play store", "appstore", "app store", "test flight", "diawi",
        "push notification", "payment", "purchase", "refund", "whatsapp",
        "twitter", "google discover", "cross platform", "ios",
    )
    if any(term in text for term in external):
        return "manual", "", "", []
    if "create account" in text or "sign up" in text or "signup" in text:
        return "manual", "", "", []
    if "logout" in text:
        return "executable", "subscriber", "Subscriber is logged in", [
            ("Run OPEN_SUBSCRIBER_HOME.yaml", "Subscriber home is visible"),
            ("TAP(nav_account)", "Account menu is visible"),
            ('Scroll down until text "LOGOUT" is visible', "Logout action is visible"),
            ("TAP_TEXT(LOGOUT)", "Anonymous home is visible"),
            ("ASSERT_VISIBLE_TEXT(SUBSCRIBE)", "User is logged out"),
        ]
    if "login" in text or "sign in" in text:
        return "executable", "subscriber", "Valid test credentials are configured", [
            ("Run OPEN_SUBSCRIBER_HOME.yaml", "Subscriber home is visible"),
            ("ASSERT_VISIBLE(screen_home)", "Login completed"),
            ("ASSERT_NOT_VISIBLE_TEXT(SUBSCRIBE)", "Subscriber state is active"),
        ]
    if "hamburger" in text or "menu" in text and "account" not in text:
        return "executable", "anonymous", "App is available", [
            ("Run OPEN_HOME_FOR_LOCATOR_SMOKE.yaml", "Home is visible"),
            ("TAP(nav_menu)", "Hamburger menu opens"),
            ("ASSERT_VISIBLE(screen_hamburger)", "Menu content is visible"),
        ]
    if "account" in text or "user menu" in text or "appearance" in text or "text size" in text:
        return "executable", "anonymous", "Anonymous home is visible", [
            ("Run OPEN_HOME_FOR_LOCATOR_SMOKE.yaml", "Home is visible"),
            ("TAP(nav_account)", "Account menu opens"),
            ("ASSERT_VISIBLE(screen_user_menu)", "Account settings are visible"),
        ]
    if "ebook" in text or "e-book" in text:
        return "executable", "anonymous", "Anonymous home is visible", [
            ("Run OPEN_HOME_FOR_LOCATOR_SMOKE.yaml", "Home is visible"),
            ("TAP(nav_ebooks)", "E-books page opens"),
            ("ASSERT_VISIBLE(screen_ebooks)", "E-books content is visible"),
        ]
    if "game" in text or "sudoku" in text or "crossword" in text:
        return "executable", "subscriber", "Subscriber account is available", [
            ("Run OPEN_SUBSCRIBER_GAMES.yaml", "Games page opens"),
            ("ASSERT_VISIBLE(screen_games)", "Games content is visible"),
        ]
    if "premium" in text:
        return "executable", "subscriber", "Subscriber account is available", [
            ("Run OPEN_SUBSCRIBER_HOME.yaml", "Subscriber home is visible"),
            ("TAP(nav_premium)", "Premium page opens"),
            ("ASSERT_VISIBLE(screen_premium)", "Premium content is visible"),
        ]
    if "trending" in text:
        return "executable", "anonymous", "Anonymous home is visible", [
            ("Run OPEN_HOME_FOR_LOCATOR_SMOKE.yaml", "Home is visible"),
            ("TAP(nav_trending)", "Trending page opens"),
            ("ASSERT_VISIBLE(screen_trending)", "Trending content is visible"),
        ]
    if "article" in text:
        return "partial", "anonymous", "Anonymous home contains an article card", [
            ("Run OPEN_HOME_FOR_LOCATOR_SMOKE.yaml", "Home is visible"),
            ("TAP(article_card)", "An article opens"),
            ("ASSERT_VISIBLE(screen_article_detail)", "Article detail is visible"),
        ]
    if "home" in text or "launch" in text or "onboarding" in text:
        return "executable", "anonymous", "App is installed", [
            ("Run OPEN_HOME_FOR_LOCATOR_SMOKE.yaml", "Home is visible"),
            ("ASSERT_VISIBLE(screen_home)", "Home content is loaded"),
        ]
    return "manual", "", "", []


source = load_workbook(SOURCE, read_only=True, data_only=True)["TH App"]
workbook = Workbook()
flows = workbook.active
flows.title = "Extracted Flows"
headers = [
    "test_case_id", "name", "step", "expected_result", "automation_intent",
    "module", "precondition", "user_state", "priority", "automatable",
    "step_number", "source_sheet", "source_row", "coverage",
]
flows.append(headers)
manual = workbook.create_sheet("Manual Review")
manual.append(["test_case_id", "module", "name", "description", "reason", "source_row"])

counts = {"executable": 0, "partial": 0, "manual": 0}
for row_number, values in enumerate(source.iter_rows(min_row=2, values_only=True), start=2):
    serial, module, scenario, description = values[:4]
    if serial is None and not any((module, scenario, description)):
        continue
    case_id = f"THCAT_{int(serial):04d}" if str(serial).replace(".0", "").isdigit() else f"THCAT_R{row_number}"
    name = str(scenario or module or f"Scenario {serial}").strip()
    module = str(module or "General").strip()
    description = str(description or "").strip()
    coverage, user_state, precondition, steps = classify(f"{module} {name} {description}")
    counts[coverage] += 1
    if not steps:
        manual.append([case_id, module, name, description,
                       "Requires external integration or a feature-specific validated flow", row_number])
        continue
    for number, (step, expected) in enumerate(steps, start=1):
        flows.append([
            case_id, name, step, "", step, module, precondition, user_state,
            "P2", "Yes" if coverage == "executable" else "Partial", number,
            "TH App", row_number, coverage,
        ])

for sheet in (flows, manual):
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="316FEA")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in ("B", "C", "D", "E"):
        sheet.column_dimensions[column].width = 55

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
workbook.save(OUTPUT)

# Create a portal-safe first batch containing ten complete cases.
source_flows = load_workbook(OUTPUT, read_only=True, data_only=True)["Extracted Flows"]
batch_workbook = Workbook()
batch_sheet = batch_workbook.active
batch_sheet.title = "Test Cases"
batch_case_ids = []
for row_number, row in enumerate(source_flows.iter_rows(values_only=True), start=1):
    if row_number == 1:
        batch_sheet.append(row)
        continue
    case_id = row[0]
    if case_id not in batch_case_ids and len(batch_case_ids) < 10:
        batch_case_ids.append(case_id)
    if case_id in batch_case_ids:
        batch_sheet.append(row)
for cell in batch_sheet[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="316FEA")
batch_sheet.freeze_panes = "A2"
batch_sheet.auto_filter.ref = batch_sheet.dimensions
batch_workbook.save(FIRST_BATCH)
print(OUTPUT)
print(FIRST_BATCH)
print(counts)

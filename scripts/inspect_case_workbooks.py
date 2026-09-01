from pathlib import Path
from openpyxl import load_workbook

for path in (
    Path("Uploads/Ready/Anonymous_Home_Approved_Test_Cases.xlsx"),
    Path("Uploads/Ready/Anonymous_Trending_Approved_Test_Cases.xlsx"),
):
    print(f"\n## {path}")
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        print(ws.title, headers)
        for values in ws.iter_rows(min_row=2, values_only=True):
            text = " | ".join("" if value is None else str(value) for value in values)
            if "ANON_HOME_" in text or "ANON_TREND_" in text:
                print(text)

"""Create the Subscriber Premium module import workbook."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = Path("Uploads/Ready/Subscriber_Premium_Approved_Test_Cases.xlsx")
PRE = ("The Hindu app is installed; network is available; configured masked credentials "
       "belong to an active subscriber with Premium entitlement.")

CASES = [
 ("SUB_PREM_001","Subscriber Premium launch and selected navigation","Critical",[
  ("Authenticate with the configured subscriber account and tap Premium.","Premium opens and its bottom-navigation item is selected.","OPEN_SUBSCRIBER_HOME; tap nav_premium; assert screen_premium and nav_premium"),
  ("Wait for content loading and capture the Premium landing page.","Readable landing-page evidence is saved.","waitForAnimationToEnd immediately before screenshot"),
  ("Verify Subscribe, Login offers, and advertisements are absent.","The entitled Premium page has no monetisation or authentication prompt.","ASSERT_SUBSCRIBER_NO_MONETIZATION; assertNotVisible Login to your account")]),
 ("SUB_PREM_002","Subscriber Premium pull to refresh","Critical",[
  ("Open Premium and pull down to refresh.","Refresh completes without leaving Premium.","swipe DOWN; extendedWaitUntil screen_premium"),
  ("Verify Premium sections reload and capture the settled page.","Current Premium content remains usable.","assert section tabs; wait; screenshot"),
  ("Recheck subscriber entitlement after refresh.","No Subscribe, login offer, paywall, or advertisement appears.","assert no monetisation/paywall/login")]),
 ("SUB_PREM_003","Subscriber Briefing unlocked entitlement","Blocker",[
  ("Open the Briefing section as the active subscriber.","Briefing content opens instead of the anonymous locked card.","tap Briefing; assert screen/content"),
  ("Verify Unlock Briefing, Subscribe, and Already a subscriber Login are absent.","Briefing is not presented as locked.","assertNotVisible Unlock Briefing/Subscribe/Already a subscriber/Login"),
  ("Scroll Briefing and capture at least two settled reading states.","Entitled Briefing content is readable and documented.","swipe UP; wait before screenshots")]),
 ("SUB_PREM_004","Subscriber Briefing complete reading","Blocker",[
  ("Open the latest Briefing item.","The selected Briefing detail opens.","tap current briefing card; assert detail"),
  ("Scroll through the available Briefing body.","Content remains readable without a subscription restriction.","repeat bounded swipe UP; assert no paywall"),
  ("Verify advertisements are absent and capture the lower reading state.","Briefing reading is ad-free and evidence is saved.","ASSERT_SUBSCRIBER_NO_MONETIZATION; wait; screenshot")]),
 ("SUB_PREM_005","Premium five-tab navigation","Critical",[
  ("Open Briefing, Specials, Packages, Webinar, and All Stories one by one.","Every Premium section opens its corresponding page.","tap each tab; assert screen_premium and selected label"),
  ("Wait and capture each selected section.","Each section is documented only after loading settles.","waitForAnimationToEnd before every screenshot"),
  ("Verify the subscriber remains authenticated and ad-free in every tab.","No login, Subscribe, paywall, or advertisement appears.","assert no monetisation/login/paywall per tab")]),
 ("SUB_PREM_006","Specials full-page subscriber coverage","High",[
  ("Open Specials and scroll from top to lower content.","Specials cards continue loading and remain usable.","tap Specials; repeat swipe UP"),
  ("Capture top, middle, and lower positions after loading.","Full-page Specials evidence is available.","wait before 3 screenshots"),
  ("Verify no advertisement or subscription prompt appears.","Specials respects subscriber entitlement.","ASSERT_SUBSCRIBER_NO_MONETIZATION; assertNotVisible Login/paywall")]),
 ("SUB_PREM_007","Specials article full access","Blocker",[
  ("Open a current article from Specials.","Article detail opens.","tap article card; assert screen_article_detail"),
  ("Scroll through the article body in bounded steps.","The complete article is readable without a paywall.","repeat swipe UP; assertNotVisible Keep reading/offer"),
  ("Capture settled article states and verify no ads.","Full, ad-free subscriber access is evidenced.","ASSERT_SUBSCRIBER_NO_MONETIZATION; wait before screenshots")]),
 ("SUB_PREM_008","Packages full-page subscriber coverage","High",[
  ("Open Packages and scroll through its available cards.","Packages content loads through the page.","tap Packages; repeat swipe UP"),
  ("Capture the top and lower Packages states after waiting.","Readable Packages evidence is saved.","wait before screenshots"),
  ("Verify no Subscribe, Login, paywall, or advertisement appears.","Packages remains entitled and ad-free.","assert no monetisation/login/paywall")]),
 ("SUB_PREM_009","Packages article full access","Blocker",[
  ("Open a current Packages article.","Article detail opens from Packages.","tap article card; assert screen_article_detail"),
  ("Scroll through its body and available post-article content.","The subscriber can read beyond anonymous restriction points.","repeat swipe UP; assert article"),
  ("Verify paywall and ads are absent; capture settled evidence.","Packages article entitlement is proven.","assert no paywall/ads; wait; screenshot")]),
 ("SUB_PREM_010","Webinar full-page subscriber coverage","Critical",[
  ("Open Webinar and scroll through current entries.","Webinar cards and metadata load correctly.","tap Webinar; repeat swipe UP"),
  ("Open a current Webinar entry.","Its corresponding article or video detail opens.","tap webinar card; assert article/player"),
  ("Verify entitled access without paywall, Subscribe, Login, or ads; capture evidence.","Webinar content is accessible to the subscriber.","assert no restrictions/ads; wait; screenshot")]),
 ("SUB_PREM_011","All Stories full-page subscriber coverage","Critical",[
  ("Open All Stories and scroll through the complete available feed.","Story cards continue loading through bounded scrolling.","tap All Stories; repeat swipe UP"),
  ("Capture top, middle, and lower feed states after waiting.","All Stories feed coverage is documented.","wait before screenshots"),
  ("Verify no inline, blank, sticky, iframe, Taboola, or interstitial ad appears.","All Stories is ad-free for the subscriber.","ASSERT_SUBSCRIBER_NO_MONETIZATION repeatedly")]),
 ("SUB_PREM_012","All Stories article full access","Blocker",[
  ("Open a current article from All Stories.","Article detail opens.","tap article card; assert screen_article_detail"),
  ("Scroll to the lower article and post-article area.","The article remains continuously readable.","repeat swipe UP"),
  ("Verify no paywall, Login offer, Subscribe action, or advertisement; capture evidence.","Full entitled access is demonstrated.","assert no restrictions/ads; wait; screenshot")]),
 ("SUB_PREM_013","Premium article reading options","Critical",[
  ("Open a Premium article and tap Text size or Reading Options.","The reading-options panel opens.","tap Text size; assert Reading Options/A-/A+"),
  ("Apply larger and smaller text settings.","The article responds and remains readable.","tap A+; close; wait; screenshot; reopen; tap A-"),
  ("Close the panel and verify entitled article access remains.","Article context remains open without monetisation.","close; assert article; assert no restrictions/ads")]),
 ("SUB_PREM_014","Premium article Bookmark","Critical",[
  ("Open a Premium article and tap Bookmark.","The bookmark action is accepted for the authenticated subscriber.","tap Bookmark"),
  ("Verify Login to your account does not appear.","Bookmark does not invoke an authentication gate.","assert article; assertNotVisible Login to your account"),
  ("Wait and capture the bookmarked article state.","Bookmark evidence is saved.","wait; screenshot")]),
 ("SUB_PREM_015","Premium article Share and Comment","Critical",[
  ("Tap Share in a Premium article and capture the settled Quick Share surface.","System sharing opens correctly.","tap Share; assert Quick Share; wait; screenshot; back"),
  ("Tap Post a comment when supported.","Comment access opens without subscriber login.","conditional tap Post a comment; assertNotVisible Login to your account"),
  ("Return and verify the same article remains open.","Article context and subscriber access are preserved.","back; assert screen_article_detail")]),
 ("SUB_PREM_016","Premium Listen and AI Summary entitlement","High",[
  ("Open an eligible Premium article and select Listen to article when available.","Audio opens without a subscription prompt.","conditional Listen; assert player; wait 30000"),
  ("Open AI Summary when available.","The entitled summary is readable without Subscribe.","conditional AI Summary; assertNotVisible Subscribe"),
  ("Capture settled media or summary evidence and return to the article.","Optional entitled features are documented safely.","wait; screenshot; back; assert article")]),
 ("SUB_PREM_017","Premium article paging and entitlement persistence","Blocker",[
  ("Open a Premium article and swipe left five times with waits.","Adjacent Premium articles can be paged left.","repeat 5 swipe LEFT; wait"),
  ("Swipe right five times with waits.","Paging works in the opposite direction.","repeat 5 swipe RIGHT; wait"),
  ("At each reached article verify no paywall, Login, Subscribe, or advertisement and capture evidence.","Entitlement persists across article paging.","assert no restrictions/ads; wait; screenshot")]),
 ("SUB_PREM_018","Subscriber Premium session persistence","Blocker",[
  ("Navigate Premium to an article, return Home, and reopen Premium.","All navigation completes without losing authentication.","Premium/article/Home/Premium round trip"),
  ("Verify Briefing remains unlocked and articles remain readable.","Premium entitlement persists throughout the round trip.","assertNotVisible Unlock Briefing/paywall/Login/Subscribe"),
  ("Verify ads remain absent and capture final Premium state after waiting.","The final Premium page is entitled and ad-free.","ASSERT_SUBSCRIBER_NO_MONETIZATION; wait; screenshot")]),
]

def main():
 wb=Workbook(); ws=wb.active; ws.title="Subscriber Premium"
 headers=["test_case_id","name","module","user_state","precondition","step_number","step","expected_result","priority","automatable","runtime_assertion"]
 ws.append(headers)
 for cid,name,priority,steps in CASES:
  for number,(step,expected,assertion) in enumerate(steps,1):
   ws.append([cid,name,"Premium","SUBSCRIBER",PRE,number,step,expected,priority,"Yes",assertion])
 for c in ws[1]: c.fill=PatternFill("solid",fgColor="17365D");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(horizontal="center")
 for i,w in enumerate([20,44,16,16,58,13,72,68,12,13,72],1): ws.column_dimensions[get_column_letter(i)].width=w
 for row in ws.iter_rows(min_row=2):
  for c in row:c.alignment=Alignment(wrap_text=True,vertical="top")
 ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions;OUTPUT.parent.mkdir(parents=True,exist_ok=True);wb.save(OUTPUT)
 print(f"Created {OUTPUT} with {len(CASES)} cases and {ws.max_row-1} steps")

if __name__=="__main__":main()

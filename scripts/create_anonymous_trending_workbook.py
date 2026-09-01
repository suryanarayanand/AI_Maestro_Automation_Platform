from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Anonymous_Trending_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
]

CASES = [
    ("ANON_TREND_001", "Anonymous Trending launch and selected navigation", [
        ("Launch the app in a fresh anonymous state and skip the welcome popup when it appears.", "Anonymous Home loads and the welcome popup is dismissed."),
        ("Tap the Trending icon in the bottom navigation.", "The Trending page opens."),
        ("Verify the Trending bottom-navigation icon and label are highlighted, then capture the Trending page.", "Trending is visibly selected and screenshot evidence is saved."),
    ]),
    ("ANON_TREND_002", "Anonymous Trending pull to refresh", [
        ("Open the Trending page as an anonymous user.", "The Trending screen and live content are visible."),
        ("Pull down from the Trending content area to refresh.", "The refresh gesture completes without leaving Trending."),
        ("Verify the Trending screen, selected Trending navigation, section tabs, and article content after refresh.", "Trending remains selected and usable with refreshed content."),
    ]),
    ("ANON_TREND_003", "Trending All feed scrolling and advertisement evidence", [
        ("Open Trending and select All.", "The All feed is visible."),
        ("Scroll through the All feed across multiple article cards and viewports, capturing representative screenshots.", "The feed scrolls continuously and article cards render."),
        ("When an Advertisement label or ad container appears, capture it before continuing.", "Rendered advertisement evidence is saved; absence of an ad in the dynamic feed is recorded as a valid conditional skip."),
        ("If the Advertisement container is blank or shows no usable creative, capture it as defect evidence.", "Blank-ad evidence is preserved for a bug report and the test does not attempt an unsafe blind tap."),
    ]),
    ("ANON_TREND_004", "Trending advertisement landing-page navigation", [
        ("Scroll the Trending All feed until a rendered, visibly clickable advertisement creative appears.", "A usable ad creative is found; if only blank ads or no ads appear, the landing-page branch is skipped with evidence."),
        ("Capture the rendered advertisement and tap its visible creative area.", "The advertisement landing page opens."),
        ("Capture the advertisement landing page and verify it is not a blank or unavailable web page.", "The external advertisement destination is visible and evidence is saved."),
        ("Press Back once.", "Only the advertisement destination closes and the Trending page is restored."),
    ]),
    ("ANON_TREND_005", "The Hindu masthead round-trip from Trending", [
        ("Open the Trending page and tap the The Hindu masthead at the top.", "The app returns to Home."),
        ("Verify the Home screen and bottom Home selection.", "Home is visible and usable."),
        ("Tap Trending again.", "The Trending page is restored with Trending selected."),
    ]),
    ("ANON_TREND_006", "Trending section-tab coverage", [
        ("Open Trending and select All, News, Business, Technology, Entertainment, and Sports one at a time.", "Each selected section loads inside the Trending screen."),
        ("For every section, verify its selected-tab indicator and capture the top of the page.", "The correct tab is visibly selected and screenshot evidence is saved for all six sections."),
        ("Scroll each section through representative content and capture any Advertisement container encountered.", "Each section remains usable while scrolling; advertisement presence or absence is recorded per section."),
    ]),
    ("ANON_TREND_007", "Trending Read Full Article and interstitial handling", [
        ("Open Trending All and tap READ FULL ARTICLE on a visible Trending card.", "The article-opening flow starts."),
        ("If an interstitial advertisement appears, capture it, wait in five-second intervals until Close becomes available, and close only the advertisement.", "The interstitial is documented and dismissed without closing the app; a delayed video close is handled safely."),
        ("Verify the article-detail screen and capture its header.", "The full article-detail page opens successfully."),
    ]),
    ("ANON_TREND_008", "Trending Premium article paywall restriction", [
        ("Open a Trending article and scroll through it while checking for a Premium badge or reader paywall.", "The article state is identified as Premium/paywalled or unrestricted."),
        ("When a paywall appears, verify an offer such as Keep reading free for 3 days or a percentage discount, plus Already a subscriber and Login or Sign in.", "The anonymous reader is blocked at the paywall and subscription/login options are visible."),
        ("Verify the Subscribe action at the bottom of the paywalled article and capture the paywall.", "Subscribe is visible and complete paywall evidence is saved."),
        ("If no paywall appears for the selected live article, record the paywall branch as a valid conditional skip and continue with unrestricted-article validation.", "Dynamic non-Premium content does not create a false failure."),
    ]),
    ("ANON_TREND_009", "Trending article Listen and AI Summary options", [
        ("Open a Trending article and inspect the article actions for Listen to article and AI Summary.", "Available optional article actions are identified."),
        ("When Listen to article is present, tap it and capture the resulting player or prompt, then return to the article.", "Listen behavior is documented without failing articles that do not support audio."),
        ("When AI Summary is present, tap it and capture either the complete summary or its anonymous subscription gate, then return to the article.", "The applicable AI Summary state is verified; absence on a short article is a valid conditional skip."),
    ]),
    ("ANON_TREND_010", "Trending article reading options and text size", [
        ("Open a Trending article and tap its reading-options or Text size control.", "The reading-options panel opens."),
        ("Verify the available reading controls and tap at least two text-size choices while observing the article text.", "Text-size controls respond and the article remains usable."),
        ("Close the reading-options panel.", "The same article-detail page is restored."),
    ]),
    ("ANON_TREND_011", "Trending article Bookmark login gate", [
        ("Open a Trending article as an anonymous user and tap Bookmark.", "Authentication is requested instead of saving an anonymous bookmark."),
        ("Verify Login to your account or equivalent sign-in text and capture the page.", "The bookmark login gate is visible and screenshot evidence is saved."),
        ("Press Back.", "The same article-detail page is restored."),
    ]),
    ("ANON_TREND_012", "Trending article Share sheet", [
        ("Open a Trending article and tap Share.", "The platform share sheet opens."),
        ("Verify a system share target such as Quick Share and capture the share sheet.", "The share interface is visible with available targets."),
        ("Press Back without sharing.", "The share sheet closes and the same article is restored."),
    ]),
    ("ANON_TREND_013", "Trending article Comment sign-in gate", [
        ("Open an unrestricted Trending article and scroll until Post a comment is visible.", "Post a comment is found; if a reader paywall blocks access first, capture the paywall and mark comment as not applicable for that article."),
        ("Tap Post a comment when available.", "The anonymous comment screen or prompt opens."),
        ("Verify sign-in wording such as SIGN IN AND JOIN THE CONVERSATION and capture it.", "Anonymous users are required to sign in before commenting."),
        ("Press Back without posting.", "The article-detail page is restored and no comment is submitted."),
    ]),
    ("ANON_TREND_014", "Trending cross-section article outcome coverage", [
        ("From News, Business, Technology, Entertainment, and Sports, open a representative article in each section.", "Each available section article opens; a section with no current article is recorded with page evidence."),
        ("For every opened article, handle any interstitial by capturing it, waiting for Close, and closing only the ad.", "Interstitials do not terminate the app or invalidate article verification."),
        ("Scroll each article and classify it as paywalled or unrestricted.", "Paywalled articles show subscription/login restrictions; unrestricted articles remain readable."),
        ("For unrestricted articles, verify available Post a comment, Related Topics, Recommended, or Headlines content near the end.", "Applicable post-article sections are documented; a paywall-blocked article is not incorrectly required to expose them."),
    ]),
    ("ANON_TREND_015", "Trending News section content and advertisement coverage", [
        ("Open Trending, select News, and verify the News selected-tab indicator.", "The News feed loads and News is visibly selected."),
        ("Scroll the News feed through multiple cards and capture top, middle, and lower-page evidence.", "News content remains usable throughout the scroll."),
        ("Capture any rendered or blank Advertisement container encountered.", "News advertisement state is documented; no live ad is a valid conditional result."),
        ("Open a News article and return with Back.", "The article opens and Back restores the News section."),
    ]),
    ("ANON_TREND_016", "Trending Business section content and advertisement coverage", [
        ("Open Trending, select Business, and verify the Business selected-tab indicator.", "The Business feed loads and Business is visibly selected."),
        ("Scroll the Business feed through multiple cards and capture top, middle, and lower-page evidence.", "Business content remains usable throughout the scroll."),
        ("Capture any rendered or blank Advertisement container encountered.", "Business advertisement state is documented; no live ad is a valid conditional result."),
        ("Open a Business article and return with Back.", "The article opens and Back restores the Business section."),
    ]),
    ("ANON_TREND_017", "Trending Technology section content and advertisement coverage", [
        ("Open Trending, select Technology, and verify the Technology selected-tab indicator.", "The Technology feed loads and Technology is visibly selected."),
        ("Scroll the Technology feed through multiple cards and capture top, middle, and lower-page evidence.", "Technology content remains usable throughout the scroll."),
        ("Capture any rendered or blank Advertisement container encountered.", "Technology advertisement state is documented; no live ad is a valid conditional result."),
        ("Open a Technology article and return with Back.", "The article opens and Back restores the Technology section."),
    ]),
    ("ANON_TREND_018", "Trending Entertainment section content and advertisement coverage", [
        ("Open Trending, select Entertainment, and verify the Entertainment selected-tab indicator.", "The Entertainment feed loads and Entertainment is visibly selected."),
        ("Scroll the Entertainment feed through multiple cards and capture top, middle, and lower-page evidence.", "Entertainment content remains usable throughout the scroll."),
        ("Capture any rendered or blank Advertisement container encountered.", "Entertainment advertisement state is documented; no live ad is a valid conditional result."),
        ("Open an Entertainment article and return with Back.", "The article opens and Back restores the Entertainment section."),
    ]),
    ("ANON_TREND_019", "Trending Sports section content and advertisement coverage", [
        ("Open Trending, select Sports, and verify the Sports selected-tab indicator.", "The Sports feed loads and Sports is visibly selected."),
        ("Scroll the Sports feed through multiple cards and capture top, middle, and lower-page evidence.", "Sports content remains usable throughout the scroll."),
        ("Capture any rendered or blank Advertisement container encountered.", "Sports advertisement state is documented; no live ad is a valid conditional result."),
        ("Open a Sports article and return with Back.", "The article opens and Back restores the Sports section."),
    ]),
    ("ANON_TREND_020", "Trending article pager in both directions", [
        ("Open a full Trending article and confirm the article-detail screen.", "A Trending article is open."),
        ("Swipe left through five article positions, handling and capturing any interstitial advertisement before closing only the ad.", "Article paging continues to the left without terminating the app."),
        ("Swipe right through five article positions with the same interstitial handling.", "Article paging continues to the right and returns toward the original position."),
        ("Capture any Premium badge or paywall encountered during paging.", "Dynamic Premium discovery is documented without requiring every feed to contain Premium content."),
    ]),
    ("ANON_TREND_021", "Trending Subscribe entry point", [
        ("Open Trending as an anonymous user and verify the top Subscribe action.", "Subscribe is visible for the anonymous state."),
        ("Tap Subscribe and capture the resulting subscription-offer page.", "A subscription plan or offer destination opens."),
        ("Press Back.", "Trending is restored with the anonymous session retained."),
    ]),
    ("ANON_TREND_022", "Trending paywall subscriber-login link", [
        ("Open and scroll Trending articles until a reader paywall with Already a subscriber is found.", "A paywalled article is found, or the dynamic search completes with valid no-paywall evidence."),
        ("Tap the Login or Sign in link associated with Already a subscriber.", "The account login page opens."),
        ("Verify the login-page identity, capture it, and press Back.", "The correct authentication destination is documented and the paywalled article is restored."),
    ]),
    ("ANON_TREND_023", "Trending section and scroll-state restoration", [
        ("Select a non-All Trending section and scroll to a clearly different article position.", "The chosen section is selected at a non-top position."),
        ("Open an article and then press Back.", "The same Trending section is restored."),
        ("Verify the selected section and inspect whether the prior feed position is retained.", "Section selection is preserved; scroll-position behavior is captured for regression comparison."),
    ]),
    ("ANON_TREND_024", "Trending interstitial video delayed-close behavior", [
        ("Open Trending articles until an interstitial advertisement appears during the bounded search.", "An interstitial is found, or the run records that none appeared within the search limit."),
        ("Immediately capture the interstitial and verify the app is not closed while Close is unavailable.", "The initial video-ad state is documented."),
        ("Wait in five-second intervals for a maximum bounded duration, then capture and tap Close when it becomes available.", "The delayed close control is handled safely."),
        ("If Close never appears, capture the blocked state and use Back only when the ad still covers the article.", "The ad activity closes without closing The Hindu, and the article screen is restored."),
    ]),
    ("ANON_TREND_025", "Trending blank and overlapping advertisement classification", [
        ("Scroll Trending feeds and articles while checking Advertisement containers and sticky ad areas.", "Advertisement states are observed across representative screens."),
        ("When an Advertisement container is blank, capture the complete screen without tapping the blank area.", "Blank-ad defect evidence is saved for review."),
        ("When a sticky advertisement appears without obscuring content or controls, capture it as an acceptable state.", "A non-blocking sticky ad is documented and does not fail the case."),
        ("When a non-sticky advertisement overlaps content or controls, capture the overlap and mark it for defect review.", "Only intrusive or broken advertising is classified as an issue."),
    ]),
    ("ANON_TREND_026", "Trending long-article reading evidence", [
        ("Open a long Trending article and capture its header and available article actions.", "The article identity and initial controls are documented."),
        ("Scroll through the article body and capture at least four screenshots distributed from top to lower content.", "Long-form rendering and continuity are documented."),
        ("During scrolling, handle any interstitial and classify the article as paywalled or unrestricted.", "Advertising and entitlement states are handled without losing the article."),
        ("For an unrestricted article, continue to the available post-article content; for a paywalled article, stop at and capture the blocking offer.", "The correct endpoint for the observed article type is verified."),
    ]),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous Trending"
    sheet.append(HEADERS)
    precondition = (
        "The Hindu app is installed; network is available; the user is signed out; "
        "dynamic advertisements and article entitlements may vary between runs."
    )
    for case_id, name, steps in CASES:
        for number, (step, expected) in enumerate(steps, 1):
            sheet.append([
                case_id, name, "Trending", "ANONYMOUS", precondition,
                number, step, expected, "High", "Yes",
            ])

    header_fill = PatternFill("solid", fgColor="A71930")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [20, 48, 18, 16, 74, 14, 88, 88, 12, 14]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()

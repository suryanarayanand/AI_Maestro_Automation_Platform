"""Normalize Anonymous Home workbook metadata after the evidence-based rerun review."""

from pathlib import Path
import json
import shutil

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Uploads" / "Anonymous_Home_Approved_Test_Cases.xlsx"
BACKUP = ROOT / "Backups" / "UploadsCleanup" / "Anonymous_Home_Approved_Test_Cases_before_home_metadata.xlsx"

CASE_METADATA = {
    "ANON_HOME_001": ("Launch anonymous Home and dismiss onboarding", "Onboarding"),
    "ANON_HOME_002": ("Refresh Home and verify stable content", "Feed"),
    "ANON_HOME_003": ("Verify advertisements across the Home feed", "Advertising"),
    "ANON_HOME_004": ("Verify the sticky Home advertisement while scrolling", "Advertising"),
    "ANON_HOME_005": ("Verify Taboola recommendations in the lower Home feed", "Recommendations"),
    "ANON_HOME_006": ("Open a Home article and validate article controls", "Article Controls"),
    "ANON_HOME_007": ("Validate article content and post-article sections", "Article Content"),
    "ANON_HOME_008": ("Page through Home articles and capture Premium evidence", "Article Paging"),
    "ANON_HOME_009": ("Open a Premium article and validate anonymous restriction", "Premium"),
    "ANON_HOME_010": ("Validate the anonymous AI Summary subscription gate", "AI Summary"),
    "ANON_HOME_011": ("Validate the anonymous Bookmark login gate", "Engagement"),
    "ANON_HOME_012": ("Validate the article Share sheet", "Engagement"),
    "ANON_HOME_013": ("Validate the anonymous Comment sign-in gate", "Engagement"),
    "ANON_HOME_014": ("Return from an article to anonymous Home", "Navigation"),
}


def main():
    BACKUP.parent.mkdir(parents=True, exist_ok=True)
    if not BACKUP.exists():
        shutil.copy2(WORKBOOK, BACKUP)

    workbook = load_workbook(WORKBOOK)
    sheet = workbook["Anonymous Home"]
    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}
    section_column = headers.get("section") or sheet.max_column + 1
    sheet.cell(1, section_column, "section")

    updated = set()
    for row in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row, headers["test_case_id"]).value or "").strip()
        if case_id not in CASE_METADATA:
            continue
        name, section = CASE_METADATA[case_id]
        sheet.cell(row, headers["name"], name)
        sheet.cell(row, headers["module"], "Home")
        sheet.cell(row, section_column, section)
        updated.add(case_id)

    workbook.save(WORKBOOK)

    suite_path = ROOT / "Suites" / "anonymous_home_page.json"
    suite = json.loads(suite_path.read_text(encoding="utf-8"))
    suite["suite"] = "Anonymous User - Home"
    suite["user_state"] = "ANONYMOUS"
    for test in suite.get("tests", []):
        metadata = CASE_METADATA.get(test.get("id"))
        if not metadata:
            continue
        test["name"], test["section"] = metadata
        test["module"] = "Home"
        test["user_state"] = "ANONYMOUS"
    suite_path.write_text(json.dumps(suite, indent=4) + "\n", encoding="utf-8")
    print(f"Updated {len(updated)} cases; backup: {BACKUP}")


if __name__ == "__main__":
    main()

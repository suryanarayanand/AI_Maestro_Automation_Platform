import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


def prepare(source, output, size=10):
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    normalized = [str(value or "").strip().casefold() for value in header]
    id_index = next(
        (index for index, value in enumerate(normalized)
         if value in {"test_case_id", "test case id", "scenario id", "scenario no"}),
        0,
    )
    selected_ids = []
    selected_rows = []
    for row in rows:
        case_id = str(row[id_index] or "").strip()
        if not case_id:
            continue
        if case_id not in selected_ids:
            if len(selected_ids) >= size:
                continue
            selected_ids.append(case_id)
        if case_id in selected_ids:
            selected_rows.append(row)
    result = Workbook()
    result.remove(result.active)
    for source_sheet in workbook.worksheets:
        source_rows = source_sheet.iter_rows(values_only=True)
        source_header = next(source_rows)
        source_normalized = [str(value or "").strip().casefold() for value in source_header]
        source_id_index = next(
            (index for index, value in enumerate(source_normalized)
             if value in {"test_case_id", "test case id", "scenario id", "scenario no"}),
            0,
        )
        target = result.create_sheet(source_sheet.title)
        target.append(source_header)
        for row in source_rows:
            if str(row[source_id_index] or "").strip() in selected_ids:
                target.append(row)
    workbook.close()
    result.save(output)
    print(f"Created {output} with {len(selected_ids)} cases and {len(selected_rows)} steps")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("output")
    parser.add_argument("--size", type=int, default=10)
    arguments = parser.parse_args()
    prepare(Path(arguments.source), Path(arguments.output), arguments.size)

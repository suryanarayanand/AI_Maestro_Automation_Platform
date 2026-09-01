"""Create the Subscriber Games module import workbook from validated SC-27 flows."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT=Path("Uploads/Ready/Subscriber_Games_Approved_Test_Cases.xlsx")
PRE=("The Hindu app is installed; network is available; configured masked credentials "
     "belong to an active subscriber with Games entitlement.")
CASES=[
 ("SUB_GAMES_001","Subscriber Games landing and entitlement","Blocker",[
  ("Authenticate with configured subscriber credentials and tap Games.","The Games screen opens and Games navigation is selected.","OPEN_SUBSCRIBER_GAMES; assert screen_games and nav_games"),
  ("Wait for the page to settle and verify the entitled Games catalogue is visible.","Playable game cards are rendered for the subscriber.","assert named game content; wait before screenshot"),
  ("Verify Login to play, Subscribe, plan offers, access errors, and advertisements are absent.","The landing page reflects active subscriber entitlement.","assertNotVisible Login to play/Subscribe/plan/access error/ads")]),
 ("SUB_GAMES_002","Subscriber Games catalogue coverage","Critical",[
  ("Scroll through the Games catalogue from top to bottom.","All available game groups load without a broken or blank page.","bounded scroll; assert screen_games"),
  ("Verify Cryptic Crossword, Sudoku, The Hindu Mini, Easy Down, Word Row, Word Flower, Word Search, and Quiz entries.","Every supported named game is represented in the catalogue.","assert or scrollUntilVisible each named game"),
  ("Capture top, middle, and lower catalogue positions after loading waits.","Complete catalogue evidence is saved.","waitForAnimationToEnd before each screenshot")]),
 ("SUB_GAMES_003","Cryptic Crossword subscriber play access","Blocker",[
  ("Open Cryptic Crossword and select a current puzzle.","The selected crossword reaches its play-mode entry.","SC_27_CRYPTIC_CROSSWORD navigation and current puzzle selector"),
  ("Choose Play solo when the play-mode prompt appears.","The crossword board opens without a purchase or login gate.","conditional Play solo; assertNotVisible access error/Login/Subscribe"),
  ("Verify Settings, Assist, and game navigation; wait and capture the board.","Core crossword controls are visible and documented.","assert Settings/Assist/main-nav; wait; screenshot")]),
 ("SUB_GAMES_004","Sudoku difficulty catalogue","Critical",[
  ("Open Sudoku from Games.","Choose Your Puzzle to Play is displayed.","tap Sudoku; assert Choose Your Puzzle to Play"),
  ("Verify Mini, Easy, Medium, Hard, and Killer difficulty options.","All expected Sudoku difficulties are available.","assert each difficulty label"),
  ("Wait and capture the difficulty screen without monetisation.","Sudoku selection evidence is saved with no restriction.","assert no Login/Subscribe/ads; wait; screenshot")]),
 ("SUB_GAMES_005","Sudoku Mini subscriber gameplay entry","Blocker",[
  ("Open Sudoku, choose Mini, and select a current Mini Sudoku puzzle.","The play-mode prompt or board opens for the selected Mini puzzle.","SC_27_SUDOKU Mini selectors"),
  ("Choose Play solo when shown.","Mini Sudoku opens without an access error.","tap Play solo; assertNotVisible access was not found"),
  ("Verify Settings, Assist, and main navigation; capture after waiting.","The entitled Mini Sudoku board and controls are usable.","assert Settings/Assist/main-nav; wait; screenshot")]),
 ("SUB_GAMES_006","Sudoku Easy subscriber gameplay entry","Blocker",[
  ("Open Sudoku, choose Easy, and select a current Easy Sudoku puzzle.","The selected Easy puzzle reaches play mode.","SC_27_SUDOKU_EASY selectors"),
  ("Choose Play solo and verify no login or access restriction.","The Easy board opens for the subscriber.","tap Play solo; assertNotVisible Login/access error/Subscribe"),
  ("Verify Settings, Assist, and game navigation; capture after waiting.","Core Easy Sudoku controls are documented.","assert Settings/Assist/main-nav; wait; screenshot")]),
 ("SUB_GAMES_007","The Hindu Mini subscriber play access","Blocker",[
  ("Scroll to The Hindu Mini, open it, and select a current dated puzzle.","The selected puzzle reaches play-mode entry.","SC_27_THE_HINDU_MINI selectors and month regex"),
  ("Dismiss optional prompts or choose Play solo.","The playable Mini board opens without entitlement errors.","conditional Dismiss/Play solo; assert no access error"),
  ("Verify Settings, Assist, and main navigation; wait and capture.","The Hindu Mini gameplay entry is documented.","assert Settings/Assist/main-nav; wait; screenshot")]),
 ("SUB_GAMES_008","Easy Down subscriber play access","Blocker",[
  ("Scroll to Easy Down, open it, and select a current dated puzzle.","The Easy Down puzzle reaches play-mode entry.","SC_27_EASY_DOWN selectors and month regex"),
  ("Dismiss optional prompts or choose Play solo.","The board opens without a login or access error.","conditional Dismiss/Play solo; assert no restriction"),
  ("Verify Settings, Assist, and main navigation; wait and capture.","Core Easy Down controls are visible.","assert Settings/Assist/main-nav; wait; screenshot")]),
 ("SUB_GAMES_009","Word Row subscriber play access","Blocker",[
  ("Scroll to Word Row, open it, and select a current dated puzzle.","The selected Word Row puzzle reaches play mode.","SC_27_WORD_ROW selectors and month regex"),
  ("Dismiss optional prompts or choose Play solo.","Word Row opens without entitlement failure.","conditional Dismiss/Play solo; assert no access error/Login"),
  ("Wait and capture the playable state with no advertisements.","Entitled Word Row evidence is saved.","assert no ads; wait; screenshot")]),
 ("SUB_GAMES_010","Word Flower subscriber play access","Blocker",[
  ("Scroll to Word Flower, open it, and select an available puzzle.","The selected Word Flower puzzle reaches play mode.","SC_27_WORD_FLOWER selectors"),
  ("Dismiss optional prompts or choose Play solo.","Word Flower opens without a subscriber restriction.","conditional Dismiss/Play solo; assert no access error/Login"),
  ("Verify Hint, Reveal, and main navigation; wait and capture.","Core Word Flower controls are visible and documented.","assert Hint/Reveal/main-nav; wait; screenshot")]),
 ("SUB_GAMES_011","Word Search subscriber play access","Blocker",[
  ("Scroll to Word Search, open it, and select a current dated puzzle.","The selected Word Search puzzle reaches play mode.","SC_27_WORD_SEARCH selectors and month regex"),
  ("Dismiss optional prompts or choose Play solo.","Word Search opens without entitlement failure.","conditional Dismiss/Play solo; assert no access error/Login"),
  ("Verify Settings, Assist, and main navigation; wait and capture.","Core Word Search controls are visible.","assert Settings/Assist/main-nav; wait; screenshot")]),
 ("SUB_GAMES_012","Quiz subscriber access","Critical",[
  ("Scroll to Quiz and open it.","The News Quiz or Quiz destination opens.","SC_27_NEWS_QUIZ ^Quiz$ selector"),
  ("Verify Quiz content is visible without Login to play, Subscribe, or access error.","The subscriber can access Quiz content.","assert Quiz; assertNotVisible restrictions"),
  ("Wait and capture the Quiz state with no advertisements.","Entitled Quiz evidence is saved.","assert no ads; wait; screenshot")]),
 ("SUB_GAMES_013","Games navigation and entitlement persistence","Critical",[
  ("Open a playable game, return to Games, navigate Home, and reopen Games.","Back and bottom navigation restore the expected screens.","game/Games/Home/Games round trip"),
  ("Open a second game after returning.","Subscriber access remains active without reauthentication.","assertNotVisible Login to play/Login to your account/Subscribe"),
  ("Verify no advertisements or access errors and capture the final settled state.","Games entitlement persists across navigation.","ASSERT_SUBSCRIBER_NO_MONETIZATION; assert no access error; wait; screenshot")]),
]

def main():
 wb=Workbook();ws=wb.active;ws.title="Subscriber Games"
 h=["test_case_id","name","module","user_state","precondition","step_number","step","expected_result","priority","automatable","runtime_assertion"]
 ws.append(h)
 for cid,name,priority,steps in CASES:
  for no,(step,expected,assertion) in enumerate(steps,1):ws.append([cid,name,"Games","SUBSCRIBER",PRE,no,step,expected,priority,"Yes",assertion])
 for c in ws[1]:c.fill=PatternFill("solid",fgColor="17365D");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(horizontal="center")
 for i,w in enumerate([20,44,16,16,58,13,72,68,12,13,72],1):ws.column_dimensions[get_column_letter(i)].width=w
 for row in ws.iter_rows(min_row=2):
  for c in row:c.alignment=Alignment(wrap_text=True,vertical="top")
 ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions;OUTPUT.parent.mkdir(parents=True,exist_ok=True);wb.save(OUTPUT)
 print(f"Created {OUTPUT} with {len(CASES)} cases and {ws.max_row-1} steps")
if __name__=="__main__":main()

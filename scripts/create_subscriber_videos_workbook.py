from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'Uploads'/'Ready'/'Subscriber_Videos_Quick_Access_Approved_Test_Cases.xlsx'
HEAD=['test_case_id','name','module','user_state','precondition','step_number','step','expected_result','priority','automatable','master_sheet_reference','existing_yaml_reference']
PRE='The app, network and Maestro device are available; valid subscriber credentials come from protected portal configuration. Use a currently playable controlled Video article. Wait before screenshots.'
MASTER='TH App Testing Scenarios_AutomationCopy.xlsx (6000+ scenarios); THG App Functionality Matrix; screen-20260827-161150-1787827145618.mp4'
REF='SC_70_Video_page_subscribed_account.yaml; anonymous Videos Quick Access YAML; subscriber Article Page YAML'
CASES=[
('SUB_VIDEO_001','Videos Quick Access landing and selected state','Blocker',[('Log in, open Quick Access, and tap Videos.','Videos opens and the Videos tab is highlighted.'),('Verify video cards and capture the settled listing.','Playable video content is displayed without layout overlap.')]),
('SUB_VIDEO_002','Video article identity and metadata','Critical',[('Open a current video card.','The Video article opens with its player and headline.'),('Verify section label, available reading time, description, author and published/updated metadata while scrolling.','Available identity and metadata remain readable.')]),
('SUB_VIDEO_003','Video play, pause and progress','Blocker',[('Tap Play and allow playback to progress.','Playback starts and elapsed time advances.'),('Tap Pause, wait, then resume.','Progress stops while paused and continues after resume.')]),
('SUB_VIDEO_004','Video 15-second seek controls','Critical',[('During playback, tap forward 15 seconds.','Elapsed playback position advances approximately 15 seconds.'),('Tap rewind 15 seconds.','Playback position moves backward without restarting or freezing.')]),
('SUB_VIDEO_005','Fullscreen entry, playback and exit','Critical',[('Enter fullscreen during playback.','The player expands to fullscreen landscape presentation.'),('Verify playback controls, then exit fullscreen.','The article returns to portrait and playback remains available inline.')]),
('SUB_VIDEO_006','Playback rate and quality settings','Critical',[('Open the player More options menu.','Playback Rates and Quality controls are visible.'),('Choose a non-default playback rate and verify its indicator, then restore Normal.','Playback speed changes and returns to Normal.'),('Open Quality and select an available quality or Auto.','The selected quality is applied without ending playback.')]),
('SUB_VIDEO_007','Timeline seek and player recovery','Major',[('Drag or tap the timeline to a safe later position.','The elapsed position updates and playback resumes from the selected point.'),('Pause, resume and close the control overlay.','The player remains responsive after seeking.')]),
('SUB_VIDEO_008','Inline playback during article scrolling','Critical',[('Exit fullscreen with video playing and scroll below the player.','The article body can be scrolled while inline playback remains stable.'),('Return to the player and inspect its state.','The player retains a valid playback position and controls.')]),
('SUB_VIDEO_009','Video article actions','Critical',[('Scroll to expose Reading Options, Bookmark, Share and Comment controls.','The floating action bar displays the available controls.'),('Exercise Reading Options, toggle Bookmark safely, open and dismiss Share, then open Comment without posting.','Each action responds without Login redirection or external mutation.')]),
('SUB_VIDEO_010','Post-comment and related-content sections','Critical',[('Scroll below the video article to Post a Comment and Related Topics.','Comment and configured topic content are accessible to the subscriber.'),('Verify Recommended and Headlines tabs.','Both related-content tabs are visible and selectable.')]),
('SUB_VIDEO_011','Recommended video navigation','Critical',[('Open a video item from Recommended or Headlines.','A different Video article and player open.'),('Start playback, then navigate Back.','The recommended video plays and Back restores the source article.')]),
('SUB_VIDEO_012','Subscriber Video entitlement and no monetisation','Blocker',[('Traverse the Videos listing, one Video article, fullscreen playback and lower article content.','The complete subscriber video journey remains accessible.'),('Verify Advertisement, ad iframe, Taboola advertising unit, Subscribe, Login and paywall controls are absent.','No monetisation surface blocks or interrupts the subscriber.')]),
('SUB_VIDEO_013','Player close and article recovery','Major',[('Start playback, open fullscreen, then use the player close control.','Fullscreen/player overlay closes without closing the app.'),('Verify the same Video article remains visible and can play again.','The article and player recover correctly.')]),
]
def main():
 wb=Workbook();ws=wb.active;ws.title='Subscriber Videos';ws.append(HEAD)
 for cid,name,pri,steps in CASES:
  for n,(step,exp) in enumerate(steps,1):ws.append([cid,name,'Videos','SUBSCRIBER',PRE,n,step,exp,pri,'Yes',MASTER,REF])
 for c in ws[1]:c.fill=PatternFill('solid',fgColor='273C75');c.font=Font(color='FFFFFF',bold=True)
 for i,w in enumerate([22,52,18,16,86,13,94,94,14,14,72,100],1):ws.column_dimensions[ws.cell(1,i).column_letter].width=w
 for row in ws.iter_rows():
  for c in row:c.alignment=Alignment(vertical='top',wrap_text=True)
 ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions;OUT.parent.mkdir(parents=True,exist_ok=True);wb.save(OUT)
 print(f'Created {OUT} with {len(CASES)} cases and {sum(len(c[3]) for c in CASES)} steps')
if __name__=='__main__':main()

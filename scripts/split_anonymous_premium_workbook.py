from copy import copy
from pathlib import Path

from openpyxl import load_workbook

source = Path("Uploads/Ready/Anonymous_Premium_Approved_Test_Cases.xlsx")
wb = load_workbook(source)
ws = wb.active
headers = {str(cell.value): cell.column for cell in ws[1]}
case_col = headers["test_case_id"]
case_ids = list(dict.fromkeys(str(ws.cell(row, case_col).value) for row in range(2, ws.max_row + 1)))

for batch_number, selected in enumerate((set(case_ids[:25]), set(case_ids[25:])), start=1):
    batch_wb = load_workbook(source)
    batch_ws = batch_wb.active
    for row in range(batch_ws.max_row, 1, -1):
        if str(batch_ws.cell(row, case_col).value) not in selected:
            batch_ws.delete_rows(row)
    output = source.with_name(f"Anonymous_Premium_Generation_Batch_{batch_number}.xlsx")
    batch_wb.save(output)
    print(output, len(selected), batch_ws.max_row - 1)

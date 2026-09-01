from pathlib import Path
from openpyxl import load_workbook

path = Path(r"C:\Users\12503\Downloads\THG App_Functionality_Matrix.xlsx")
wb = load_workbook(path, read_only=True, data_only=True)
print("workbook", path, "sheets", wb.sheetnames)
for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    print(f"\nSHEET {ws.title} rows={ws.max_row} cols={ws.max_column}")
    for number, row in enumerate(rows[:12], start=1):
        values = [str(value).strip() if value is not None else "" for value in row]
        print(number, " | ".join(values))

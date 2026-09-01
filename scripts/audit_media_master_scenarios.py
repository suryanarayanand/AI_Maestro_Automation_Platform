from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


path = Path("Uploads/Source/TH App Testing Scenarios_AutomationCopy.xlsx")
workbook = load_workbook(path, read_only=True, data_only=True)
matches = []
for sheet in workbook.worksheets:
    for number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        values = [str(value).strip() for value in row if value not in (None, "")]
        text = " | ".join(values)
        lower = text.lower()
        if "video" in lower or "photo" in lower:
            matches.append((sheet.title, number, text))
workbook.close()
print("matches", len(matches), "by_sheet", dict(Counter(item[0] for item in matches)))

categories = {
    "navigation_landing": ("navigate", "landing", "video page", "photos page", "photo page"),
    "refresh_scroll_pagination": ("refresh", "scroll", "show more", "pagination"),
    "open_content": ("open", "click", "tap", "thumbnail"),
    "playback_controls": ("play", "pause", "rewind", "forward", "seek"),
    "fullscreen_orientation": ("full screen", "fullscreen", "orientation", "landscape"),
    "speed_quality": ("speed", "quality", "resolution"),
    "gallery_swipe": ("swipe", "next photo", "previous photo", "gallery"),
    "metadata": ("caption", "credit", "headline", "author", "date"),
    "share_comment": ("share", "comment"),
    "access_subscription": ("subscribe", "subscriber", "paywall", "premium", "login"),
    "advertising": ("advertisement", " ads", "ad ", "interstitial"),
    "theme_visual": ("dark theme", "light theme", "ui", "layout"),
    "error_recovery": ("error", "timeout", "retry", "network"),
}
for name, terms in categories.items():
    found = [item for item in matches if any(term in item[2].lower() for term in terms)]
    refs = ", ".join(f"{s}!{n}" for s, n, _ in found[:5])
    print(f"{name}: {len(found)} rows; samples: {refs or '-'}")

print("\nRepresentative rows:")
seen = set()
for sheet, number, text in matches:
    normalized = text.lower()[:100]
    if normalized in seen:
        continue
    seen.add(normalized)
    print(f"{sheet}!{number}: {text[:350]}")
    if len(seen) == 20:
        break

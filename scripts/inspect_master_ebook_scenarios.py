from collections import Counter
from openpyxl import load_workbook

path = "Uploads/Source/TH App Testing Scenarios_AutomationCopy.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)
matches = []
for ws in wb.worksheets:
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    for number, values in enumerate(rows, start=2):
        record = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
        text = " | ".join(str(value or "") for value in values)
        if "ebook" in text.casefold() or "e-book" in text.casefold():
            matches.append((ws.title, number, record))

print("matches", len(matches))
for sheet, number, record in matches:
    useful = [
        str(value).strip() for key, value in record.items()
        if value and any(token in key.casefold() for token in (
            "component", "module", "scenario", "description", "expected", "priority", "type"
        ))
    ]
    print(sheet, number, " | ".join(dict.fromkeys(useful)))

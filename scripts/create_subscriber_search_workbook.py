from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Subscriber_Search_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "master_sheet_reference", "existing_yaml_reference",
]
PRE = (
    "The Hindu app is installed; network and Maestro device are available; valid subscriber credentials "
    "are supplied through protected portal configuration. Login before evidence capture."
)
REF = "SC_35_SearchFuction_subscriber.yaml; SC_34_Search_functionality_anonymous_account.yaml; SC_40_search_specialCharacter.yaml; CPLX_001.yaml; TH_0012.yaml"
MASTER = "TH App Testing Scenarios_AutomationCopy.xlsx (6000+ scenarios)"
CASES = [
    ("SUB_SEARCH_001", "Hamburger Search entry and selected screen", "Blocker", [
        ("Log in as the subscriber, open Hamburger, and verify the Search action.", "Hamburger opens and the Search action is visible and enabled."),
        ("Tap Search, wait for loading, and capture the screen.", "Search opens with an editable search input and no app exit."),
    ]),
    ("SUB_SEARCH_002", "Search close and Home recovery", "Critical", [
        ("Open Search from Hamburger and navigate Back without entering text.", "Home is restored and remains interactive."),
        ("Open Hamburger again.", "Hamburger and Search remain reusable after recovery."),
    ]),
    ("SUB_SEARCH_003", "Valid city keyword result list", "Blocker", [
        ("Search for Chennai and submit from the keyboard.", "A results page or valid no-results state loads without an error."),
        ("Wait and capture the result state.", "Search response evidence is saved."),
    ]),
    ("SUB_SEARCH_004", "Valid topic keyword result list", "Critical", [
        ("Search for the common news word flood and submit.", "Flood-related result cards or a valid current no-results state are displayed."),
        ("Clear the query, search for war, and scroll the refreshed results.", "War-related results replace the earlier query and scrolling remains responsive."),
    ]),
    ("SUB_SEARCH_005", "Search result article opening", "Blocker", [
        ("Run a valid search and open the first currently visible result card.", "The selected article detail page opens."),
        ("Verify the article screen, wait, and capture it.", "Article evidence is saved and the app remains stable."),
    ]),
    ("SUB_SEARCH_006", "Subscriber full article entitlement from Search", "Blocker", [
        ("Open an article from Search results and scroll through its body.", "The subscriber can read available article content."),
        ("Verify Subscribe, Already a subscriber Login, and paywall offer controls are absent.", "No anonymous monetisation restriction blocks the entitled subscriber."),
    ]),
    ("SUB_SEARCH_007", "Search article toolbar controls", "Critical", [
        ("Open an article from Search and verify Text size, Bookmark, and Share controls.", "The article toolbar controls are present and enabled."),
        ("Open and close Text size, then open Share and return without sharing externally.", "Both controls respond and the article is restored."),
    ]),
    ("SUB_SEARCH_008", "Search article Bookmark subscriber behavior", "Critical", [
        ("Open a Search result article and tap Bookmark.", "Bookmark responds without redirecting the subscriber to Login."),
        ("Capture the bookmarked state and restore it when safely supported.", "Bookmark evidence is saved without ending the subscriber session."),
    ]),
    ("SUB_SEARCH_009", "Search article AI Summary conditional behavior", "Major", [
        ("Open a sufficiently long Search result article and look for AI Summary in bounded scrolling.", "AI Summary may appear only for eligible article length."),
        ("If available, open it and verify full subscriber summary with no Subscribe restriction; otherwise capture the eligible skip state.", "Conditional AI Summary behavior is correctly documented."),
    ]),
    ("SUB_SEARCH_010", "No-result keyword handling", "Critical", [
        ("Search for a unique nonsense term such as zxqv987654321.", "A clear no-results or empty state appears without stale previous results."),
        ("Wait and capture the state.", "No-result behavior is recorded without a crash."),
    ]),
    ("SUB_SEARCH_011", "Special-character input safety", "Critical", [
        ("Enter a bounded special-character string and submit it.", "Search handles the input safely and shows results, validation, or a no-results state."),
        ("Verify the app remains responsive and capture the outcome.", "No crash, script injection, or malformed layout occurs."),
    ]),
    ("SUB_SEARCH_012", "Empty and whitespace search validation", "Major", [
        ("Submit Search with an empty value, then with whitespace only.", "The app prevents submission or shows a controlled empty/no-results state."),
        ("Verify no stale article result opens.", "Empty input does not navigate to unrelated content."),
    ]),
    ("SUB_SEARCH_013", "Long search input boundary", "Critical", [
        ("Enter a long but bounded text query and submit it.", "The input is accepted or limited without freezing, clipping, or crashing."),
        ("Wait and capture the handled result.", "Boundary behavior is documented."),
    ]),
    ("SUB_SEARCH_014", "Clear and replace search query", "Major", [
        ("Search for Chennai, return to the input, clear it, and enter Bangalore.", "The old query is removed and the new query is accepted."),
        ("Submit and verify the updated result state.", "Results correspond to the replacement query without stale content."),
    ]),
    ("SUB_SEARCH_015", "Search result back-navigation state", "Major", [
        ("Search for a valid term, open one result, and navigate Back.", "The Search results page is restored."),
        ("Navigate Back again and verify Home, then reopen Hamburger Search.", "The navigation stack recovers correctly and Search remains reusable."),
    ]),
    ("SUB_SEARCH_016", "Subscriber Search no-advertisement contract", "Blocker", [
        ("Search, scroll the result list, open an article, and scroll it.", "Search and article screens remain usable across representative content."),
        ("Verify Advertisement, ad iframe, Taboola advertising unit, and Subscribe controls are absent.", "Subscriber no-monetisation behavior is enforced throughout Search."),
    ]),
    ("SUB_SEARCH_017", "Emoji search input safety", "Critical", [
        ("Open Search, enter a bounded emoji query such as a globe and newspaper, and submit it.", "Search accepts or safely validates emoji input without crashing or corrupting the field."),
        ("Verify a controlled result or no-results state, wait, and capture it.", "Emoji handling is documented and the layout remains intact."),
    ]),
    ("SUB_SEARCH_018", "Sort Search results by Newest", "Critical", [
        ("Search for flood, open Sort, and select Newest.", "Newest becomes selected and the result list refreshes."),
        ("Verify visible result-date ordering where dates are available and capture the list.", "Results are presented newest-first without stale overlays."),
    ]),
    ("SUB_SEARCH_019", "Sort Search results by Oldest", "Critical", [
        ("Search for flood, open Sort, and select Oldest.", "Oldest becomes selected and the result list refreshes."),
        ("Verify visible result-date ordering where dates are available and capture the list.", "Results are presented oldest-first without a navigation failure."),
    ]),
    ("SUB_SEARCH_020", "Sort Search results by Relevance", "Critical", [
        ("Search for flood, change Sort to Oldest, then select Relevance.", "Relevance becomes selected and the list refreshes from the prior order."),
        ("Verify results remain related to flood and capture the final list.", "The relevance sort returns query-related results without stale ordering."),
    ]),
    ("SUB_SEARCH_021", "Multi-word news phrase search", "Critical", [
        ("Search for the phrase climate change and submit it.", "Results related to the complete phrase or its meaningful terms are displayed."),
        ("Open one visible result and verify the article page.", "A relevant result opens without losing the subscriber session."),
    ]),
    ("SUB_SEARCH_022", "Uppercase and lowercase query consistency", "Major", [
        ("Search for INDIA, capture the result state, then clear the query.", "Uppercase input returns a controlled result set."),
        ("Search for india and compare the result state.", "Search is case-insensitive or any intentional difference is consistently presented."),
    ]),
    ("SUB_SEARCH_023", "Partial-word search behavior", "Major", [
        ("Search for a partial news term such as elect and submit it.", "Matching results, suggestions, or a valid no-results state appears."),
        ("Replace it with election and submit again.", "The complete query refreshes the result set without retaining stale partial-query content."),
    ]),
    ("SUB_SEARCH_024", "Numeric and year query", "Major", [
        ("Search for the current year 2026 and submit it.", "The numeric query is accepted and a controlled result state is shown."),
        ("Open one result when available, then navigate Back.", "Numeric-search navigation works without a crash or malformed input."),
    ]),
    ("SUB_SEARCH_025", "Tamil Unicode query safety", "Critical", [
        ("Enter a Tamil news query such as சென்னை and submit it.", "Unicode text remains intact and Search returns results or a controlled no-results state."),
        ("Wait and capture the query and result state.", "Tamil characters render correctly without replacement symbols or layout corruption."),
    ]),
    ("SUB_SEARCH_026", "Leading and trailing whitespace normalization", "Major", [
        ("Enter spaces before and after Chennai and submit it.", "Search trims or safely handles surrounding whitespace."),
        ("Verify the returned state is not treated as an unrelated blank query.", "Whitespace does not create stale or malformed results."),
    ]),
    ("SUB_SEARCH_027", "Recent search history and reuse", "Major", [
        ("Search for Chennai, return to Search, and inspect the recent-search area when available.", "The recent query appears when search history is supported."),
        ("Tap the Chennai recent query when available.", "The saved query reruns and restores its result page; unsupported history is documented as a conditional skip."),
    ]),
    ("SUB_SEARCH_028", "Repeated identical query stability", "Major", [
        ("Search for flood, return to the input, and submit flood again.", "The repeated query refreshes normally without duplicating overlays or freezing."),
        ("Scroll the result page and capture it.", "The repeated-query results remain responsive."),
    ]),
    ("SUB_SEARCH_029", "Deep result scrolling and duplicate-card check", "Critical", [
        ("Search for war and scroll through multiple result-page viewports with waits.", "Additional results load or the end of results is reached cleanly."),
        ("Compare visible headlines across consecutive viewports.", "The UI does not repeat the same result card because of a pagination defect."),
    ]),
    ("SUB_SEARCH_030", "Sort persistence after article navigation", "Critical", [
        ("Search for flood, select Newest, and open a visible result article.", "The selected article opens from the Newest-sorted list."),
        ("Navigate Back and inspect the Search result state.", "The query and Newest selection remain active, or a documented default reset occurs consistently."),
    ]),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Subscriber Search"
    sheet.append(HEADERS)
    for case_id, name, priority, steps in CASES:
        for number, (step, expected) in enumerate(steps, 1):
            sheet.append([case_id, name, "Search", "SUBSCRIBER", PRE, number, step,
                          expected, priority, "Yes", MASTER, REF])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="273C75")
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [22, 52, 18, 16, 82, 13, 94, 94, 14, 14, 62, 100]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT)
    print(f"Created {OUTPUT} with {len(CASES)} cases and {sum(len(case[3]) for case in CASES)} steps")


if __name__ == "__main__":
    main()

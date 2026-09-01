from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Subscriber_Login_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "master_sheet_reference", "existing_yaml_reference",
]

PRECONDITION = (
    "The Hindu app is installed; network is available; the app starts signed out; "
    "the Subscriber profile contains valid active-subscriber TEST_EMAIL and "
    "TEST_PASSWORD credentials. Never place credential values in Excel or YAML."
)

CASES = [
    (
        "SUB_LOGIN_001",
        "Subscriber login from Let's Get Started",
        "Blocker",
        [
            (
                "Launch the app with cleared state and wait for the Let's Get Started page.",
                "The first-use page is visible with Create a free account and Have an account? Login.",
            ),
            (
                "Wait briefly for the page to settle, capture a screenshot, and tap Login in Have an account? Login.",
                "Login to your account opens.",
            ),
            (
                "Enter ${TEST_EMAIL} in the email field and tap Next.",
                "The password step opens for the supplied subscriber email without exposing its value in evidence.",
            ),
            (
                "Enter ${TEST_PASSWORD} in the password field and tap Login.",
                "Authentication succeeds and the subscriber enters the app.",
            ),
            (
                "After authentication reaches Account, press Back to return to Home, wait for Home to settle, capture a screenshot, and verify Subscribe is absent.",
                "Subscriber Home is visible and no Subscribe action or upsell is shown.",
            ),
        ],
    ),
    (
        "SUB_LOGIN_002",
        "Subscriber login from Account Settings after onboarding skip",
        "Blocker",
        [
            (
                "Launch the app with cleared state and wait for the Let's Get Started page.",
                "The first-use page is visible.",
            ),
            (
                "Tap Skip and wait until the anonymous Home page finishes loading.",
                "Home opens in the signed-out state.",
            ),
            (
                "Tap Account in the bottom navigation and wait for Account Settings.",
                "The anonymous Account page opens and shows Login.",
            ),
            (
                "Wait briefly, capture the anonymous Account page, and tap Login.",
                "Login to your account opens.",
            ),
            (
                "Enter ${TEST_EMAIL} in the email field and tap Next.",
                "The password step opens for the supplied subscriber email without exposing its value in evidence.",
            ),
            (
                "Enter ${TEST_PASSWORD} in the password field and tap Login.",
                "Authentication succeeds.",
            ),
            (
                "On the authenticated Account page, verify Account and scroll until Log Out is visible, capture the settled Account page, press Back, then verify Home and capture it.",
                "Authenticated Account and Log Out are visible; Back returns to Subscriber Home where Subscribe is absent.",
            ),
        ],
    ),
    (
        "SUB_LOGIN_003",
        "Subscriber login from article floating Subscribe",
        "Critical",
        [
            (
                "Launch the app with cleared state, tap Skip on the initial screen, and wait for anonymous Home.",
                "Anonymous Home opens with article cards and the signed-out Subscribe state.",
            ),
            (
                "Open the first available Home article and wait for the Article Page to settle, closing only a true interstitial when required.",
                "The selected Article Page opens.",
            ),
            (
                "Locate and tap the floating Subscribe action on the article.",
                "The subscription plan surface opens.",
            ),
            (
                "Verify the Yearly and Monthly plans, scroll when required, and capture the settled plan surface.",
                "Yearly, Monthly, and Already a subscriber? Login are available.",
            ),
            (
                "Tap Login in Already a subscriber? Login.",
                "Login to your account opens.",
            ),
            (
                "Enter ${TEST_EMAIL}, tap Next, enter ${TEST_PASSWORD}, and tap Login without re-entering either credential.",
                "Subscriber authentication succeeds through the article subscription entry point.",
            ),
            (
                "Wait for the originating article to return, capture it, and verify the anonymous Subscribe restriction is absent.",
                "The user returns authenticated to the Article Page and the anonymous subscription restriction is no longer shown.",
            ),
        ],
    ),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Subscriber Login"
    sheet.append(HEADERS)

    references = "THG App_Functionality_Matrix.xlsx; TH App Testing Scenarios_AutomationCopy.xlsx"
    yaml_references = "Common/SUBSCRIBER_LOGIN_ONCE.yaml; LOGIN_FROM_ARTICLE_SUBSCRIBE.yaml; ARTICLE_SUBSCRIBE_LOGIN_RETURN.yaml; SC_32_AccountSettingPage_subscriberAccount.yaml"
    for case_id, name, priority, steps in CASES:
        for number, (step, expected) in enumerate(steps, 1):
            sheet.append([
                case_id, name, "Login", "SUBSCRIBER", PRECONDITION,
                number, step, expected, priority, "Yes", references, yaml_references,
            ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    widths = [18, 48, 18, 16, 72, 14, 78, 78, 12, 14, 58, 58]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()

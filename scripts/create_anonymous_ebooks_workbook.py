from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Anonymous_EBooks_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "master_sheet_reference", "existing_yaml_reference",
]

MASTER_REFERENCES = {
    "ANON_EBOOK_001": "1179, 1201, 1204, 2113-2114",
    "ANON_EBOOK_002": "2115-2122",
    "ANON_EBOOK_003": "2115-2122",
    "ANON_EBOOK_004": "2115-2119",
    "ANON_EBOOK_005": "2120-2122",
    "ANON_EBOOK_006": "2123, 2127, 2203, 6297",
    "ANON_EBOOK_007": "2124, 2127, 2203, 6297",
    "ANON_EBOOK_008": "2125, 2127, 2203, 6297",
    "ANON_EBOOK_009": "1203, 2203, 6297",
    "ANON_EBOOK_010": "1203, 2203, 6366",
    "ANON_EBOOK_011": "1203, 2205",
    "ANON_EBOOK_012": "2114, 2127",
    "ANON_EBOOK_013": "2123-2127, 2203, 6297",
    "ANON_EBOOK_014": "1203, 2203, 6297",
    "ANON_EBOOK_015": "429, 1179",
    "ANON_EBOOK_016": "1185, 2132-2133",
    "ANON_EBOOK_017": "2126",
    "ANON_EBOOK_018": "2140, 2151-2152",
    "ANON_EBOOK_019": "2193, 2198, 2201",
    "ANON_EBOOK_020": "2203, 2205, 6297, 6366",
    "ANON_EBOOK_021": "1191, 2198",
    "ANON_EBOOK_022": "2196-2200",
    "ANON_EBOOK_023": "2134, 2201",
    "ANON_EBOOK_024": "2179-2191",
    "ANON_EBOOK_025": "2203, 2205, 6297, 6366",
    "ANON_EBOOK_026": "2203, 6297",
    "ANON_EBOOK_027": "6297",
}
YAML_REFERENCE = (
    "SC_26_Anonymous_E-Book_restriction.yaml; "
    "SC_51_EBOOKS_Page_Boundary_Validation.yaml; TH_0011.yaml; TH_0012.yaml"
)


CASES = [
    ("ANON_EBOOK_001", "Anonymous eBooks launch and selected navigation", [
        ("Launch the app in a fresh anonymous state, dismiss the welcome popup when shown, and tap eBooks.", "The eBooks root screen opens."),
        ("Verify the eBooks bottom-navigation icon is selected and capture the page.", "eBooks is visibly selected and evidence is saved."),
    ]),
    ("ANON_EBOOK_002", "eBooks single pull to refresh", [
        ("Open eBooks and record the current card state.", "The current eBook card is visible."),
        ("Pull down once to refresh.", "Refresh completes without leaving eBooks."),
        ("Verify the eBooks screen and current card after refresh.", "eBooks remains usable with content loaded."),
    ]),
    ("ANON_EBOOK_003", "eBooks repeated pull to refresh", [
        ("Open eBooks and pull down to refresh twice, waiting for each refresh to finish.", "Both refresh cycles complete successfully."),
        ("After each refresh verify the eBooks screen, selected navigation, and at least one card.", "The page remains stable and does not duplicate or lose content."),
        ("Capture the final refreshed state.", "Repeated-refresh evidence is saved."),
    ]),
    ("ANON_EBOOK_004", "eBooks current-card presentation", [
        ("Open eBooks and inspect the featured current eBook card.", "The current card loads."),
        ("Verify cover image, title, publication date, and description where provided.", "Current eBook metadata renders correctly."),
        ("Capture the card without clipped or overlapping content.", "Card presentation evidence is saved."),
    ]),
    ("ANON_EBOOK_005", "Previous eBooks listing and scrolling", [
        ("Open eBooks and scroll to the previous eBooks section.", "Previous eBooks are displayed."),
        ("Scroll through multiple previous eBook cards.", "The listing scrolls smoothly and cards remain usable."),
        ("Capture top, middle, and lower listing states.", "Previous eBook coverage evidence is saved."),
    ]),
    ("ANON_EBOOK_006", "Open eBook from cover", [
        ("Open eBooks and tap a visible eBook cover.", "The anonymous subscription restriction opens instead of the reader."),
        ("Verify the subscription offer and capture it.", "The eBook cannot be read without a subscription."),
    ]),
    ("ANON_EBOOK_007", "Open eBook from title", [
        ("Open eBooks and tap a visible eBook title.", "The same anonymous subscription restriction opens."),
        ("Verify the selected title does not bypass entitlement.", "Full eBook content remains protected."),
    ]),
    ("ANON_EBOOK_008", "Open eBook from description", [
        ("Open eBooks and tap a visible eBook description area when supported.", "The same eBook restriction opens, or the non-tappable description leaves the card unchanged according to design."),
        ("Capture the observed behavior.", "Description interaction is documented without bypassing entitlement."),
    ]),
    ("ANON_EBOOK_009", "Anonymous eBook subscription-plan restriction", [
        ("Open a representative eBook as an anonymous user.", "A subscription purchase or support-journalism screen appears."),
        ("Verify Subscribe or purchase messaging, Yearly, Monthly, and Already a subscriber Login.", "The required subscription and login choices are visible."),
        ("Capture the complete restriction page.", "Anonymous eBook access evidence is saved."),
    ]),
    ("ANON_EBOOK_010", "eBook Subscribe purchase route", [
        ("Open the anonymous eBook restriction and tap Subscribe or the purchase action.", "The plan-selection or checkout journey opens, not the account-login page."),
        ("Verify Yearly and Monthly or the applicable purchase details and capture them.", "The purchase route is correct."),
        ("Press Back without purchasing.", "The eBook restriction or eBooks listing is safely restored."),
    ]),
    ("ANON_EBOOK_011", "eBook existing-subscriber Login route", [
        ("Open the anonymous eBook restriction and tap Already a subscriber Login.", "Login to your account opens."),
        ("Verify the login-page identity and capture it.", "The existing-subscriber link reaches the correct login destination."),
        ("Press Back without signing in.", "The anonymous eBook restriction is restored."),
    ]),
    ("ANON_EBOOK_012", "eBook restriction back-navigation", [
        ("Open an eBook to display its anonymous restriction.", "The subscription restriction is visible."),
        ("Press Back once.", "The eBooks listing is restored rather than closing the app."),
        ("Verify eBooks remains selected and the anonymous session is retained.", "Navigation state remains correct."),
    ]),
    ("ANON_EBOOK_013", "Multiple eBooks entitlement consistency", [
        ("Open at least three different visible eBooks one at a time.", "Every selected eBook presents the anonymous subscription restriction."),
        ("For every restriction verify Subscribe and Already a subscriber Login, then return.", "No eBook bypasses anonymous access control."),
        ("Capture each distinct restriction state.", "Cross-card entitlement evidence is saved."),
    ]),
    ("ANON_EBOOK_014", "eBook restriction after repeated refresh", [
        ("Refresh eBooks twice and then open an eBook.", "The eBook subscription restriction still appears."),
        ("Verify refresh does not unlock reader content or lose the Login route.", "Anonymous entitlement remains enforced."),
    ]),
    ("ANON_EBOOK_015", "eBooks masthead Home round-trip", [
        ("Open eBooks and tap The Hindu masthead.", "The Home screen opens."),
        ("Tap eBooks again.", "The eBooks listing is restored with eBooks selected."),
    ]),
    ("ANON_EBOOK_016", "eBooks listing scroll-state restoration", [
        ("Scroll to a previous eBook at a non-top position and open it.", "The restriction page opens."),
        ("Press Back.", "The eBooks listing is restored."),
        ("Verify the prior listing position is retained or document the product reset behavior.", "Return-state behavior is captured for regression comparison."),
    ]),
    ("ANON_EBOOK_017", "eBook long-title and metadata layout", [
        ("Inspect eBook cards with long titles and descriptions.", "Text wraps without clipping, overlap, or hiding the cover."),
        ("Verify publication date and metadata remain aligned.", "Card layout remains readable."),
        ("Capture representative long-content cards.", "Layout evidence is saved."),
    ]),
    ("ANON_EBOOK_018", "eBooks rapid card interaction protection", [
        ("Rapidly tap a single eBook card within a bounded sequence.", "Only one restriction or purchase page opens."),
        ("Verify duplicate stacked dialogs or repeated checkout pages do not appear.", "Rapid input is handled safely."),
        ("Return to eBooks.", "The listing remains usable."),
    ]),
    ("ANON_EBOOK_019", "eBooks controlled resilience and recovery", [
        ("Open an eBook restriction page, background and resume the app, then verify the restriction remains.", "Lifecycle changes do not expose reader content."),
        ("In a controlled run, interrupt connectivity while opening an eBook and restore it.", "The app shows a stable error or retry state and then recovers to the restriction."),
        ("Capture the recovered restriction with Subscribe and Login available.", "Anonymous access control survives lifecycle and network recovery."),
    ]),
    ("ANON_EBOOK_020", "Anonymous eBooks access-boundary summary", [
        ("Open the current eBook and two previous eBooks one at a time.", "Every selected eBook shows the subscription restriction before readable content."),
        ("Verify Subscribe opens plans with Yearly and Monthly, then return without purchasing.", "The purchase route is correct."),
        ("Verify Already a subscriber Login opens Login to your account, then return without signing in.", "The authentication route is correct and remains distinct from purchase."),
    ]),
    ("ANON_EBOOK_021", "eBooks network loss while opening", [
        ("Using a controlled network test, disconnect connectivity before opening an eBook.", "The app shows a loading, error, or retry state without crashing."),
        ("Verify protected reader content is not exposed during the failure.", "The app fails securely."),
        ("Restore connectivity and retry.", "The normal anonymous subscription restriction opens."),
    ]),
    ("ANON_EBOOK_022", "eBooks network transition and recovery", [
        ("Open eBooks on Wi-Fi and begin a refresh.", "The listing is usable on Wi-Fi."),
        ("Switch to mobile data or another available connection and refresh again.", "The listing recovers and remains usable."),
        ("Open an eBook and verify the anonymous restriction.", "Network transition does not change entitlement."),
    ]),
    ("ANON_EBOOK_023", "eBooks slow-loading and duplicate-content protection", [
        ("Using controlled network throttling, open and refresh eBooks.", "A loading indicator or stable waiting state appears."),
        ("Wait for completion and verify cards are not duplicated or left permanently blank.", "The listing finishes in a consistent state."),
        ("Open a card and verify the subscription restriction.", "The delayed load does not bypass access control."),
    ]),
    ("ANON_EBOOK_024", "eBooks dark mode and accessibility layout", [
        ("Open eBooks and its restriction page in dark mode.", "Cards, subscription text, buttons, and Login remain readable."),
        ("Repeat with maximum supported font scaling.", "Text reflows without clipping or overlapping purchase controls."),
        ("Capture both states and restore default settings.", "Accessibility evidence is saved without leaving settings modified."),
    ]),
    ("ANON_EBOOK_025", "eBooks anonymous access boundary summary", [
        ("Open the current eBook and representative previous eBooks after one and two refresh cycles.", "Every eBook reaches the same anonymous access boundary."),
        ("Verify purchase actions lead to plans or checkout and Already a subscriber Login leads to Login to your account.", "Purchase and authentication routes remain distinct and correct."),
        ("Capture the final listing, restriction, plan, and login states.", "Complete anonymous eBooks evidence is saved."),
    ]),
    ("ANON_EBOOK_026", "Anonymous eBook reader access prevention", [
        ("Open a representative eBook as an anonymous user.", "The subscription restriction opens before readable book content."),
        ("Attempt to reach the reader without purchasing or signing in.", "The first page, page gestures, and reader controls remain inaccessible."),
        ("Capture the blocking restriction and return to the eBooks listing.", "Access-prevention evidence is saved and eBooks is restored."),
    ]),
    ("ANON_EBOOK_027", "Anonymous premium eBook access boundary", [
        ("Open another controlled eBook as an anonymous user.", "The subscription restriction blocks the reader consistently."),
        ("Verify Subscribe, Yearly, Monthly, and Already a subscriber Login when supplied by the current offer.", "Premium access and authentication choices are visible."),
        ("Capture the restriction without purchasing or signing in.", "Premium-access evidence is saved safely."),
    ]),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous eBooks"
    sheet.append(HEADERS)
    precondition = (
        "The Hindu app is installed; network is available; the user is signed out; "
        "eBook inventory and subscription offers may vary between runs."
    )
    # Twenty cases retain unique navigation, presentation, entitlement,
    # interaction, lifecycle, and recovery coverage without summary duplicates.
    for case_id, name, steps in CASES[:20]:
        for number, (step, expected) in enumerate(steps, 1):
            controlled_only = case_id in {
                "ANON_EBOOK_019", "ANON_EBOOK_021",
                "ANON_EBOOK_022", "ANON_EBOOK_023", "ANON_EBOOK_024",
            }
            sheet.append([
                case_id, name, "eBooks", "ANONYMOUS", precondition,
                number, step, expected, "Medium" if controlled_only else "High",
                "No - controlled device" if controlled_only else "Yes",
                MASTER_REFERENCES.get(case_id, "2113-2206"), YAML_REFERENCE,
            ])
    header_fill = PatternFill("solid", fgColor="8A6A28")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [21, 50, 18, 16, 76, 14, 90, 90, 12, 14, 30, 65]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()

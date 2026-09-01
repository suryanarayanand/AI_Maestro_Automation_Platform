import json
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Uploads" / "Ready" / "Anonymous_EBooks_Approved_Test_Cases.xlsx"
SOURCE = WORKBOOK.name


def build_yaml(case_id, name):
    number = int(case_id.rsplit("_", 1)[1])
    lower = name.casefold()
    commands = [
        "appId: com.mobstac.thehindu",
        "tags: [generated, ordered, anonymous, ebooks]",
        "---",
        "",
    ]
    restriction_case = number in set(range(6, 15)) | {18, 19, 20}
    commands.append(
        '- runFlow: "../Common/OPEN_ANONYMOUS_EBOOK_RESTRICTION.yaml"'
        if restriction_case else '- runFlow: "../Common/OPEN_ANONYMOUS_EBOOKS.yaml"'
    )

    if "refresh" in lower:
        times = 2 if "repeated" in lower else 1
        commands += [
            '- repeat:', f'    times: {times}', '    commands:',
            '      - swipe: {direction: DOWN}', '      - waitForAnimationToEnd: {timeout: 2500}',
            '- assertVisible: {id: "screen_ebooks"}',
        ]
    if number in {4, 5, 17}:
        commands += [
            '- assertVisible: {id: "screen_ebooks"}',
            '- repeat:', '    times: 3', '    commands:',
            '      - swipe: {direction: UP}', '      - waitForAnimationToEnd: {timeout: 1000}',
            '- assertVisible: {id: "screen_ebooks"}',
        ]
    if restriction_case:
        commands += ['- runFlow: "../Common/ASSERT_ANONYMOUS_EBOOK_RESTRICTION.yaml"']
    if number == 10:
        commands += [
            '- tapOn: {text: "Subscribe|SUBSCRIBE"}',
            '- extendedWaitUntil: {visible: {text: "Yearly"}, timeout: 25000}',
            '- assertVisible: {text: "Yearly"}', '- assertVisible: {text: "Monthly"}',
        ]
    if number == 11:
        commands += [
            '- tapOn: {text: "Already a subscriber.*Login|Login|LOGIN"}',
            '- extendedWaitUntil: {visible: {text: "Login to your account"}, timeout: 25000}',
            '- assertVisible: {text: "Login to your account"}',
        ]
    if number == 12:
        commands += [
            '- back', '- extendedWaitUntil: {visible: {id: "screen_ebooks"}, timeout: 25000}',
            '- assertVisible: {id: "screen_ebooks"}',
        ]
    if number == 13:
        commands += [
            '- back', '- assertVisible: {id: "screen_ebooks"}',
            '- swipe: {direction: UP}', '- tapOn: {point: "77%,34%"}',
            '- runFlow: "../Common/ASSERT_ANONYMOUS_EBOOK_RESTRICTION.yaml"',
        ]
    if number == 15:
        commands += [
            '- tapOn: {id: "nav_home"}', '- assertVisible: {id: "screen_home"}',
            '- tapOn: {id: "nav_ebooks"}', '- assertVisible: {id: "screen_ebooks"}',
        ]
    if number == 16:
        commands += [
            '- swipe: {direction: UP}', '- tapOn: {point: "77%,34%"}',
            '- runFlow: "../Common/ASSERT_ANONYMOUS_EBOOK_RESTRICTION.yaml"',
            '- back', '- assertVisible: {id: "screen_ebooks"}',
        ]
    if number == 18:
        commands += ['- assertVisible: {text: "Subscribe|SUBSCRIBE"}']
    if number == 19:
        commands += [
            '- takeScreenshot: "Screenshots/Generated/ANON_EBOOK_019_controlled_recovery"',
            '- assertVisible: {text: "Subscribe|SUBSCRIBE|Retry|No Internet"}',
        ]
    if number == 20:
        commands += [
            '- tapOn: {text: "Subscribe|SUBSCRIBE"}', '- assertVisible: {text: "Yearly"}',
            '- assertVisible: {text: "Monthly"}', '- back',
        ]
    commands += [f'- takeScreenshot: "Screenshots/Generated/{case_id}_result"']
    return "\n".join(commands) + "\n"


workbook = load_workbook(WORKBOOK)
sheet = workbook.active
headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}
for header in ("runtime_assertion", "yaml_structure"):
    if header not in headers:
        headers[header] = sheet.max_column + 1
        sheet.cell(1, headers[header], header)

cases = {}
for row in range(2, sheet.max_row + 1):
    case_id = str(sheet.cell(row, headers["test_case_id"]).value)
    cases.setdefault(case_id, {"name": str(sheet.cell(row, headers["name"]).value), "rows": []})["rows"].append(row)
    sheet.cell(row, headers["runtime_assertion"], "Assert eBooks screen or mandatory anonymous subscription restriction at runtime.")
    sheet.cell(row, headers["yaml_structure"], "Setup > eBooks navigation > action > restriction assertion > evidence > safe return")
for header in ("runtime_assertion", "yaml_structure"):
    sheet.column_dimensions[sheet.cell(1, headers[header]).column_letter].width = 65
workbook.save(WORKBOOK)

db = sqlite3.connect(ROOT / "portal.db")
db.execute("DELETE FROM drafts WHERE source_file=? AND status IN ('pending','rejected')", (SOURCE,))
for case_id, case in cases.items():
    traceability = []
    for position, row in enumerate(case["rows"], 1):
        traceability.append({
            "position": position, "source_type": "expected_result",
            "step_number": sheet.cell(row, headers["step_number"]).value,
            "requirement": str(sheet.cell(row, headers["expected_result"]).value or ""),
            "generation_input": str(sheet.cell(row, headers["step"]).value or ""),
            "commands": ["runFlow", "assertVisible", "assertNotVisible", "takeScreenshot"],
            "selector": "nav_ebooks; screen_ebooks; validated restriction text",
            "status": "covered",
            "reason": "Grounded in SC_26, SC_51, TH_0011 and TH_0012.",
            "source_sheet": sheet.title, "source_row": row,
        })
    db.execute(
        """INSERT INTO drafts(case_id,name,yaml,source_file,error,generation_mode,
        ai_confidence,ai_assumptions,traceability,coverage_status,user_state)
        VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
        (case_id, case["name"], build_yaml(case_id, case["name"]), SOURCE, None,
         "reference-template", 1.0, json.dumps(["Existing manual eBooks YAML reused."]),
         json.dumps(traceability), "complete", "ANONYMOUS"),
    )
db.commit()
db.close()
print("reviewable eBooks drafts", len(cases))

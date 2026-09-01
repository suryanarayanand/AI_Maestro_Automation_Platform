"""Create the Subscriber eBooks module import workbook."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT=Path("Uploads/Ready/Subscriber_EBooks_Approved_Test_Cases.xlsx")
PRE=("The Hindu app is installed; network is available; configured masked credentials "
     "belong to an active subscriber with the applicable eBooks entitlement.")

CASES=[
 ("SUB_EBOOK_001","Subscriber eBooks launch and selected navigation","Critical",[
  ("Authenticate with configured subscriber credentials and tap eBooks.","The eBooks page opens and its navigation item is selected.","OPEN_SUBSCRIBER_HOME; tap nav_ebooks; assert eBooks screen/navigation"),
  ("Wait for the current eBook content to load and capture the landing page.","Readable eBooks landing evidence is saved.","waitForAnimationToEnd immediately before screenshot"),
  ("Verify Subscribe, Login, purchase restriction, and advertisements are absent.","The entitled eBooks landing page has no monetisation gate.","ASSERT_SUBSCRIBER_NO_MONETIZATION; assertNotVisible Login/Choose plan")]),
 ("SUB_EBOOK_002","Subscriber eBooks single refresh","Critical",[
  ("Open eBooks and perform one pull-to-refresh gesture.","Refresh completes without leaving eBooks.","swipe DOWN; extendedWaitUntil eBooks screen"),
  ("Verify current cover, title, and metadata reload correctly.","The current eBook card remains complete and usable.","assert cover/title/metadata"),
  ("Wait, capture the refreshed page, and verify entitlement remains active.","Refresh evidence is saved with no restriction or advertisement.","assert no monetisation/restriction; wait; screenshot")]),
 ("SUB_EBOOK_003","Subscriber eBooks repeated-refresh stability","High",[
  ("Refresh the eBooks page twice with a loading wait between gestures.","Both refreshes complete without duplicate or broken content.","repeat 2 swipe DOWN; wait"),
  ("Verify the eBooks page and current card remain visible.","The page remains stable after repeated refresh.","assert eBooks screen/current card"),
  ("Verify the subscriber session and entitlement are retained.","No Login, Subscribe, plan, or advertisement appears.","assert no monetisation/login/plan")]),
 ("SUB_EBOOK_004","Current eBook card presentation","Critical",[
  ("Inspect the current eBook card after loading settles.","Cover, title, description, and available issue metadata are readable.","assert card elements"),
  ("Check the card for clipping, overlap, or missing primary content.","The current card layout is visually usable.","visual assertions; wait; screenshot"),
  ("Verify no advertisement overlays the card.","The subscriber card remains ad-free.","ASSERT_SUBSCRIBER_NO_MONETIZATION")]),
 ("SUB_EBOOK_005","Previous eBooks listing and scrolling","High",[
  ("Scroll from the current eBook into the previous-books listing.","Previous eBook cards load progressively.","repeat bounded swipe UP"),
  ("Verify visible covers, titles, and issue metadata remain aligned.","Previous entries are readable without overlap.","assert visible card content"),
  ("Capture top, middle, and lower listing states after waits and verify no ads.","The listing is fully documented and ad-free.","wait before screenshots; ASSERT_SUBSCRIBER_NO_MONETIZATION")]),
 ("SUB_EBOOK_006","Open eBook from cover","Blocker",[
  ("Tap the current eBook cover.","The selected eBook reader or entitled detail opens.","tap cover; assert reader/detail"),
  ("Verify no plan selector, Subscribe action, or Login page blocks access.","Cover entry respects the active entitlement.","assertNotVisible plan/Subscribe/Login"),
  ("Wait and capture the opened eBook state.","Entitled cover-entry evidence is saved.","wait; screenshot")]),
 ("SUB_EBOOK_007","Open eBook from title","Blocker",[
  ("Return to eBooks and tap the current title.","The same corresponding eBook reader or detail opens.","tap title; assert reader/detail"),
  ("Verify the destination matches the selected title and is not a purchase gate.","Title entry resolves to entitled content.","assert title/context; assertNotVisible plan/Subscribe/Login"),
  ("Capture the settled destination and return safely.","Evidence is saved and Back restores eBooks.","wait; screenshot; back; assert eBooks")]),
 ("SUB_EBOOK_008","Open eBook from previous listing","Blocker",[
  ("Scroll to a previous eBook and tap its card.","The selected previous eBook opens.","scroll; tap previous card; assert reader/detail"),
  ("Verify entitled access without a subscription or authentication prompt.","The previous eBook is accessible under subscriber entitlement.","assertNotVisible plan/Subscribe/Login"),
  ("Capture the opened state after waiting and return.","Previous-eBook access is documented.","wait; screenshot; back")]),
 ("SUB_EBOOK_009","Open eBook and navigate through all pages","Blocker",[
  ("Open an entitled eBook and verify its first readable page is displayed.","The book reader opens at a valid page without a purchase or login restriction.","tap book; assert reader/page; assertNotVisible plan/Subscribe/Login; wait; screenshot"),
  ("Swipe left one page at a time, waiting after every swipe, until the reader displays its final-page or end-of-book indicator; use a safe maximum-page bound to prevent an infinite run.","Every forward swipe advances the page and the final page is reached within the safety bound.","repeatUntil final-page/end indicator with max bound; swipe LEFT; wait after each swipe"),
  ("Wait and capture the final page, then swipe right through multiple pages and verify the page number decreases or earlier content returns.","The final page is documented and reverse page navigation works without losing subscriber entitlement.","assert final-page/end indicator; wait; screenshot; repeat swipe RIGHT; wait; assert earlier page; assert no restriction/ads")]),
 ("SUB_EBOOK_010","Multiple eBooks entitlement consistency","Blocker",[
  ("Open three different current or previous eBooks using bounded selection.","Each selected eBook reaches its reader or entitled detail.","bounded 3-card loop"),
  ("For every selection, verify no purchase plan, Subscribe, Login, or advertisement appears.","Entitlement is consistent across multiple books.","assert no restriction/monetisation per book"),
  ("Capture one settled state per selected eBook and return to the listing.","Multiple-book evidence and safe Back navigation are available.","wait; screenshot; back; assert eBooks")]),
 ("SUB_EBOOK_011","eBooks masthead Home round trip","High",[
  ("From eBooks, tap The Hindu masthead.","Subscriber Home opens.","tap masthead; assert screen_home"),
  ("Return to eBooks using its navigation control.","The eBooks page reopens without reauthentication.","tap nav_ebooks; assert eBooks"),
  ("Verify entitlement and no-ad state persist; capture after waiting.","The session remains active and ad-free.","assert no login/restriction/ads; wait; screenshot")]),
 ("SUB_EBOOK_012","eBooks scroll-state restoration","High",[
  ("Scroll to a lower previous-eBooks position and open a card.","The selected eBook opens from the lower listing.","repeat swipe UP; tap card; assert reader/detail"),
  ("Return using Back.","The eBooks listing is restored.","back; assert eBooks screen"),
  ("Verify the previous-books context is retained and capture it after waiting.","Return navigation preserves useful listing context.","assert previous listing; wait; screenshot")]),
 ("SUB_EBOOK_013","eBooks long-title and metadata layout","High",[
  ("Locate an eBook card with a long title or extended metadata.","The applicable card is visible within bounded scrolling.","bounded scroll/card selection"),
  ("Verify title, description, cover, and metadata do not overlap or truncate primary actions.","The complex card remains readable and actionable.","visual assertions"),
  ("Wait and capture the card with no advertisement overlay.","Layout evidence is saved in an ad-free state.","ASSERT_SUBSCRIBER_NO_MONETIZATION; wait; screenshot")]),
 ("SUB_EBOOK_014","Subscriber eBooks session and access persistence","Blocker",[
  ("Navigate Home to eBooks, open a book, return Home, and reopen eBooks.","The complete round trip succeeds without losing authentication.","Home/eBooks/book/Home/eBooks round trip"),
  ("Open another eBook and verify access remains entitled.","The second book opens without plan, Subscribe, or Login prompts.","assert reader/detail; assertNotVisible restrictions"),
  ("Verify ads remain absent and capture the final settled eBooks state.","Subscriber access and no-ad behavior persist.","ASSERT_SUBSCRIBER_NO_MONETIZATION; wait; screenshot")]),
]

def main():
 wb=Workbook();ws=wb.active;ws.title="Subscriber eBooks"
 headers=["test_case_id","name","module","user_state","precondition","step_number","step","expected_result","priority","automatable","runtime_assertion"]
 ws.append(headers)
 for cid,name,priority,steps in CASES:
  for number,(step,expected,assertion) in enumerate(steps,1):ws.append([cid,name,"eBooks","SUBSCRIBER",PRE,number,step,expected,priority,"Yes",assertion])
 for c in ws[1]:c.fill=PatternFill("solid",fgColor="17365D");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(horizontal="center")
 for i,w in enumerate([20,44,16,16,58,13,72,68,12,13,72],1):ws.column_dimensions[get_column_letter(i)].width=w
 for row in ws.iter_rows(min_row=2):
  for c in row:c.alignment=Alignment(wrap_text=True,vertical="top")
 ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions;OUTPUT.parent.mkdir(parents=True,exist_ok=True);wb.save(OUTPUT)
 print(f"Created {OUTPUT} with {len(CASES)} cases and {ws.max_row-1} steps")

if __name__=="__main__":main()

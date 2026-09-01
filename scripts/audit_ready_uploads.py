import json
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
READY = ROOT / "Uploads" / "Ready"


def workbook_case_ids(path):
    result = set()
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip().casefold() for value in next(rows)]
        except StopIteration:
            continue
        column = next((headers.index(name) for name in ("test_case_id", "test case id", "case_id", "id") if name in headers), None)
        if column is None:
            continue
        for row in rows:
            if column < len(row) and row[column]:
                result.add(str(row[column]).strip())
    workbook.close()
    return result


def main():
    db = sqlite3.connect(ROOT / "portal.db")
    db.row_factory = sqlite3.Row
    draft_rows = db.execute("SELECT source_file,case_id FROM drafts").fetchall()
    drafts_by_source = {}
    all_draft_ids = set()
    for row in draft_rows:
        drafts_by_source.setdefault(row["source_file"], set()).add(row["case_id"])
        all_draft_ids.add(row["case_id"])
    results = []
    for path in sorted(READY.glob("*.xlsx")):
        ids = workbook_case_ids(path)
        direct = drafts_by_source.get(path.name, set())
        covered = ids & all_draft_ids
        if ids and ids <= direct:
            status = "uploaded_directly"
        elif ids and ids <= all_draft_ids:
            status = "covered_by_another_uploaded_workbook"
        elif covered:
            status = "partially_uploaded"
        else:
            status = "not_uploaded"
        results.append({
            "file": path.name, "cases": len(ids), "status": status,
            "direct_cases": len(ids & direct), "portal_cases": len(covered),
            "missing_ids": sorted(ids - all_draft_ids),
        })
    catalog = [dict(row) for row in db.execute(
        "SELECT source_file,COUNT(DISTINCT case_id) cases FROM atomic_flow_steps GROUP BY source_file"
    )]
    behavior_rules = db.execute("SELECT COUNT(*) FROM app_behavior_rules WHERE status='approved'").fetchone()[0]
    db.close()
    print(json.dumps({"ready_workbooks": results, "catalog_sources": catalog, "approved_behavior_rules": behavior_rules}, indent=2))


if __name__ == "__main__":
    main()

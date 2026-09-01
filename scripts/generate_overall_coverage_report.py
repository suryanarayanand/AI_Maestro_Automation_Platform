import argparse, html, json, sqlite3, subprocess
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Reports" / "Overall_Coverage"
EDGE = Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe")
SUITES = [
    "user_anonymous.json", "user_subscriber.json", "user_registered.json", "user_expired.json",
    "user_anonymous__module__premium.json", "user_anonymous__module__ebooks.json",
]
READY = [
    "Anonymous_Games_Hamburger_Account_Settings_Approved_Test_Cases.xlsx",
]

def esc(value): return html.escape(str(value or ""))
def state(value):
    value = str(value or "UNKNOWN").upper().replace(" ", "_").replace("-", "_")
    if "ANON" in value: return "ANONYMOUS"
    if "EXPIRED" in value: return "EXPIRED_USER"
    if "REGISTER" in value: return "REGISTERED_USER"
    if "SUBSCR" in value: return "SUBSCRIBER"
    return value

def collect_cases():
    cases = {}
    for filename in SUITES:
        path = ROOT / "Suites" / filename
        if not path.is_file(): continue
        data = json.loads(path.read_text(encoding="utf-8"))
        for test in data.get("tests", []):
            cid = str(test.get("id", "")).strip()
            if not cid: continue
            cases[cid] = {"id":cid,"name":test.get("name",cid),"module":test.get("display_module") or test.get("module") or "Unassigned","user_state":state(test.get("user_state") or filename),"yaml":test.get("yaml", ""),"source":filename,"readiness":"Executable YAML"}
    for filename in READY:
        path = ROOT / "Uploads" / "Ready" / filename
        if not path.is_file(): continue
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True); headers = [str(v or "").strip() for v in next(rows)]; ix = {v:i for i,v in enumerate(headers)}
            for row in rows:
                cid = str(row[ix["test_case_id"]] or "").strip()
                if not cid or cid in cases: continue
                cases[cid] = {"id":cid,"name":row[ix["name"]],"module":row[ix["module"]],"user_state":state(row[ix["user_state"]]),"yaml":"Pending generation/review","source":filename,"readiness":"Designed; pending YAML approval"}
        wb.close()
    return cases

def collect_results(cases):
    db = sqlite3.connect(ROOT / "portal.db"); db.row_factory = sqlite3.Row
    jobs = [dict(r) for r in db.execute("SELECT id,suite,status,completed,total,created_at,finished_at FROM jobs ORDER BY id")]
    history = Counter(str(r[0]).upper() for r in db.execute("SELECT status FROM job_results"))
    latest = {}
    for row in db.execute("""SELECT r.* FROM job_results r JOIN (SELECT case_id,MAX(id) id FROM job_results GROUP BY case_id) x ON x.id=r.id"""):
        latest[row["case_id"]] = dict(row)
    db.close()
    for cid,item in cases.items(): item["status"] = str(latest.get(cid, {}).get("status") or "PENDING").upper()
    return jobs, history

def render(cases, jobs, history):
    grouped=defaultdict(list)
    for item in cases.values(): grouped[(item["user_state"],item["module"])].append(item)
    latest_counts=Counter(item["status"] for item in cases.values()); job_counts=Counter(j["status"].upper() for j in jobs)
    status_order = ["PASS", "FAIL", "NEEDS_REVIEW", "CANCELLED", "PENDING"]
    cards="".join(f'<div class="metric"><b>{esc(k)}</b><strong>{latest_counts.get(k, 0)}</strong></div>' for k in status_order)
    planned_states = ["ANONYMOUS", "SUBSCRIBER", "REGISTERED_USER", "EXPIRED_USER"]
    state_rows = "".join(
        f'<tr><td>{esc(user)}</td><td>{sum(1 for x in cases.values() if x["user_state"] == user)}</td><td>{"Covered" if any(x["user_state"] == user for x in cases.values()) else "Not yet covered"}</td></tr>'
        for user in planned_states
    )
    module_counts = Counter(item["module"] for item in cases.values())
    module_rows = "".join(f'<tr><td>{esc(module)}</td><td>{count}</td></tr>' for module,count in sorted(module_counts.items()))
    history_rows = "".join(f'<tr><td>{label}</td><td>{history.get(label, 0)}</td></tr>' for label in status_order)
    job_rows = "".join(f'<tr><td>{label}</td><td>{job_counts.get(label, 0)}</td></tr>' for label in ("PASSED","FAILED","NEEDS_REVIEW","CANCELLED","RUNNING","QUEUED"))
    sections=[]
    for (user,module),items in sorted(grouped.items()):
        counts=Counter(x["status"] for x in items)
        rows="".join(f'<tr><td>{esc(x["id"])}</td><td>{esc(x["name"])}</td><td>{esc(x["yaml"])}</td><td>{esc(x["readiness"])}</td><td><span class="s {x["status"].lower()}">{esc(x["status"])}</span></td></tr>' for x in sorted(items,key=lambda y:y["id"]))
        summary=" · ".join(f"{k}: {v}" for k,v in sorted(counts.items()))
        sections.append(f'<section><h2>{esc(user)} / {esc(module)}</h2><p class="summary">{len(items)} flows · {esc(summary)}</p><table><thead><tr><th>Case</th><th>Flow</th><th>YAML</th><th>Readiness</th><th>Latest</th></tr></thead><tbody>{rows}</tbody></table></section>')
    generated=datetime.now().strftime("%d %B %Y, %I:%M %p IST")
    return f'''<!doctype html><html><head><meta charset="utf-8"><style>@page{{size:A4 landscape;margin:10mm}}*{{box-sizing:border-box}}body{{font:10px Arial;color:#172033;margin:0}}header{{padding:18px 22px;background:#172640;color:white;border-radius:10px}}h1{{margin:0;font-size:25px}}header p{{margin:6px 0 0;color:#b9c9e5}}.metrics{{display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin:12px 0}}.metric{{padding:10px;border:1px solid #dbe3ef;border-radius:8px;background:#f7f9fc}}.metric b{{display:block;font-size:8px;color:#68758a}}.metric strong{{display:block;font-size:19px;margin-top:4px}}.overview{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}section{{break-inside:avoid;margin:14px 0}}h2{{margin:0;padding-bottom:5px;border-bottom:2px solid #316fea;font-size:15px}}.summary{{color:#68758a}}table{{width:100%;border-collapse:collapse}}th,td{{padding:5px;border:1px solid #dce3ed;text-align:left;vertical-align:top}}th{{background:#edf3ff}}.s{{font-weight:bold}}.pass,.passed{{color:#08784d}}.fail,.failed{{color:#bd2435}}.needs_review{{color:#9a6500}}.pending,.cancelled{{color:#64748b}}footer{{margin-top:20px;color:#68758a}} </style></head><body><header><h1>AI Maestro — Overall Test Coverage</h1><p>Generated {generated} · Latest case result per flow plus cumulative execution history</p></header><div class="metrics">{cards}</div><div class="overview"><section><h2>User-state coverage</h2><table><tr><th>User state</th><th>Flows</th><th>Coverage</th></tr>{state_rows}</table></section><section><h2>Module coverage</h2><table><tr><th>Module</th><th>Flows</th></tr>{module_rows}</table></section><section><h2>Cumulative case executions</h2><table><tr><th>Verdict</th><th>Result records</th></tr>{history_rows}</table></section><section><h2>Execution jobs</h2><table><tr><th>Status</th><th>Jobs</th></tr>{job_rows}</table></section></div>{''.join(sections)}<footer>Sources: curated user suites, approved Ready workbooks, scenario YAML metadata, and portal.db execution results.</footer></body></html>'''

def main():
    parser=argparse.ArgumentParser(); parser.add_argument("--pdf",action="store_true"); args=parser.parse_args()
    OUT.mkdir(parents=True,exist_ok=True); cases=collect_cases(); jobs,history=collect_results(cases); content=render(cases,jobs,history)
    html_path=OUT/"AI_Maestro_Overall_Coverage_Report.html"; pdf_path=OUT/"AI_Maestro_Overall_Coverage_Report.pdf"; html_path.write_text(content,encoding="utf-8")
    latest_manifest = Counter(item["status"] for item in cases.values())
    state_manifest = Counter(item["user_state"] for item in cases.values())
    for label in ("PASS", "FAIL", "NEEDS_REVIEW", "CANCELLED", "PENDING"):
        latest_manifest.setdefault(label, 0)
    for label in ("ANONYMOUS", "SUBSCRIBER", "REGISTERED_USER", "EXPIRED_USER"):
        state_manifest.setdefault(label, 0)
    manifest = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "distinct_flows": len(cases),
        "latest_case_status": dict(latest_manifest),
        "cumulative_result_records": dict(history),
        "job_lifecycle": dict(Counter(job["status"].upper() for job in jobs)),
        "user_states": dict(state_manifest),
        "modules": dict(Counter(item["module"] for item in cases.values())),
    }
    manifest_path = OUT / "AI_Maestro_Overall_Coverage_Report.audit.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    if args.pdf:
        subprocess.run([str(EDGE),"--headless","--disable-gpu",f"--print-to-pdf={pdf_path}",html_path.resolve().as_uri()],check=True,timeout=90)
    print(json.dumps({"cases":len(cases),"html":str(html_path),"pdf":str(pdf_path) if pdf_path.exists() else "","audit":str(manifest_path)}))

if __name__ == "__main__": main()

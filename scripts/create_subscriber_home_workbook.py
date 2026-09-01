from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Subscriber_Home_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
]

CASES = [
    ("SUB_HOME_001", "Fresh subscriber launch and Home entitlement", [
        ("Launch the app in a fresh state and sign in with the configured active subscriber credentials.", "Subscriber authentication succeeds and Home loads."),
        ("Wait for the Home screen.", "The Home screen is visible."),
        ("After a short Home load wait, verify the Subscribe action, Advertisement label, and Google sticky advertisement container.", "Subscribe, Advertisement, and the sticky advertisement container are not visible for the active subscriber."),
    ]),
    ("SUB_HOME_002", "Subscriber Home pull to refresh", [
        ("Establish the active subscriber Home state.", "Home is visible without Subscribe or Advertisement."),
        ("Pull down on the Home feed to refresh.", "Refresh completes successfully."),
        ("After refresh, verify the Home identity, The Hindu logo, hamburger menu, and refreshed article content.", "Home remains visible with The Hindu logo, hamburger menu, and at least one article card available."),
    ]),
    ("SUB_HOME_003", "No subscriber monetisation while scrolling Home", [
        ("Establish the active subscriber Home state.", "Home is visible."),
        ("Scroll the Home feed slowly through multiple viewports.", "Home scrolls without interruption."),
        ("Verify inline advertisements, sticky advertisements, Taboola advertising, and Subscribe upsells.", "None of those monetisation elements is visible for the subscriber."),
    ]),
    ("SUB_HOME_004", "Subscriber Home content and navigation", [
        ("Establish the active subscriber Home state.", "Home is visible."),
        ("Scroll through Home sections, widgets, timestamps, and article cards.", "Applicable Home content renders and remains usable."),
        ("Open one Home section and return to Home.", "Section navigation works and Home is restored with the subscriber session retained."),
    ]),
    ("SUB_HOME_005", "Subscriber full article access", [
        ("Establish the active subscriber Home state and open a verified article card.", "The article detail screen opens."),
        ("Scroll through the full article.", "The complete entitled article remains readable."),
        ("Verify paywall, Subscribe prompt, and Advertisement.", "Paywall, Subscribe prompt, and Advertisement are not visible."),
    ]),
    ("SUB_HOME_006", "Subscriber premium article access", [
        ("Open a Home article as the active subscriber, then swipe left five times and right five times to search Home articles for a Premium badge.", "Home article paging works; Premium discovery is conditional because the live Home feed is dynamic."),
        ("If a Premium badge is found, scroll through that Home-opened article and capture its header plus three body screenshots.", "A discovered Premium article remains fully readable for the subscriber; no Premium result is a valid skip with search evidence."),
        ("During every article swipe and Premium evidence scroll, verify subscription restrictions, advertisements, and interstitials are absent.", "No paywall, Subscribe upsell, Advertisement, or advertising interstitial interrupts the subscriber Home article flow."),
    ]),
    ("SUB_HOME_007", "Subscriber entitled AI Summary", [
        ("Open the controlled long-form article that is confirmed to support AI Summary as an active subscriber.", "The eligible article detail screen and AI Summary action are visible."),
        ("Tap AI Summary.", "The entitled AI Summary sheet opens."),
        ("Verify Summary and Article FAQs.", "Summary and Article FAQs are visible, and no subscription prompt appears."),
    ]),
    ("SUB_HOME_008", "Subscriber article Bookmark", [
        ("Open a verified article as an active subscriber.", "The article detail screen is visible."),
        ("Tap Bookmark.", "The article becomes bookmarked and its selected state or success confirmation is visible."),
        ("Verify authentication behavior.", "Login to your account is not shown because the subscriber is authenticated."),
    ]),
    ("SUB_HOME_009", "Subscriber article Share", [
        ("Open a verified article as an active subscriber.", "The article detail screen is visible."),
        ("Tap Share.", "The platform share sheet opens with available share targets."),
        ("Close the share sheet.", "The same article detail screen is restored."),
    ]),
    ("SUB_HOME_010", "Subscriber article Comment access", [
        ("Open a verified article as an active subscriber and scroll until Post a comment is visible.", "Post a comment is visible."),
        ("Open Post a comment and enter non-destructive test text without submitting it.", "Comment entry is available and accepts the test text."),
        ("Verify authentication behavior and exit without posting.", "Login to your account is not shown, and no public comment is submitted."),
    ]),
    ("SUB_HOME_011", "Subscriber post-article sections", [
        ("Open controlled article test data that supports post-article sections.", "The article detail screen is visible."),
        ("Scroll to the end of the article.", "The post-article region is reached."),
        ("Verify the sections required by the controlled article data.", "Related Topics, Recommended, and Headlines appear when specified for that article type."),
    ]),
    ("SUB_HOME_012", "Subscriber article pager session", [
        ("Scroll to a randomly selected bounded position in the live Home feed and open a currently visible article without using a fixed headline.", "A dynamically selected Home article opens in the active subscriber session."),
        ("Swipe left through five articles and then swipe right through five articles.", "Article paging works five times in both directions."),
        ("Verify subscriber state and interruptions after each swipe.", "The subscriber session remains active and no Subscribe upsell or advertising interstitial appears."),
    ]),
    ("SUB_HOME_013", "Subscriber plan entitlements", [
        ("Establish the active subscriber Home state using credentials with known plan entitlements.", "The subscribed session is active."),
        ("Open each feature guaranteed by the configured plan: Games, ePaper, or Ebooks.", "Each plan-entitled feature opens successfully."),
        ("Verify access restrictions.", "No Subscribe upsell is shown for features included in that subscriber plan."),
    ]),
    ("SUB_HOME_014", "Subscriber logout state boundary", [
        ("Establish the active subscriber Home state and open Account.", "The subscriber user menu is visible."),
        ("Scroll to Logout, tap it, and confirm when required.", "The subscriber session ends."),
        ("Open Account again.", "The anonymous Login and Create account actions are visible."),
    ]),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscriber Home"
    ws.append(HEADERS)
    precondition = (
        "The Hindu app is installed; network is available; configured credentials have "
        "an active subscription and known plan entitlements."
    )
    for case_id, name, steps in CASES:
        for number, (step, expected) in enumerate(steps, 1):
            ws.append([case_id, name, "Subscriber Home", "SUBSCRIBER", precondition,
                       number, step, expected, "High", "Yes"])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [18, 42, 24, 16, 66, 14, 72, 72, 12, 14]
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, column).column_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()

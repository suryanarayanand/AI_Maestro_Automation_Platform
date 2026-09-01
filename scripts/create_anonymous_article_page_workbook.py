"""Create the portal-ready anonymous Article Page test workbook."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Anonymous_Article_Page_Approved_Test_Cases.xlsx"
MATRIX = ROOT / "Uploads" / "Source" / "THG App_Functionality_Matrix.xlsx"
MASTER_SCENARIOS = ROOT / "Uploads" / "Source" / "TH App Testing Scenarios_AutomationCopy.xlsx"
HEADERS = (
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "runtime_assertion",
)
PRECONDITION = (
    "The Hindu app is installed; network is available; the user is signed out; "
    "use a controlled active Article Library URL matching the case entitlement. "
    "Skip the welcome popup when present. For an interstitial, capture evidence, "
    "wait until its close control is available (at least 5 seconds), and close only the ad."
)


CASES = [
    ("ANON_ARTICLE_001", "Article identity, header and metadata", "Critical", [
        ("Open a controlled free or unblocked article from an approved section.", "The Article Page opens without leaving the app.", "assertVisible screen_article_detail and headline"),
        ("Verify the section label, headline, publication/update metadata and estimated reading time when supplied by the article.", "Article identity and available metadata are readable and not overlapped.", "assert headline; conditional section/time assertions; wait; screenshot"),
    ]),
    ("ANON_ARTICLE_002", "Readable article content and vertical scrolling", "Critical", [
        ("Open a controlled free or unblocked article and scroll through at least four viewports.", "Body content remains readable and scrolling works to the lower article sections.", "repeat swipe UP 4; assert article body"),
        ("Capture settled evidence at the top, middle and lower article positions.", "Three distinct readable article positions are recorded.", "wait before each of 3 screenshots"),
    ]),
    ("ANON_ARTICLE_003", "Inline, sticky and blank advertisement handling", "High", [
        ("Scroll an anonymous readable article and inspect configured inline and bottom sticky advertisement slots.", "Eligible advertising may render without blocking article navigation; overlapping sticky advertising is acceptable.", "conditional assert ADVERTISEMENT; wait; screenshot"),
        ("If an advertisement slot is blank or a non-sticky ad obscures content, wait for the UI to settle and capture evidence without marking normal sticky overlap as a defect.", "Blank or obstructive non-sticky advertising is preserved as NEEDS_REVIEW bug evidence.", "conditional blank/obstruction screenshot; review classification"),
    ]),
    ("ANON_ARTICLE_004", "Interstitial advertisement safe close", "Critical", [
        ("Continue article interactions until an interstitial appears or the controlled observation window ends.", "The app remains responsive whether the frequency-capped interstitial appears or not.", "conditional interstitial detection"),
        ("When it appears, wait at least 5 seconds and continue waiting for the close control on video ads; capture evidence before closing only the advertisement.", "The ad is closed only after its close control becomes available; the application is not closed.", "wait 5000; extended conditional wait; screenshot; tap ad close"),
        ("Verify the same Article Page is visible after dismissal.", "The reader returns to the original article.", "assertVisible screen_article_detail and headline"),
    ]),
    ("ANON_ARTICLE_005", "Floating article action bar", "Critical", [
        ("Open an article, scroll enough to expose the floating action bar, and inspect its controls.", "Reading options, Bookmark, Share, Comment and Subscribe controls are available as applicable.", "assert floating controls and Subscribe; wait; screenshot"),
        ("Verify Gift this article is not offered to the anonymous user.", "Subscriber-only gifting is absent.", "assertNotVisible Gift this article"),
    ]),
    ("ANON_ARTICLE_006", "Reading options and text-size changes", "High", [
        ("Open Reading Options from the article floating bar.", "The reading options panel opens with text-size controls.", "tap Reading Options; assert text-size controls; screenshot"),
        ("Apply a larger text size and then a smaller text size to the same paragraph, allowing the UI to settle before each capture.", "The article typography visibly changes without clipping or losing position.", "tap larger; wait; screenshot; tap smaller; wait; screenshot"),
        ("Close Reading Options.", "The panel closes and the Article Page remains active.", "close panel; assertVisible screen_article_detail"),
    ]),
    ("ANON_ARTICLE_007", "Same article across light and dark themes", "High", [
        ("Open a controlled article in the current theme and capture the settled article header and body.", "Baseline theme evidence is recorded.", "wait; screenshot"),
        ("Change to Dark theme, return to the same article, and verify headline, body, images and controls remain legible.", "The same article renders correctly in Dark theme.", "select Dark; reopen same article; assert headline; wait; screenshot"),
        ("Change to Light theme and recheck the same content before restoring the original setting.", "The same article renders correctly in Light theme and the test does not leave a changed preference behind.", "select Light; assert headline; wait; screenshot; restore theme"),
    ]),
    ("ANON_ARTICLE_008", "Listen to Article playback for accessible content", "Critical", [
        ("Open a controlled free or unblocked article that supports Listen to Article and tap the audio control.", "The audio player opens and playback starts.", "assert/tap Listen to article; assert player/playback state"),
        ("Listen for at least 30 seconds, then pause playback.", "Playback progresses for at least 30 seconds and can be paused.", "wait 30000; tap Pause; assert paused/progress; wait; screenshot"),
        ("Resume briefly and stop or close the player without leaving the app.", "Audio controls remain responsive and the Article Page is retained.", "tap Play; wait; close player; assert screen_article_detail"),
    ]),
    ("ANON_ARTICLE_009", "Audio entitlement follows paywall state", "Critical", [
        ("Open a controlled anonymous paywall-blocked article and inspect the available reader actions.", "The paywall blocks the body content.", "assert paywall/Subscribe"),
        ("Verify Listen to Article is unavailable while the article is blocked.", "TTS is hidden for anonymous paywall-blocked premium or metered content.", "assertNotVisible Listen to article; wait; screenshot"),
    ]),
    ("ANON_ARTICLE_010", "AI Summary conditional entitlement", "Critical", [
        ("Open a controlled eligible article published after the feature launch with sufficient content and tap AI Summary.", "AI Summary opens when the article supports it; unsupported short/old article data is recorded as SKIPPED, not failed.", "conditional assert/tap AI Summary"),
        ("For an unblocked article, verify complete Summary content and available Article FAQs; for a blocked article, verify a subscription message and Subscribe button instead of summary content.", "AI Summary follows the article entitlement without exposing blocked content.", "conditional unblocked Summary/FAQs OR blocked subscription assertion; wait; screenshot"),
        ("Close the AI Summary panel.", "The same Article Page remains visible.", "close summary; assert screen_article_detail"),
    ]),
    ("ANON_ARTICLE_011", "Premium article reader paywall", "Blocker", [
        ("Open a controlled article carrying the Premium badge and scroll toward its restricted body.", "The Premium article opens but full reading is blocked for the anonymous user.", "assert Premium; swipe UP until paywall"),
        ("Verify the paywall offer, Subscribe action, and Already a subscriber? Login link; scroll when required to expose the Login link.", "Subscription offers and anonymous Login recovery are visible, while restricted body content is not exposed.", "assert paywall/offer and Subscribe; scroll; assert Already a subscriber? Login; wait; screenshot"),
    ]),
    ("ANON_ARTICLE_012", "Metered non-premium article paywall", "Critical", [
        ("Open a controlled non-premium article known to be blocked by the current metering state and scroll toward its restricted body.", "The configured metering paywall blocks continued reading even without a Premium badge.", "assertNotVisible Premium badge; assert paywall/Keep reading/offer"),
        ("Verify Subscribe and Already a subscriber? Login are available and capture the settled restriction.", "Anonymous subscription and login recovery actions are visible.", "assert Subscribe and Already a subscriber? Login; wait; screenshot"),
    ]),
    ("ANON_ARTICLE_013", "Anonymous Bookmark redirects to sign-in", "Critical", [
        ("Tap Bookmark from an article without authenticating.", "The app opens Login to your account instead of saving anonymously.", "tap Bookmark; assert Login to your account; wait; screenshot"),
        ("Press Back once.", "The original Article Page is restored.", "back; assert screen_article_detail and headline"),
    ]),
    ("ANON_ARTICLE_014", "Article sharing without forced authentication", "High", [
        ("Tap Share from the article action bar.", "The Android share sheet opens with the article share payload.", "tap Share; assert system share sheet; wait; screenshot"),
        ("Dismiss the share sheet without selecting a target.", "The same Article Page is restored and remains usable.", "back; assert screen_article_detail and headline"),
    ]),
    ("ANON_ARTICLE_015", "Comment authentication and paywall dependency", "Critical", [
        ("On a controlled unblocked article, scroll to Post a Comment and tap it.", "The comment surface requests anonymous users to sign in.", "assert/tap Post a Comment; assert Sign in/Login; wait; screenshot"),
        ("Return to the article; on a paywall-blocked article verify that lower comment, related and headline sections are not falsely required beyond the paywall.", "Unreachable lower sections are treated as correctly blocked, not automation failures.", "back; conditional paywall assertion; no forced lower-section search"),
    ]),
    ("ANON_ARTICLE_016", "Related headlines and recommended navigation", "High", [
        ("On a controlled unblocked article, scroll to Related Stories, Related Headlines or Recommended Articles.", "At least one supported recommendation section is visible below the readable article.", "scrollUntilVisible related/recommended; assert section; wait; screenshot"),
        ("Open one recommended article and verify its Article Page identity differs from the source headline.", "The selected recommendation opens as a new valid Article Page.", "capture source headline; tap recommendation; assert screen_article_detail and changed headline; wait; screenshot"),
        ("Press Back and verify the source article is restored.", "Navigation history returns to the original article.", "back; assert original headline"),
    ]),
    ("ANON_ARTICLE_017", "Taboola recommendation and external return", "High", [
        ("On a controlled unblocked article, scroll to the end and inspect the Taboola or recommendation advertising area.", "Anonymous users can see the configured recommendation advertising area when inventory is available.", "conditional scrollUntilVisible Taboola/recommendation ad; wait; screenshot"),
        ("When an external recommendation is available, open it, capture the settled destination, and press Back.", "The destination opens and Back returns safely to The Hindu Article Page; missing live inventory is NEEDS_REVIEW, not an app failure.", "conditional tap ad; wait; screenshot; back; assert screen_article_detail"),
    ]),
    ("ANON_ARTICLE_018", "Swipe-to-read collection and read meter", "High", [
        ("Open an article from a swipe-enabled collection and verify the owl-dot or article-position indicator when supplied.", "The available collection position indicator is visible and aligned.", "conditional assert owl dots/article position; wait; screenshot"),
        ("Swipe left through three articles and then right through three articles, handling any interstitial with the shared safe-close rule.", "Article transitions work smoothly in both directions and each destination remains a valid Article Page.", "repeat swipe LEFT 3 and RIGHT 3; assert screen_article_detail after each"),
        ("Scroll the current article down and up and inspect the reading-progress meter.", "The available read meter updates with scroll position without visual corruption.", "conditional assert read meter; swipe UP/DOWN; wait; screenshot"),
    ]),
    ("ANON_ARTICLE_019", "Article Back navigation", "Critical", [
        ("Open an article from a known source section and tap the article Back icon.", "The app returns to the same source section rather than exiting.", "tap article Back; assert source section"),
        ("Reopen the article and press the Android system Back button.", "System Back also returns to the same source section.", "open same article; back; assert source section; wait; screenshot"),
    ]),
    ("ANON_ARTICLE_020", "Related Topics navigation", "High", [
        ("On a controlled unblocked article, scroll to Related Topics and capture the settled topic list.", "One or more related topic links are displayed when configured for the article.", "conditional scrollUntilVisible Related Topics; assert topic; wait; screenshot"),
        ("Tap one related topic and verify the matching topic or section page opens, then return.", "The selected topic routes to the correct listing and Back restores the source article.", "tap related topic; assert topic title/listing; wait; screenshot; back; assert source headline"),
    ]),
    ("ANON_ARTICLE_021", "Archive article paywall", "Critical", [
        ("Open a controlled Article Library reference published more than one year ago as an anonymous user.", "The archive article opens with an archive restriction instead of unrestricted full content.", "assert archive article headline; scrollUntilVisible paywall"),
        ("Verify the restriction offers Subscribe and Already a subscriber? Login.", "The anonymous archive paywall presents subscription and login recovery actions.", "assert Subscribe and Already a subscriber? Login; wait; screenshot"),
    ]),
    ("ANON_ARTICLE_022", "Paywall plans and safe dismissal", "Blocker", [
        ("From an anonymous article paywall, tap Subscribe.", "The subscription plan surface opens.", "tap Subscribe; assert plan surface"),
        ("Verify yearly and monthly plan choices and Already a subscriber? Login, scrolling when required.", "Both plan choices and the Login recovery link are available.", "assert yearly; assert monthly; scroll; assert Already a subscriber? Login; wait; screenshot"),
        ("Close the plan surface without purchasing.", "The plan surface closes and returns to the same blocked article without changing entitlement.", "tap close; assert original headline and paywall"),
    ]),
]


def validate_matrix():
    """Fail loudly if the source behavior matrix is unavailable or no longer relevant."""
    workbook = load_workbook(MATRIX, read_only=True, data_only=True)
    text = " ".join(str(cell) for row in workbook.active.iter_rows(values_only=True) for cell in row if cell)
    required = ("Article Page", "Paywall", "AI Summary", "TTS")
    missing = [term for term in required if term.lower() not in text.lower()]
    if missing:
        raise RuntimeError(f"Functionality matrix is missing required behavior: {missing}")

    master = load_workbook(MASTER_SCENARIOS, read_only=True, data_only=True)
    master_text = " ".join(
        str(cell) for row in master.active.iter_rows(values_only=True) for cell in row if cell
    )
    master_required = (
        "Article page_Swipe to read", "Article page_Read Meter", "Archive Paywall",
        "Related Topics", "Article page_Back icon navigation",
    )
    missing_master = [term for term in master_required if term.lower() not in master_text.lower()]
    if missing_master:
        raise RuntimeError(f"6,638-scenario workbook is missing required coverage: {missing_master}")


def build():
    validate_matrix()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous Article Page"
    sheet.append(HEADERS)
    for case_id, name, priority, steps in CASES:
        for number, (step, expected, assertion) in enumerate(steps, 1):
            sheet.append((case_id, name, "Article Page", "ANONYMOUS", PRECONDITION,
                          number, step, expected, priority, "Yes", assertion))

    header_fill = PatternFill("solid", fgColor="17324D")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {"A": 21, "B": 43, "C": 18, "D": 14, "E": 68, "F": 13,
              "G": 78, "H": 70, "I": 12, "J": 13, "K": 62}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Created {OUTPUT} with {len(CASES)} cases and {sheet.max_row - 1} executable steps")


if __name__ == "__main__":
    build()

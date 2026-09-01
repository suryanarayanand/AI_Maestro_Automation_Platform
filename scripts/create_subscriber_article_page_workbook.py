from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'Uploads'/'Ready'/'Subscriber_Article_Page_Approved_Test_Cases.xlsx'
HEAD=['test_case_id','name','module','user_state','precondition','step_number','step','expected_result','priority','automatable','master_sheet_reference','existing_yaml_reference']
PRE='The app, network and Maestro device are available; valid subscriber credentials come from protected portal configuration. Use a controlled current Article Library URL when article-specific eligibility is required. Wait before every screenshot.'
REF='Anonymous_Article_Page_Approved_Test_Cases.xlsx; SC_12_Subscriber_Article_Full_Read_Access.yaml; SC_Paywall_Verification_subsscribedAccount.yaml; SC_89_Opinion_Article_EndToEnd_Subscriber_Feature_Validation.yaml; SC_74_indiaPage_specific_article_SubscriberAccount.yaml'
MASTER='TH App Testing Scenarios_AutomationCopy.xlsx (6000+ scenarios); THG App Functionality Matrix'
CASES=[
('SUB_ARTICLE_001','Article identity, header and metadata','Blocker',[('Open a current article after subscriber login.','Article detail opens with headline, section and available publication/read-time metadata.'),('Wait and capture the header.','Header evidence is readable and unobstructed.')]),
('SUB_ARTICLE_002','Full article reading and vertical scrolling','Blocker',[('Open a controlled article and scroll through at least four viewports.','Body content remains readable throughout scrolling.'),('Capture top, middle and lower positions after waits.','Three settled reading positions are recorded.')]),
('SUB_ARTICLE_003','Subscriber no-advertisement contract','Blocker',[('Scroll the article from header toward its end.','Article content remains usable without monetisation overlays.'),('Verify Advertisement, ad iframe, Taboola advertising unit and Subscribe controls are absent.','No subscriber advertising or purchase prompt is rendered.')]),
('SUB_ARTICLE_004','Floating article action bar','Critical',[('Scroll enough to expose the floating action bar.','Reading options, Bookmark, Share and Comment controls are available.'),('Verify subscriber-only actions such as Gift this article when supported.','Entitled actions are displayed according to article eligibility.')]),
('SUB_ARTICLE_005','Reading options and text sizes','Critical',[('Open Reading Options and increase text size.','Text grows without clipping or losing the article position.'),('Decrease and restore the original size, then close the panel.','Controls respond and the Article Page remains active.')]),
('SUB_ARTICLE_006','Same article in Light and Dark themes','Critical',[('Capture a controlled article in Light theme.','Headline, body, media and controls are legible.'),('Open the same article in Dark theme, capture it, then restore the original theme.','The same content remains legible and the test restores its preference.')]),
('SUB_ARTICLE_007','Listen to Article playback','Critical',[('Open an eligible article and start Listen to Article.','Audio player opens and playback begins.'),('Listen for at least 30 seconds, pause, resume briefly and close.','Progress advances and playback controls remain responsive.')]),
('SUB_ARTICLE_008','AI Summary eligible article','Critical',[('Open a sufficiently long eligible article and locate AI Summary with bounded scrolling.','AI Summary is available for eligible content; short ineligible content is a valid skip.'),('Open Summary and verify full content without a Subscribe restriction.','The subscriber receives the entitled summary and available FAQs.')]),
('SUB_ARTICLE_009','Short article without AI Summary','Major',[('Open a controlled short article below the AI Summary eligibility threshold.','The article remains fully readable.'),('Verify missing AI Summary is treated as expected eligibility behavior.','The case passes without forcing an unavailable feature.')]),
('SUB_ARTICLE_010','Premium article full entitlement','Blocker',[('Open an article carrying a Premium badge and scroll through it.','The subscriber can read the Premium article.'),('Verify paywall, Keep reading offer and Already a subscriber Login are absent.','No anonymous restriction blocks entitled content.')]),
('SUB_ARTICLE_011','Archive article subscriber access','Critical',[('Open a controlled article older than one year and scroll its body.','The subscriber can access the entitled archive article.'),('Verify archive Subscribe/Login restrictions are absent.','Archive entitlement is honored.')]),
('SUB_ARTICLE_012','Bookmark add and remove','Critical',[('Tap Bookmark on an article.','Bookmark state changes without Login redirection.'),('Capture the state, remove the bookmark and verify restoration.','Bookmark toggling works and leaves no unwanted saved state.')]),
('SUB_ARTICLE_013','Share sheet and safe return','Critical',[('Tap Share on an article.','The Android share sheet opens with article content.'),('Dismiss it without sharing externally.','The same Article Page is restored.')]),
('SUB_ARTICLE_014','Post a Comment subscriber access','Critical',[('Scroll to Post a Comment and open it on an eligible readable article.','Comment UI opens without anonymous Login restriction.'),('Verify the comment input, then return without posting.','Comment access works without external mutation.')]),
('SUB_ARTICLE_015','Related Topics and recommended article','Major',[('Scroll to Related Topics or Recommended on a controlled article.','A supported related-content section appears when configured.'),('Open one item and navigate Back.','A new valid article opens and Back restores the source article.')]),
('SUB_ARTICLE_016','Swipe-to-read article session','Critical',[('From a swipe-enabled article collection, swipe left three times.','Each transition remains on a valid Article Page.'),('Swipe right three times and inspect the position/read meter when available.','Reverse paging works and optional progress UI updates correctly.')]),
('SUB_ARTICLE_017','Article Back navigation','Critical',[('Open an article from a known source and use the article Back icon.','The source section is restored.'),('Reopen it and use Android Back.','System Back also restores the source section without closing the app.')]),
('SUB_ARTICLE_018','Live article embedded-frame scrolling','Critical',[('Open a controlled live article and perform bounded scrolling within its embedded content.','The live updates frame scrolls without freezing the app.'),('Do not require automation to reach the absolute end; capture multiple settled positions.','Live-frame evidence is recorded without an unbounded scroll failure.')]),
('SUB_ARTICLE_019','Article media and image rendering','Major',[('Open an article containing hero and inline media and scroll through it.','Images/media placeholders resolve without broken or overlapping content.'),('Wait at representative media positions and capture them.','Loaded media evidence is recorded.')]),
('SUB_ARTICLE_020','Article session and entitlement recovery','Blocker',[('Open an article, use an action panel, close it, and navigate to a recommended article.','The subscriber session remains active across interactions.'),('Verify no Subscribe/Login/paywall controls appear after navigation.','Subscriber entitlement persists through the complete article journey.')]),
]
def main():
 wb=Workbook();ws=wb.active;ws.title='Subscriber Article Page';ws.append(HEAD)
 for cid,name,pri,steps in CASES:
  for n,(step,exp) in enumerate(steps,1):ws.append([cid,name,'Article Page','SUBSCRIBER',PRE,n,step,exp,pri,'Yes',MASTER,REF])
 for c in ws[1]:c.fill=PatternFill('solid',fgColor='273C75');c.font=Font(color='FFFFFF',bold=True)
 for i,w in enumerate([22,52,20,16,88,13,94,94,14,14,68,105],1):ws.column_dimensions[ws.cell(1,i).column_letter].width=w
 for row in ws.iter_rows():
  for c in row:c.alignment=Alignment(vertical='top',wrap_text=True)
 ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions;OUT.parent.mkdir(parents=True,exist_ok=True);wb.save(OUT)
 print(f'Created {OUT} with {len(CASES)} cases and {sum(len(x[3]) for x in CASES)} steps')
if __name__=='__main__':main()

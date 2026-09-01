from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Anonymous_Games_Approved_Test_Cases.xlsx"
MASTER = ROOT / "Uploads" / "Source" / "TH App Testing Scenarios_AutomationCopy.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "master_sheet_reference", "existing_yaml_reference",
]

COMMON_YAML = "SC_28_Anonymous_Games_validation.yaml; SC_48_Games_page_verification.yaml; SC_55_login_from_gamepage.yaml"

CASES = [
    ("ANON_GAMES_001", "Games page launch and selected navigation", "Blocker", ["games", "game page"], COMMON_YAML, [
        ("Launch the app signed out, dismiss the welcome popup when shown, and tap Games in bottom navigation.", "The Games page opens without an app crash or indefinite loading."),
        ("Verify the Games navigation icon is highlighted and capture the landing page.", "Games is visibly selected and screenshot evidence is saved."),
    ]),
    ("ANON_GAMES_002", "Anonymous Games landing access prompt", "Major", ["unlock", "free games", "login"], COMMON_YAML, [
        ("Open Games as an anonymous user and inspect the top access prompt.", "Unlock and play all free games and a Login action are visible."),
        ("Capture the prompt and verify it does not overlap the game catalogue.", "The anonymous access message is readable and the catalogue remains usable."),
    ]),
    ("ANON_GAMES_003", "Free Games catalogue availability", "Critical", ["sudoku", "word flower", "word search", "quiz"], COMMON_YAML, [
        ("Open Games and scroll through the free-games catalogue.", "Sudoku, The Hindu Mini, Easy Down, Word Flower, Word Search, and Quiz are present."),
        ("Verify each game card displays Login to play and capture the catalogue in multiple screenshots.", "Every free-game card clearly communicates the anonymous login restriction."),
    ]),
    ("ANON_GAMES_004", "Sudoku anonymous Login gate", "Critical", ["sudoku", "login to play"], "SC_27_SUDOKU.yaml; SC_27_SUDOKU_EASY.yaml; SC_27_SUDOKU_MINI.yaml; SC_55_login_from_gamepage.yaml", [
        ("Open Games, locate Sudoku, and verify Login to play.", "Sudoku is protected by the login gate."),
        ("Tap Sudoku or its Login to play action.", "Login to your account opens and Sudoku gameplay does not start."),
        ("Capture Login, press Back, and verify Games is restored.", "The anonymous session and Games context are retained."),
    ]),
    ("ANON_GAMES_005", "The Hindu Mini anonymous Login gate", "Critical", ["hindu mini", "login to play"], "SC_27_THE_HINDU_MINI.yaml; SC_55_login_from_gamepage.yaml", [
        ("Open Games, locate The Hindu Mini, and verify Login to play.", "The Hindu Mini is protected by the login gate."),
        ("Tap The Hindu Mini or its Login to play action.", "Login to your account opens and gameplay does not start."),
        ("Capture Login and return to Games.", "Back navigation restores Games in the anonymous state."),
    ]),
    ("ANON_GAMES_006", "Easy Down anonymous Login gate", "Critical", ["easy down", "login to play"], "SC_27_EASY_DOWN.yaml; SC_55_login_from_gamepage.yaml", [
        ("Open Games, locate Easy Down, and verify Login to play.", "Easy Down is protected by the login gate."),
        ("Tap Easy Down or its Login to play action.", "Login to your account opens and gameplay does not start."),
        ("Capture Login and return to Games.", "Back navigation restores Games in the anonymous state."),
    ]),
    ("ANON_GAMES_007", "Word Flower anonymous Login gate", "Critical", ["word flower", "login to play"], "SC_27_WORD_FLOWER.yaml; SC_55_login_from_gamepage.yaml", [
        ("Open Games, locate Word Flower, and verify Login to play.", "Word Flower is protected by the login gate."),
        ("Tap Word Flower or its Login to play action.", "Login to your account opens and gameplay does not start."),
        ("Capture Login and return to Games.", "Back navigation restores Games in the anonymous state."),
    ]),
    ("ANON_GAMES_008", "Word Search anonymous Login gate", "Critical", ["word search", "login to play"], "SC_27_WORD_SEARCH.yaml; SC_55_login_from_gamepage.yaml", [
        ("Open Games, locate Word Search, and verify Login to play.", "Word Search is protected by the login gate."),
        ("Tap Word Search or its Login to play action.", "Login to your account opens and gameplay does not start."),
        ("Capture Login and return to Games.", "Back navigation restores Games in the anonymous state."),
    ]),
    ("ANON_GAMES_009", "Quiz anonymous Login gate", "Critical", ["quiz", "login to play"], "SC_27_NEWS_QUIZ.yaml; SC_55_login_from_gamepage.yaml", [
        ("Open Games, locate Quiz, and verify Login to play.", "Quiz is protected by the login gate."),
        ("Tap Quiz or its Login to play action.", "Login to your account opens and the Quiz does not start."),
        ("Capture Login and return to Games.", "Back navigation restores Games in the anonymous state."),
    ]),
    ("ANON_GAMES_010", "Premium Games section presentation", "Major", ["exclusive play", "premium", "cryptic crossword"], "SC_27_CRYPTIC_CROSSWORD.yaml; SC_28_Anonymous_Games_validation.yaml", [
        ("Open Games and scroll to the premium-games area.", "Exclusive Play, the Premium Way is visible."),
        ("Verify Cryptic Crossword is present and capture the complete section.", "The premium title and its access presentation are readable without overlap."),
    ]),
    ("ANON_GAMES_011", "Cryptic Crossword subscription gate", "Critical", ["cryptic crossword", "subscription", "plan"], "SC_27_CRYPTIC_CROSSWORD.yaml; SC_28_Anonymous_Games_validation.yaml", [
        ("From Exclusive Play, the Premium Way, tap Cryptic Crossword.", "A subscription-plan screen opens and the crossword does not start."),
        ("Verify the available subscription plan choices and Already a subscriber? Login.", "Purchase and existing-subscriber paths are visible."),
        ("Capture the complete plan screen without purchasing.", "Anonymous premium-game restriction evidence is saved."),
    ]),
    ("ANON_GAMES_012", "Cryptic Crossword existing-subscriber Login route", "Critical", ["cryptic crossword", "already a subscriber", "login"], "SC_27_CRYPTIC_CROSSWORD.yaml; SC_55_login_from_gamepage.yaml", [
        ("Open the Cryptic Crossword subscription gate and tap Already a subscriber? Login.", "Login to your account opens."),
        ("Capture the Login page and press Back without authenticating.", "The subscription gate or Games page is restored safely."),
    ]),
    ("ANON_GAMES_013", "Anonymous Games access-boundary enforcement", "Blocker", ["games", "anonymous", "login", "subscription"], "SC_28_Anonymous_Games_validation.yaml; SC_55_login_from_gamepage.yaml", [
        ("Attempt to open representative free games and return after each Login gate.", "No free game starts before authentication."),
        ("Attempt to open Cryptic Crossword and return from its subscription gate.", "The premium game does not start without an eligible subscription."),
        ("Verify the user remains anonymous throughout the checks.", "No game or navigation path bypasses the required entitlement."),
    ]),
    ("ANON_GAMES_014", "Games refresh and navigation recovery", "Major", ["games", "refresh", "back navigation"], "SC_48_Games_page_verification.yaml; SC_28_Anonymous_Games_validation.yaml", [
        ("Open Games and pull down to refresh once, waiting for content to finish loading.", "The Games catalogue reloads without duplicate or missing sections."),
        ("Open a free-game Login gate, press Back, then open the Cryptic Crossword plan gate and press Back.", "Each Back action restores the Games page rather than closing the app."),
        ("Verify Games remains highlighted and capture the restored page.", "Navigation state and anonymous restrictions remain correct after refresh and round trips."),
    ]),
]


def master_reference(keywords):
    if not MASTER.is_file():
        return "Games scenarios in 6000+ master workbook"
    workbook = load_workbook(MASTER, read_only=True, data_only=True)
    matches = []
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            text = " ".join(str(value) for value in row if value is not None).lower()
            if any(keyword.lower() in text for keyword in keywords):
                matches.append(f"{sheet.title}!{row_number}")
                if len(matches) == 8:
                    workbook.close()
                    return ", ".join(matches)
    workbook.close()
    return ", ".join(matches) or "Games scenarios in 6000+ master workbook"


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous Games"
    sheet.append(HEADERS)
    precondition = (
        "The Hindu app is installed; network is available; user is signed out. "
        "Dismiss the welcome popup when shown. If an interstitial advertisement appears, "
        "capture it, wait until Close becomes available, close only the ad, and continue."
    )
    for case_id, name, priority, keywords, yaml_reference, steps in CASES:
        reference = master_reference(keywords)
        for number, (step, expected) in enumerate(steps, 1):
            sheet.append([
                case_id, name, "Games", "ANONYMOUS", precondition, number,
                step, expected, priority, "Yes", reference, yaml_reference,
            ])

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="273C75")
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [21, 52, 16, 16, 85, 14, 92, 92, 14, 14, 55, 75]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()

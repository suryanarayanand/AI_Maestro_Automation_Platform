from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Subscriber_Account_Settings_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "master_sheet_reference", "existing_yaml_reference",
]
PRECONDITION = (
    "The Hindu app is installed; network and Maestro device are available; valid subscriber "
    "credentials are supplied through protected portal configuration. Login before evidence capture."
)
REFERENCE = "SC_32_AccountSettingPage_subscriberAccount.yaml; ANON_ACCOUNT_*.yaml; Common/OPEN_SUBSCRIBER_HOME.yaml"
CASES = [
    ("SUB_ACCOUNT_001", "Subscriber Account page and identity state", "Blocker", [
        ("Log in as the configured subscriber and open Account.", "Account opens in an authenticated subscriber state."),
        ("Verify Account and Logout are available and Login, Create account, and Subscribe are absent.", "Subscriber identity and entitlement controls are correct."),
        ("Wait for loading and capture the Account page.", "Authenticated Account evidence is saved."),
    ]),
    ("SUB_ACCOUNT_002", "My Account profile field presentation", "Critical", [
        ("Open My Account and wait for profile content.", "My Account opens without authentication prompts."),
        ("Verify First name, Last name, Email, Password or Change password, and profile information controls while scrolling.", "Subscriber profile fields render and remain readable."),
        ("Capture top and lower profile evidence after waits.", "Profile presentation evidence is saved."),
    ]),
    ("SUB_ACCOUNT_003", "Phone field safe edit and clear", "Critical", [
        ("In My Account, scroll to the phone or mobile-number input.", "The phone input is visible and enabled."),
        ("Tap the input, enter a safe test number, and verify the typed value is shown.", "The field accepts numeric input without UI corruption."),
        ("Clear the entered test number and verify the field is empty; leave without saving.", "Temporary data is removed and no profile mutation is committed."),
    ]),
    ("SUB_ACCOUNT_004", "Subscription details for entitled subscriber", "Blocker", [
        ("Open Account and select Subscriptions.", "The subscriber subscription-details page opens."),
        ("Verify an active-plan or entitlement detail is visible and purchase/login prompts are absent.", "The active subscriber is not treated as anonymous or expired."),
        ("Wait and capture the subscription details.", "Entitlement evidence is saved."),
    ]),
    ("SUB_ACCOUNT_005", "Vouchers page navigation", "Major", [
        ("Open Vouchers from the account area.", "The Vouchers page opens without losing the subscriber session."),
        ("Verify the page title or valid voucher empty state, capture it, and return.", "Voucher behavior is documented and Account is restored."),
    ]),
    ("SUB_ACCOUNT_006", "Notification Inbox authenticated behavior", "Major", [
        ("Open Notification Inbox from Account.", "Notification Inbox opens for the subscriber."),
        ("Verify inbox content or a valid empty state, capture it, and return.", "No anonymous login restriction appears and Account is restored."),
    ]),
    ("SUB_ACCOUNT_007", "Subscriber History access", "Critical", [
        ("Open History from Account.", "History opens for the authenticated subscriber."),
        ("Verify history content or a valid empty state, capture it, and return.", "No login restriction appears and Account is restored."),
    ]),
    ("SUB_ACCOUNT_008", "Subscriber Bookmarks access", "Critical", [
        ("Open Bookmarks from Account.", "Bookmarks opens for the authenticated subscriber."),
        ("Verify saved content or a valid empty state, capture it, and return.", "No login restriction appears and Account is restored."),
    ]),
    ("SUB_ACCOUNT_009", "Text Size controls and restoration", "Major", [
        ("Open Text Size and exercise decrease and increase controls.", "The controls respond and the preview remains readable."),
        ("Restore the original size, wait, capture it, and return.", "The test leaves the text-size preference unchanged."),
    ]),
    ("SUB_ACCOUNT_010", "Appearance controls and restoration", "Major", [
        ("Open Appearance and switch to the alternate Light or Dark mode.", "The selected theme applies without unreadable or overlapping UI."),
        ("Wait, capture the alternate mode, restore the original mode, and return.", "Theme evidence is saved and the original preference is restored."),
    ]),
    ("SUB_ACCOUNT_011", "Account scrolling and navigation recovery", "Major", [
        ("Scroll Account from top to bottom, waiting before top, middle, and lower screenshots.", "All account rows remain readable and tappable."),
        ("Open and return from My Account, History, and one application setting.", "Back navigation restores Account without closing the app or losing login."),
    ]),
    ("SUB_ACCOUNT_012", "Logout control presence without session mutation", "Critical", [
        ("Scroll to Logout and verify it is visible and enabled.", "The authenticated account exposes a Logout control."),
        ("Capture the control without confirming logout.", "Logout coverage is recorded without ending the shared subscriber session."),
    ]),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Subscriber Account Settings"
    sheet.append(HEADERS)
    for case_id, name, priority, steps in CASES:
        for number, (step, expected) in enumerate(steps, 1):
            sheet.append([
                case_id, name, "Account Settings", "SUBSCRIBER", PRECONDITION,
                number, step, expected, priority, "Yes",
                "THG App Functionality Matrix and 6000+ scenario reference", REFERENCE,
            ])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="273C75")
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [22, 48, 22, 16, 82, 13, 90, 90, 14, 14, 58, 90]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT)
    print(f"Created {OUTPUT} with {len(CASES)} cases and {sum(len(c[3]) for c in CASES)} steps")


if __name__ == "__main__":
    main()

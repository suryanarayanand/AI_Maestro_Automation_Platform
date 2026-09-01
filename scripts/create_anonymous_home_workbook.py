from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Anonymous_Home_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
]

CASES = [
    ("ANON_HOME_001", "Fresh anonymous launch and Home", [
        ("Launch the app with cleared state without signing in and complete or skip onboarding.", "Home loads in anonymous state."),
        ("Wait for the Home screen.", "The Home screen is visible."),
        ("Verify the Subscribe action in the header.", "Subscribe is visible for the anonymous user."),
    ]),
    ("ANON_HOME_002", "Anonymous Home pull to refresh", [
        ("Establish fresh anonymous Home.", "Home and Subscribe are visible."),
        ("Pull down on the Home feed to refresh.", "Refresh completes successfully."),
        ("Verify Home content after refresh.", "Home remains visible and at least one article card is available."),
    ]),
    ("ANON_HOME_003", "Anonymous Home inline advertisement", [
        ("Establish fresh anonymous Home with ad-network connectivity.", "Home is visible."),
        ("Scroll the Home feed slowly through multiple viewports.", "Home scrolls without leaving the screen."),
        ("Verify a loaded inline advertisement slot.", "The validated Advertisement label is visible when test ad inventory is available."),
    ]),
    ("ANON_HOME_004", "Anonymous Home sticky advertisement", [
        ("Establish fresh anonymous Home with ad-network connectivity.", "Home is visible."),
        ("Scroll Home until the sticky advertisement is displayed.", "The verified Home sticky-ad container is visible."),
        ("Capture screenshot evidence of the sticky advertisement.", "A screenshot is saved without interacting with the advertisement."),
    ]),
    ("ANON_HOME_005", "Anonymous Home Taboola region", [
        ("Establish fresh anonymous Home and scroll toward the lower recommendation region.", "The lower Home feed is reached."),
        ("Verify the current validated Taboola or recommendation marker.", "The recommendation advertising region is visible when configured for the build."),
        ("Capture screenshot evidence.", "A screenshot of the recommendation region is saved."),
    ]),
    ("ANON_HOME_006", "Open Home article and validate controls", [
        ("Establish fresh anonymous Home and tap a verified article card.", "The article detail screen opens."),
        ("Dismiss an interstitial or subscription sheet only when it appears.", "The article detail screen is restored."),
        ("Verify applicable AI Summary, Bookmark, Share, Comment, and Subscribe controls.", "Controls required by the selected article test data are visible."),
    ]),
    ("ANON_HOME_007", "Anonymous article content and post-article sections", [
        ("Open controlled article test data from anonymous Home.", "The article detail screen is visible."),
        ("Scroll slowly through the article to the end.", "Article content, applicable Advertisement, and in-content Subscribe placements are encountered."),
        ("Verify post-article sections required by the controlled article.", "Post a comment, Related Topics, Recommended, and Headlines appear when supported by that article type."),
    ]),
    ("ANON_HOME_008", "Anonymous article left and right paging", [
        ("Open a verified article from anonymous Home.", "The article detail screen is visible."),
        ("Swipe left to the next article and then right to the previous article.", "Article paging works in both directions."),
        ("Verify anonymous state after each swipe.", "The article remains accessible under anonymous restrictions and Subscribe remains applicable."),
    ]),
    ("ANON_HOME_009", "Anonymous premium article and interstitial recovery", [
        ("Open controlled premium-badge article test data as an anonymous user.", "The premium article detail screen opens with the applicable restriction."),
        ("When an advertising interstitial appears, capture a screenshot and close it using the verified control.", "Evidence is saved and the article detail screen is restored."),
        ("Scroll the premium article and verify its restriction.", "The anonymous paywall or Subscribe restriction appears, and loaded Advertisement is allowed."),
    ]),
    ("ANON_HOME_010", "Anonymous AI Summary subscription gate", [
        ("Open the controlled long-form article confirmed to support AI Summary.", "The article detail screen and AI Summary action are visible."),
        ("Tap AI Summary.", "The anonymous AI Summary prompt opens with Subscribe access rather than entitled summary content."),
        ("Tap Subscribe in the prompt and verify the bottom paywall sheet without purchasing.", "The paywall sheet appears; no purchase or authentication is performed."),
    ]),
    ("ANON_HOME_011", "Anonymous Bookmark login gate", [
        ("Open a verified article as an anonymous user.", "The article detail screen is visible."),
        ("Tap Bookmark.", "The app navigates to Login to your account."),
        ("Verify Login to your account and capture a screenshot.", "The login page evidence is saved and the case ends without signing in."),
    ]),
    ("ANON_HOME_012", "Anonymous article Share", [
        ("Open a verified article as an anonymous user.", "The article detail screen is visible."),
        ("Tap Share.", "The platform share sheet opens with available share targets."),
        ("Capture evidence and close the share sheet.", "The same article detail screen is restored without sharing externally."),
    ]),
    ("ANON_HOME_013", "Anonymous Comment login gate", [
        ("Open a verified article and scroll until Post a comment is visible.", "Post a comment is visible."),
        ("Open Post a comment and enter non-destructive test text without submitting it.", "The comment entry accepts safe test text or proceeds to its authentication gate."),
        ("Verify Login to your account and capture a screenshot.", "The login page evidence is saved; no comment is submitted and no sign-in occurs."),
    ]),
    ("ANON_HOME_014", "Anonymous Home state after article return", [
        ("Open an article from fresh anonymous Home.", "The article detail screen is visible."),
        ("Navigate back to Home.", "The Home screen is restored."),
        ("Verify anonymous state.", "Subscribe remains visible and the user is not authenticated."),
    ]),
]


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Anonymous Home"
    ws.append(HEADERS)
    precondition = (
        "The Hindu app is installed; network is available; the case starts with cleared "
        "anonymous state. Controlled article data is supplied where the case requires it."
    )
    for case_id, name, steps in CASES:
        for number, (step, expected) in enumerate(steps, 1):
            ws.append([case_id, name, "Anonymous Home", "ANONYMOUS", precondition,
                       number, step, expected, "High", "Yes"])
    header_fill = PatternFill("solid", fgColor="9E480E")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    widths = [18, 44, 24, 16, 68, 14, 74, 74, 12, 14]
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, column).column_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()

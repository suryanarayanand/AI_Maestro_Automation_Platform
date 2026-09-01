"""Create the reviewed Subscriber Trending import workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = Path("Uploads/Ready/Subscriber_Trending_Approved_Test_Cases.xlsx")
PRECONDITION = (
    "The Hindu app is installed; network is available; configured credentials belong to an "
    "active subscriber; login is performed from local masked portal credentials."
)


CASES = [
    ("SUB_TREND_001", "Subscriber Trending launch and selected navigation", "Critical", [
        ("Authenticate with the configured active subscriber account and open Trending from the bottom navigation.", "Trending opens and its bottom-navigation item is selected.", "run OPEN_SUBSCRIBER_HOME; tap nav_trending; assert screen_trending and nav_trending"),
        ("Wait for Trending content to settle and capture the landing page.", "Loaded Subscriber Trending evidence is saved.", "waitForAnimationToEnd immediately before screenshot"),
        ("Verify Subscribe and advertisement surfaces are absent.", "No Subscribe action, Advertisement label, sticky ad, iframe ad, or Taboola unit is visible.", "run ASSERT_SUBSCRIBER_NO_MONETIZATION"),
    ]),
    ("SUB_TREND_002", "Subscriber Trending pull to refresh", "Critical", [
        ("Open Subscriber Trending and pull down from the content area.", "The refresh gesture completes without leaving Trending.", "swipe DOWN; extendedWaitUntil screen_trending"),
        ("Verify current Trending content reloads and capture it after loading settles.", "Trending remains usable and refreshed content evidence is saved.", "assert screen_trending and READ FULL ARTICLE; wait; screenshot"),
        ("Recheck the subscriber no-ad state after refresh.", "No monetisation surface is introduced by refresh.", "run ASSERT_SUBSCRIBER_NO_MONETIZATION"),
    ]),
    ("SUB_TREND_003", "Subscriber Trending All feed and no-ad entitlement", "Blocker", [
        ("Open the All tab and scroll through the feed in bounded steps.", "The All feed remains responsive while additional stories load.", "tap All; repeat bounded swipe UP"),
        ("At the top, middle, and lower feed positions, verify subscriber monetisation is absent.", "No inline, sticky, blank, interstitial, iframe, or Taboola advertisement is visible.", "repeat ASSERT_SUBSCRIBER_NO_MONETIZATION"),
        ("Capture the loaded lower-feed state after waiting.", "Evidence shows a usable ad-free subscriber feed.", "waitForAnimationToEnd; screenshot"),
    ]),
    ("SUB_TREND_004", "Trending masthead round trip for subscriber", "High", [
        ("From Trending, tap The Hindu masthead.", "Home opens for the authenticated subscriber.", "tap masthead; assert screen_home"),
        ("Return using the Trending bottom-navigation item.", "Trending reopens and remains selected.", "tap nav_trending; assert screen_trending and nav_trending"),
        ("Verify no login, Subscribe, or advertisement regression occurred.", "The subscriber session and no-ad entitlement remain active.", "assertNotVisible Login to your account; run ASSERT_SUBSCRIBER_NO_MONETIZATION"),
    ]),
    ("SUB_TREND_005", "Trending section-tab navigation", "Critical", [
        ("Open All, News, Business, Technology, Entertainment, and Sports one by one.", "Every selected tab loads its corresponding Trending feed.", "tap each visible tab; assert screen_trending and selected tab"),
        ("Wait and capture each selected section.", "Readable evidence is captured only after each section settles.", "waitForAnimationToEnd before every screenshot"),
        ("Verify each section remains free of subscriber advertisements.", "No monetisation UI is visible in any section.", "run ASSERT_SUBSCRIBER_NO_MONETIZATION after each section"),
    ]),
    ("SUB_TREND_006", "Subscriber opens Trending full article", "Blocker", [
        ("Open a current Trending card using READ FULL ARTICLE.", "The article detail screen opens.", "tap READ FULL ARTICLE; handle sheet; assert screen_article_detail"),
        ("Scroll through the article in bounded steps.", "The subscriber can continue through the article body without a paywall.", "repeat swipe UP; assert screen_article_detail"),
        ("Verify paywall, Subscribe, login offer, and advertisements are absent; capture the article.", "Full subscriber access is proven with no monetisation blockage.", "assertNotVisible Keep reading/Already a subscriber/Subscribe/Login; assert no ads; wait; screenshot"),
    ]),
    ("SUB_TREND_007", "Subscriber Premium Trending article entitlement", "Blocker", [
        ("Open a Trending article displaying a Premium badge when one is available within a bounded search.", "A Premium article opens, or bounded unavailability is documented without a false failure.", "bounded article selection; conditional Premium branch"),
        ("Scroll through the Premium article and capture multiple settled reading states.", "The active subscriber can read the Premium article body.", "repeat swipe UP; wait before 3 screenshots"),
        ("Verify no paywall, offer, Login link, Subscribe action, or advertisement blocks reading.", "Premium entitlement is active and ad-free.", "assertNotVisible paywall/login/Subscribe/ADVERTISEMENT/ad_iframe/Taboola"),
    ]),
    ("SUB_TREND_008", "Trending Listen to Article", "High", [
        ("Open a Trending article and select Listen to article when the control is available.", "The audio player opens for an eligible article.", "conditional tap Listen to article"),
        ("Play or continue listening for at least 30 seconds.", "Audio remains active without an authentication or subscription prompt.", "wait 30000; assert player/article state"),
        ("Capture the settled player state and return to the same article.", "Audio evidence is saved and Back restores the article.", "wait; screenshot; back; assert screen_article_detail"),
    ]),
    ("SUB_TREND_009", "Trending AI Summary subscriber access", "High", [
        ("Open a suitable Trending article and tap AI Summary when available.", "The AI Summary surface opens for the subscriber.", "conditional tap AI Summary"),
        ("Verify readable summary content is shown without a Subscribe button.", "The summary is entitled and not subscription-blocked.", "assertNotVisible Subscribe; assert summary surface"),
        ("Wait, capture the summary, and return to the article.", "Summary evidence is readable and article context is restored.", "wait; screenshot; back; assert screen_article_detail"),
    ]),
    ("SUB_TREND_010", "Trending reading options and text sizes", "Critical", [
        ("Open a Trending article and tap Text size or Reading Options.", "The reading-options panel opens.", "tap Text size; assert Reading Options and A-/A+"),
        ("Apply larger and smaller text settings and capture each settled article state.", "The article remains readable and the text setting responds.", "tap A+ and A-; close panel; wait before screenshots"),
        ("Close the panel and verify the article remains open without monetisation.", "Article context is preserved and no subscriber restriction appears.", "tap Close; assert screen_article_detail; assert no monetisation"),
    ]),
    ("SUB_TREND_011", "Subscriber Trending article Bookmark", "Critical", [
        ("Open a Trending article and tap Bookmark.", "The authenticated subscriber bookmark action is accepted.", "tap Bookmark"),
        ("Verify the article remains open and Login to your account does not appear.", "Bookmark does not invoke an authentication gate.", "assert screen_article_detail; assertNotVisible Login to your account"),
        ("Wait and capture the bookmarked state.", "Bookmark evidence is saved.", "waitForAnimationToEnd; screenshot"),
    ]),
    ("SUB_TREND_012", "Subscriber Trending article Share", "Critical", [
        ("Open a Trending article and tap Share.", "The system Quick Share surface opens.", "tap Share; assert Quick Share"),
        ("Wait and capture the share sheet.", "Share destinations are visible in evidence.", "waitForAnimationToEnd; screenshot"),
        ("Close Share and verify the same article is restored.", "Back returns safely to article detail.", "back; assert screen_article_detail"),
    ]),
    ("SUB_TREND_013", "Subscriber Trending comment access", "Critical", [
        ("Open an unrestricted Trending article and scroll to Post a comment.", "The comment action becomes available when the article supports comments.", "bounded swipe UP; conditional Post a comment"),
        ("Tap Post a comment and verify the subscriber is not sent to account login.", "The comment surface opens without Login to your account.", "tap Post a comment; assertNotVisible Login to your account"),
        ("Wait, capture the comment state, and return to the article.", "Comment access evidence is saved and article context is restored.", "wait; screenshot; back; assert screen_article_detail"),
    ]),
    ("SUB_TREND_014", "Trending cross-section article entitlement", "Blocker", [
        ("Open one current article from News, Business, Technology, Entertainment, and Sports.", "Each selected article opens from its correct section.", "bounded loop by section; tap READ FULL ARTICLE; assert article"),
        ("For every opened article, verify no paywall, Subscribe, login offer, or advertisement appears.", "Subscriber entitlement and no-ad behavior are consistent across sections.", "assertNotVisible paywall/Subscribe/Login/ads"),
        ("Return after each article and verify the originating section remains selected.", "Navigation context is restored correctly.", "back; assert screen_trending and section label"),
    ]),
    ("SUB_TREND_015", "Trending article paging in both directions", "High", [
        ("Open a Trending article and swipe left five times, waiting after every swipe.", "Adjacent articles can be paged in the left direction.", "repeat 5 swipe LEFT; wait"),
        ("Swipe right five times, waiting after every swipe.", "Article paging also works in the right direction.", "repeat 5 swipe RIGHT; wait"),
        ("Verify each reached article remains readable and free of paywall and advertisements; capture settled evidence.", "Subscriber paging preserves full, ad-free article access.", "assert article; assert no monetisation; wait; screenshot"),
    ]),
    ("SUB_TREND_016", "Trending section and scroll-state restoration", "High", [
        ("Select Business and scroll to a lower feed position.", "Business remains selected at the lower position.", "tap Business; repeat swipe UP"),
        ("Open a current article and return using Back.", "The Business Trending feed is restored.", "tap READ FULL ARTICLE; assert article; back; assert screen_trending"),
        ("Verify Business remains selected and capture the restored state after waiting.", "Section context is retained without reauthentication or monetisation.", "assert Business; assert no monetisation; wait; screenshot"),
    ]),
    ("SUB_TREND_017", "Trending long-article complete reading", "Blocker", [
        ("Open a long-form Trending article and capture its header after loading.", "The article header is readable.", "assert article; wait; screenshot"),
        ("Scroll through the article body in bounded steps, capturing at least three settled positions.", "The subscriber can read continuously through the article.", "repeat swipe UP; wait before screenshots"),
        ("Verify no paywall or advertisement interrupts reading and that post-article content is reachable when supplied.", "Complete entitled reading is proven without monetisation blockage.", "assertNotVisible restrictions/ads; conditional Related/Latest/Post a comment"),
    ]),
    ("SUB_TREND_018", "Subscriber session persistence across Trending", "Blocker", [
        ("Navigate repeatedly between Home, Trending, a Trending article, and back.", "Navigation succeeds without losing the authenticated state.", "Home/Trending/article round trip"),
        ("Verify Login to your account and Subscribe never appear during the round trip.", "The subscriber session remains active.", "assertNotVisible Login to your account and Subscribe"),
        ("Verify advertisements remain absent and capture the final Trending state after waiting.", "The final page remains ad-free and usable.", "run ASSERT_SUBSCRIBER_NO_MONETIZATION; wait; screenshot"),
    ]),
]


def main():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Subscriber Trending"
    headers = [
        "test_case_id", "name", "module", "user_state", "precondition",
        "step_number", "step", "expected_result", "priority", "automatable",
        "runtime_assertion",
    ]
    sheet.append(headers)
    for case_id, name, priority, steps in CASES:
        for number, (step, expected, assertion) in enumerate(steps, 1):
            sheet.append([
                case_id, name, "Trending", "SUBSCRIBER", PRECONDITION,
                number, step, expected, priority, "Yes", assertion,
            ])
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [20, 44, 16, 16, 60, 13, 72, 68, 12, 13, 72]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Created {OUTPUT} with {len(CASES)} cases and {sheet.max_row - 1} steps")


if __name__ == "__main__":
    main()

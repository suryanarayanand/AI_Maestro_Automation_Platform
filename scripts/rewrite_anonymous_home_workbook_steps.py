"""Rewrite Anonymous Home Excel steps to match the validated Maestro behavior."""

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Uploads" / "Anonymous_Home_Approved_Test_Cases.xlsx"

CASES = {
    "ANON_HOME_001": ("Launch anonymous Home and dismiss onboarding", [
        ("Launch The Hindu with cleared app data and complete anonymous onboarding.", "Anonymous Home opens without authentication."),
        ("If the Welcome to a new The Hindu experience popup appears, capture it and tap Skip.", "The popup closes and does not block Home."),
        ("Verify Home, Subscribe, and at least one article card.", "Home is ready with anonymous controls and article content."),
    ]),
    "ANON_HOME_002": ("Refresh Home and verify stable content", [
        ("Open anonymous Home and pull down once to refresh.", "The refresh gesture completes without leaving Home."),
        ("Wait for refreshed content to finish loading.", "Home and article cards become visible within the readiness timeout."),
        ("Verify Home, Subscribe, and article content remain available.", "The refreshed Home feed is usable and remains anonymous."),
    ]),
    "ANON_HOME_003": ("Verify advertisements across the Home feed", [
        ("Open anonymous Home and scroll through the first advertisement viewport.", "An ADVERTISEMENT placement becomes visible within the maximum wait."),
        ("Assert ADVERTISEMENT is visible and capture the first viewport.", "The first advertisement placement has evidence; an empty container is marked for review."),
        ("Continue scrolling, assert a later ADVERTISEMENT placement, and capture it.", "A second advertisement placement is evidenced and Home remains open."),
    ]),
    "ANON_HOME_004": ("Verify the sticky Home advertisement while scrolling", [
        ("Open anonymous Home and scroll upward through six feed positions.", "Home remains responsive while the feed moves."),
        ("Wait for and assert the sticky ADVERTISEMENT placement.", "The sticky advertisement label is visible within the readiness timeout."),
        ("Capture the sticky advertisement and verify Home remains open.", "Evidence is saved; blank or overlapping ad content is marked for review."),
    ]),
    "ANON_HOME_005": ("Verify Taboola recommendations in the lower Home feed", [
        ("Open anonymous Home and scroll toward the lower feed.", "The lower Home feed loads without leaving the app."),
        ("When a Taboola advertising unit appears, verify and capture it; otherwise continue the bounded search.", "Taboola evidence is captured when supplied; absence after the bounded search is a valid content-data skip."),
        ("Capture the final lower-feed viewport and verify Home.", "The feed remains usable after the recommendation search."),
    ]),
    "ANON_HOME_006": ("Open a Home article and validate article controls", [
        ("Open the first available Home article and close any interstitial safely.", "The article detail screen opens and the app remains active."),
        ("Conditionally verify AI Summary, Bookmark, Share, and Comment controls.", "Each control supplied by the selected article is visible; unsupported optional controls are skipped."),
        ("Verify the anonymous Subscribe control on the article.", "Subscribe is visible and the article remains in anonymous state."),
    ]),
    "ANON_HOME_007": ("Validate article content and post-article sections", [
        ("Open an anonymous Home article and verify article detail.", "The article detail screen is visible."),
        ("Scroll through the article to its end, capturing content, advertisements, or Subscribe placements encountered.", "The article scroll completes; visible dynamic sections have evidence."),
        ("Verify applicable Post a comment, Related Topics, Recommended, or Headlines sections; if a reader paywall appears, capture its offer and Already a subscriber Login instead.", "Supported post-article sections are visible, or the paywall is evidence that those blocked sections are not applicable."),
    ]),
    "ANON_HOME_008": ("Page through Home articles and capture Premium evidence", [
        ("Open a Home article, swipe left five times, then swipe right five times.", "Article paging works in both directions and article detail remains open."),
        ("For every interstitial, capture it, wait in 5-second windows, close it when possible, and verify article recovery.", "Only the ad closes; delayed/no-close video ads safely return to the article."),
        ("When Premium or ADVERTISEMENT appears, wait 5 seconds and capture the header/body or advertisement placement.", "Premium evidence is conditional. A blank ad is captured for a bug; sticky-ad overlap is allowed, while other ad overlap requires review."),
    ]),
    "ANON_HOME_009": ("Discover Premium articles while paging from Home", [
        ("Open a Home article without using the Premium navigation tab.", "A Home article detail screen opens."),
        ("Search five articles left and five right, safely handling every interstitial and waiting 5 seconds for visible advertisements.", "The search completes without opening an ad destination; visible advertisement placements are captured after their loading window."),
        ("Conditionally capture Premium evidence and any blank advertisement remaining after the wait.", "No Premium article is a valid skip. Blank ads become bug evidence; sticky-ad overlap is allowed, while other ad overlap requires review."),
    ]),
    "ANON_HOME_010": ("Validate conditional AI Summary and article paywall states", [
        ("Open an anonymous Home article and conditionally open AI Summary.", "No AI Summary on a short article is a valid skip; otherwise the summary opens."),
        ("Identify and capture either the AI Summary Subscribe gate or the complete summary.", "The observed gated or full-summary state is evidenced without forcing the other state."),
        ("Return to the article and detect Keep reading, discount, or Already a subscriber reader-paywalls; when shown, tap Login or Sign in.", "A displayed paywall blocks content, shows its offer and subscriber-login control, and that control opens Login to your account; unrestricted articles may complete without a paywall."),
    ]),
    "ANON_HOME_011": ("Validate the anonymous Bookmark login gate", [
        ("Open an anonymous Home article and verify Bookmark.", "Bookmark is visible on article detail."),
        ("Tap Bookmark.", "The anonymous login gate opens instead of saving the article."),
        ("Assert Login to your account and capture the gate.", "The login requirement is clearly displayed with evidence."),
    ]),
    "ANON_HOME_012": ("Validate the article Share sheet", [
        ("Open a Home article and verify Share.", "Share is visible on article detail."),
        ("Tap Share and wait for the Android share sheet.", "Quick Share becomes visible."),
        ("Capture the share sheet, press Back, and verify article detail.", "The share sheet closes and the same article is restored."),
    ]),
    "ANON_HOME_013": ("Validate the anonymous Comment sign-in gate", [
        ("Open a Home article and search for Post a comment while also detecting Keep reading, discount, or Already a subscriber paywalls.", "Either the comment action is found or a reader paywall is identified."),
        ("If Post a comment is available, tap it; if a paywall blocks the article, assert its Login or Sign in link and skip the comment action.", "The anonymous comment sign-in gate appears, or the paywall is recorded as the valid reason the comment is not applicable."),
        ("Capture SIGN IN AND JOIN THE CONVERSATION or the blocking paywall and its Already a subscriber Login evidence.", "The applicable anonymous restriction is saved without treating blocked article sections as a failure."),
    ]),
    "ANON_HOME_014": ("Return from an article to anonymous Home", [
        ("Open an article from anonymous Home.", "Article detail is visible."),
        ("Press Back once.", "The article closes and Home is restored."),
        ("Verify Home and Subscribe.", "The user returns to the anonymous Home state."),
    ]),
}


def main():
    workbook = load_workbook(WORKBOOK)
    sheet = workbook["Anonymous Home"]
    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}
    rows = {}
    for row in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row, headers["test_case_id"]).value or "").strip()
        rows.setdefault(case_id, []).append(row)

    for case_id, (name, steps) in CASES.items():
        target_rows = rows.get(case_id, [])
        if len(target_rows) != len(steps):
            raise ValueError(f"{case_id}: expected {len(steps)} rows, found {len(target_rows)}")
        for number, (row, (action, expected)) in enumerate(zip(target_rows, steps), 1):
            sheet.cell(row, headers["name"], name)
            sheet.cell(row, headers["module"], "Home")
            sheet.cell(row, headers["user_state"], "ANONYMOUS")
            sheet.cell(row, headers["step_number"], number)
            sheet.cell(row, headers["step"], action)
            sheet.cell(row, headers["expected_result"], expected)

    workbook.save(WORKBOOK)
    print(f"Rewrote {sum(len(value[1]) for value in CASES.values())} steps across {len(CASES)} cases")


if __name__ == "__main__":
    main()

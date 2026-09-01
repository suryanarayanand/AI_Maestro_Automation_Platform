from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT = Path("Uploads/Ready/Anonymous_Photos_Quick_Access_Approved_Test_Cases.xlsx")
HEADERS = ["test_case_id", "name", "module", "user_state", "precondition", "step_number", "step", "expected_result", "priority", "automatable", "runtime_assertion"]
PRE = "The Hindu app is installed; network is available; the user is signed out; gallery content and subscription prompts may vary between runs."

cases = [
    ("ANON_PHOTO_001", "Photos quick section launch and selected tab", "Critical", [
        ("Launch anonymously, skip the welcome popup when present, and tap Photos in Quick Access.", "The Photos grid opens.", "assertVisible Photos and screen_home"),
        ("Verify Photos is highlighted and capture the top of the grid.", "Photos is visibly selected and evidence is saved.", "assert selected Photos; screenshot"),
    ]),
    ("ANON_PHOTO_002", "Photos refresh and gallery reload", "High", [
        ("Open Photos and pull down once from the gallery grid.", "Refresh completes without leaving Photos.", "swipe DOWN; wait"),
        ("Verify gallery cards, images and count badges reload without overlap.", "The refreshed grid remains usable.", "assert Photos and gallery cards; screenshot"),
    ]),
    ("ANON_PHOTO_003", "Photos grid layout and count badges", "High", [
        ("Inspect multiple visible cards across at least three viewports.", "Mixed-size gallery cards render with headline overlays and images.", "repeat swipe UP 3 with waits"),
        ("Verify each available gallery count badge is a positive number.", "Count badges such as 1 or a multi-image total are readable.", "assert numeric gallery badge; screenshot"),
    ]),
    ("ANON_PHOTO_004", "Photos listing has no inline advertisements", "Critical", [
        ("Scroll through at least five Photos-grid viewports with waits between gestures.", "The gallery grid continues loading normally.", "repeat swipe UP 5 with waits"),
        ("Verify no Advertisement card, inline ad or Support Journalism ad block appears between gallery cards.", "The Photos listing contains gallery content without inline advertisements.", "assertNotVisible Advertisement and Support Journalism between cards"),
        ("If an advertisement is found, capture it and classify the result as a product bug requiring review.", "Unexpected listing-ad evidence is saved as NEEDS_REVIEW, not an automation error.", "conditional screenshot; NEEDS_REVIEW"),
    ]),
    ("ANON_PHOTO_005", "Single-photo article presentation", "High", [
        ("Open a current gallery card whose count badge is 1.", "A single-photo detail view opens.", "select dynamic count-1 card"),
        ("Verify one image, its article description/headline and available photo credit are displayed.", "Single-photo content is presented as an article/image without gallery paging.", "assert image and description/headline; screenshot"),
        ("Verify a left or right swipe does not advance to a second image counter.", "The counter remains single-image and no false gallery page is created.", "swipe LEFT; assert counter remains 1/1 or no multi-page counter"),
    ]),
    ("ANON_PHOTO_006", "Multi-photo bidirectional swipe and counter", "Blocker", [
        ("Open a current gallery card with a count greater than 1 and close a subscription modal if it covers the gallery.", "The first gallery image and current/total counter are visible.", "select multi-image card; conditionally close modal; assert 1/N"),
        ("Swipe left twice, waiting after each swipe.", "The image changes and the current counter increments by one on each swipe.", "swipe LEFT; assert 2/N; swipe LEFT; assert 3/N"),
        ("Swipe right once.", "The previous image returns and the counter decrements to 2/N.", "swipe RIGHT; assert 2/N; screenshot"),
    ]),
    ("ANON_PHOTO_007", "Gallery image description and photo credit", "High", [
        ("Open a multi-photo gallery and inspect the first image.", "An image-specific description and available PHOTO credit are readable.", "assert description and PHOTO/credit"),
        ("Swipe to the next image and compare its supporting text.", "The next image renders with its corresponding description/credit without overlap.", "swipe LEFT; assert text and image; screenshot"),
    ]),
    ("ANON_PHOTO_008", "Gallery subscription popup handling", "Critical", [
        ("Open a gallery and detect the optional subscription popup over the image.", "The flow branches safely whether the popup appears or not.", "conditional popup branch"),
        ("When present, capture the popup and verify its Subscribe action and close icon.", "The subscription prompt is documented and actionable.", "assert Subscribe and Close; screenshot"),
        ("Tap only the popup close icon and verify the gallery remains open at its original counter.", "The popup closes without closing the gallery or app.", "tap popup Close; assert current/total counter"),
    ]),
    ("ANON_PHOTO_009", "Read Full Article expansion", "Critical", [
        ("Open a gallery, dismiss an optional subscription popup, and swipe when necessary until Read Full Article is visible.", "READ FULL ARTICLE is available below the gallery description.", "scrollUntilVisible Read Full Article"),
        ("Tap Read Full Article.", "The complete article text expands within the gallery view.", "tapOn Read Full Article; assert expanded content; screenshot"),
    ]),
    ("ANON_PHOTO_010", "Read Less article collapse", "Critical", [
        ("Expand the gallery using Read Full Article.", "The full article body is visible.", "tapOn Read Full Article"),
        ("Scroll through the expanded content until Read Less is visible.", "READ LESS becomes reachable after the expanded article content.", "scrollUntilVisible Read Less direction DOWN"),
        ("Tap Read Less and verify the content collapses back to the gallery summary.", "The shortened gallery article presentation is restored.", "tapOn Read Less; assert Read Full Article or summary; screenshot"),
    ]),
    ("ANON_PHOTO_011", "Gallery Bookmark anonymous login gate", "Critical", [
        ("Open a gallery and tap the Bookmark icon in the fixed header.", "Anonymous bookmarking invokes authentication.", "tap header Bookmark"),
        ("Verify Login to your account or a sign-in prompt and capture it.", "Bookmark access is correctly restricted.", "assert Login/Sign in; screenshot; back"),
    ]),
    ("ANON_PHOTO_012", "Gallery Share functionality", "High", [
        ("Open a gallery and tap Share in the fixed header.", "The system share sheet opens.", "tap header Share; assert share sheet"),
        ("Capture the share sheet, then close it with Back.", "The same gallery and counter are restored.", "screenshot; back; assert current/total counter"),
    ]),
    ("ANON_PHOTO_013", "Gallery Comment anonymous login gate", "Critical", [
        ("Open a gallery and tap Comment in the fixed header.", "The comment surface or authentication requirement opens.", "tap header Comment"),
        ("Verify Sign in/Login is required for the anonymous user and capture it.", "Anonymous commenting is protected.", "assert Sign in/Login; screenshot; close/back"),
    ]),
    ("ANON_PHOTO_014", "Gallery Subscribe and Close actions", "Critical", [
        ("Open a gallery and tap Subscribe in the fixed header.", "The subscription offer/plan page opens.", "tap header Subscribe"),
        ("Verify monthly/yearly plan or offer information, then return.", "The Subscribe destination is valid and the gallery is restored.", "assert Monthly/Yearly/offer; back"),
        ("Tap the gallery Close icon.", "The Photos grid returns with Photos still selected.", "tap header Close; assert Photos selected; screenshot"),
    ]),
]

wb = Workbook()
ws = wb.active
ws.title = "Anonymous Photos Quick"
ws.append(HEADERS)
for case_id, name, priority, steps in cases:
    for number, (step, expected, assertion) in enumerate(steps, 1):
        ws.append([case_id, name, "Photos Quick Access", "ANONYMOUS", PRE, number, step, expected, priority, "Yes", assertion])

for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor="9C0006")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
for index, width in enumerate([20, 48, 22, 16, 74, 14, 88, 88, 12, 14, 65], 1):
    ws.column_dimensions[chr(64 + index)].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
ws.row_dimensions[1].height = 28
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(f"Created {OUTPUT} with {len(cases)} cases and {ws.max_row - 1} executable steps")

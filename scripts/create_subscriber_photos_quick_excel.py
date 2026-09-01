from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT = Path("Uploads/Ready/Subscriber_Photos_Quick_Access_Approved_Test_Cases.xlsx")
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "runtime_assertion",
]
PRE = (
    "The Hindu app is installed; network is available; valid subscriber credentials "
    "are configured securely; the subscriber is signed in; current gallery content may vary."
)

cases = [
    ("SUB_PHOTO_001", "Photos launch, selection and subscriber entitlement", "Blocker", [
        ("Open subscriber Home and tap Photos in Quick Access.", "The Photos listing opens and Photos is highlighted.", "assert Photos selected; screenshot after wait"),
        ("Verify subscriber identity remains active on the Photos listing.", "No Subscribe, login gate or paywall is shown.", "assertNotVisible Subscribe/Login/Keep reading"),
    ]),
    ("SUB_PHOTO_002", "Photos refresh and content recovery", "High", [
        ("Pull down to refresh Photos twice, waiting for loading to settle each time.", "The listing reloads without leaving Photos or duplicating cards.", "swipe DOWN twice; wait; assert Photos"),
        ("Verify visible cards have usable images/headlines and no request-timeout or blank-content state.", "Refreshed gallery content renders correctly.", "assertNotVisible Request timeout; screenshot after wait"),
    ]),
    ("SUB_PHOTO_003", "Photos grid layout, badges and pagination", "High", [
        ("Inspect at least three listing viewports and record available photo-count badges.", "Cards, headlines, images and positive count badges do not overlap or clip.", "repeat scroll 3 with wait; screenshot"),
        ("When SHOW MORE is available, tap it and wait for the listing to extend.", "Additional unique gallery cards load without resetting the section.", "conditional tap SHOW MORE; assert Photos"),
    ]),
    ("SUB_PHOTO_004", "Subscriber Photos contains no monetization", "Blocker", [
        ("Scroll through at least five Photos listing viewports.", "The listing continues loading gallery cards.", "repeat scroll 5 with waits"),
        ("Check each viewport for inline, sticky, interstitial, blank or Taboola advertising and subscription prompts.", "No advertising or Subscribe/paywall surface is present for the subscriber.", "assertNotVisible Advertisement/ad_iframe/Taboola/Subscribe; screenshots"),
        ("If monetization appears, capture evidence and mark the result as a product bug requiring review.", "Entitlement leakage is reported separately from automation failure.", "conditional screenshot; NEEDS_REVIEW"),
    ]),
    ("SUB_PHOTO_005", "Single-photo gallery presentation", "Critical", [
        ("Open a current card with a count of 1.", "A single-photo detail opens with its image, headline/description and available credit.", "select dynamic count-1 gallery; assert detail; screenshot after wait"),
        ("Swipe left and right once.", "No second image is created and the counter stays 1/1 or remains absent.", "swipe LEFT; swipe RIGHT; assert single-photo state"),
        ("Verify the image and supporting text remain aligned and readable.", "No blank image, clipping or overlay collision occurs.", "visual assertion; screenshot after wait"),
    ]),
    ("SUB_PHOTO_006", "Multi-photo gallery forward paging and counter", "Critical", [
        ("Open a current card with a count greater than 1.", "The gallery opens at the first image with a valid current/total counter.", "select dynamic multi-photo gallery; capture initial counter"),
        ("Swipe left through at least three available images with a wait after each swipe.", "The image changes once per swipe and the current counter increments without skipping or duplication.", "repeat swipe LEFT 3; compare counters; screenshots"),
    ]),
    ("SUB_PHOTO_007", "Multi-photo gallery reverse paging and boundaries", "Critical", [
        ("From a later gallery image, swipe right back toward the first image.", "The counter decrements once per swipe and matching images return.", "repeat swipe RIGHT; compare counters"),
        ("Attempt one extra right swipe at the first image and one extra left swipe at the last image.", "Paging does not underflow, overflow, crash or show an invalid counter.", "assert boundary counter and detail visible; screenshot"),
    ]),
    ("SUB_PHOTO_008", "Gallery image metadata synchronization", "High", [
        ("Record the visible description/caption and available photo credit on one gallery image.", "Metadata belongs to the displayed image and is readable.", "assert image plus description/credit"),
        ("Swipe to the next image and inspect the supporting text.", "Image, counter and metadata update together without stale or overlapping text.", "swipe LEFT; visual compare; screenshot after wait"),
    ]),
    ("SUB_PHOTO_009", "Rapid paging and image-load resilience", "High", [
        ("In a multi-photo gallery, perform three deliberate forward swipes followed by three reverse swipes.", "The gallery remains responsive and returns to the expected counter.", "swipe LEFT 3 and RIGHT 3 with bounded waits"),
        ("Check for blank, stretched, duplicated or permanently loading images after paging.", "Every settled page has a rendered image and valid counter.", "visual assertion; screenshot after wait"),
    ]),
    ("SUB_PHOTO_010", "Subscriber gallery header controls", "Critical", [
        ("Open a gallery and inspect its fixed header.", "Close, Bookmark, Share and Comment controls are visible and usable.", "assert Close/Bookmark/Share/Comment; screenshot"),
        ("Verify Subscribe is absent from the subscriber gallery header.", "Subscriber entitlement is reflected in the header.", "assertNotVisible Subscribe"),
    ]),
    ("SUB_PHOTO_011", "Subscriber gallery bookmark persistence", "Critical", [
        ("Tap Bookmark in a gallery and wait for the saved state.", "The gallery is bookmarked without opening Login or a subscription page.", "tap Bookmark; assert saved; assertNotVisible Login/Subscribe"),
        ("Close and reopen the same gallery, then remove the bookmark.", "Saved state persists and can be removed successfully.", "reopen gallery; assert saved; tap Bookmark; assert removed"),
    ]),
    ("SUB_PHOTO_012", "Gallery share sheet and safe return", "High", [
        ("Tap Share in a multi-photo gallery.", "The system share sheet opens over the app.", "tap Share; assert share sheet; screenshot after wait"),
        ("Dismiss the share sheet with Back.", "The same gallery image and counter are restored without reload.", "back; assert captured counter/detail"),
    ]),
    ("SUB_PHOTO_013", "Subscriber gallery comment access", "Critical", [
        ("Tap Comment from a gallery header.", "The comment surface opens without a Login gate.", "tap Comment; assert comment UI; assertNotVisible Login"),
        ("Return without submitting text.", "The gallery is restored and no comment is posted.", "back/close; assert gallery detail"),
    ]),
    ("SUB_PHOTO_014", "Read Full Article expansion for subscriber", "Blocker", [
        ("Scroll within a gallery until Read Full Article is visible and tap it.", "The complete article expands in the current gallery experience.", "scrollUntilVisible Read Full Article; tap"),
        ("Scroll through expanded content and verify it is fully readable.", "No paywall, Subscribe prompt or Already a subscriber login gate interrupts the article.", "assertNotVisible paywall/Subscribe/Login; screenshots after waits"),
    ]),
    ("SUB_PHOTO_015", "Read Less collapse and state restoration", "Critical", [
        ("Expand the gallery article, then scroll until Read Less is visible.", "Read Less is reachable after expanded content.", "tap Read Full Article; scrollUntilVisible Read Less"),
        ("Tap Read Less.", "The body collapses to its summary and Read Full Article becomes available again.", "tap Read Less; assert Read Full Article; screenshot after wait"),
        ("Verify the gallery image/counter remains at the pre-expansion state.", "Expanding and collapsing does not reset gallery paging.", "assert captured counter"),
    ]),
    ("SUB_PHOTO_016", "Gallery close, Back and listing-position recovery", "High", [
        ("Open a gallery from a lower listing viewport and close it with the header Close icon.", "The Photos listing returns near its prior scroll position with Photos selected.", "tap Close; assert Photos and prior card"),
        ("Reopen the gallery and use Android Back.", "Back also restores the Photos listing without app crash or Home redirection.", "back; assert Photos; assertNotVisible crash"),
    ]),
    ("SUB_PHOTO_017", "Photos light-theme visual integrity", "Medium", [
        ("Set Light mode, return to Photos and inspect the listing and a gallery detail.", "Images, counters, headlines, metadata and controls have readable contrast.", "set Light mode; screenshots after waits"),
        ("Expand and collapse Read Full Article in Light mode.", "Text and controls remain visible with no theme-related layout break.", "tap Read Full Article/Read Less; visual assertion"),
    ]),
    ("SUB_PHOTO_018", "Photos dark-theme visual integrity", "Medium", [
        ("Set Dark mode, return to Photos and inspect the listing and a gallery detail.", "Images, counters, headlines, metadata and controls have readable contrast.", "set Dark mode; screenshots after waits"),
        ("Swipe between images and expand the article in Dark mode.", "No white flash, invisible control, clipped text or stale theme appears.", "swipe; expand; visual assertion; screenshot"),
    ]),
]

wb = Workbook()
ws = wb.active
ws.title = "Subscriber Photos Quick"
ws.append(HEADERS)
for case_id, name, priority, steps in cases:
    for number, (step, expected, assertion) in enumerate(steps, 1):
        ws.append([
            case_id, name, "Photos", "SUBSCRIBER", PRE, number, step, expected,
            priority, "Yes", assertion,
        ])

for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor="9C0006")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
for index, width in enumerate([20, 52, 20, 16, 78, 14, 90, 90, 12, 14, 68], 1):
    ws.column_dimensions[chr(64 + index)].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
ws.row_dimensions[1].height = 28
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(f"Created {OUTPUT} with {len(cases)} cases and {ws.max_row - 1} executable steps")

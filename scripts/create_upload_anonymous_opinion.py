from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from web.services.generation_service import create_drafts

OUT = ROOT / "Uploads" / "Ready" / "Anonymous_Opinion_Quick_Access_Approved_Test_Cases.xlsx"
MODULE = "Opinion Quick Access"
STATE = "ANONYMOUS"
PRE = ("The Hindu app is installed; network is available; the user is signed out; "
       "Opinion article entitlement and optional features may vary by live content.")

cases = [
    ("ANON_OPINION_001", "Opinion landing and selected Quick Access tab", "Critical", [
        ("Launch anonymously, skip the welcome popup when present, and open Opinion from Quick Access.", "Opinion opens and its Quick Access label is selected.", "assert Opinion and screen_home; screenshot"),
        ("Verify the Opinion page title and initial content are visible.", "Opinion content loads without a blank or error state.", "assert Opinion and visible content"),
    ]),
    ("ANON_OPINION_002", "Opinion refresh retains section context", "Major", [
        ("Open Opinion and refresh the page twice with a settle wait after each refresh.", "Both refreshes complete and Opinion remains selected.", "swipe DOWN twice; wait; assert Opinion; screenshot"),
    ]),
    ("ANON_OPINION_003", "Opinion feed scrolling and visual integrity", "Major", [
        ("Scroll through at least five Opinion feed viewports, waiting before each screenshot.", "Headlines and cards remain readable with no overlap, clipping, blank block, or stuck loader.", "repeat swipe UP 5; wait and screenshot each viewport"),
    ]),
    ("ANON_OPINION_004", "Opinion advertisement handling", "Critical", [
        ("Scroll the Opinion feed and observe inline, sticky, blank, or interstitial advertisements.", "Sticky ads may remain; blank or obstructive ads are captured as bug evidence.", "scroll; conditional screenshot for Advertisement"),
        ("If an interstitial appears, capture it, wait at least five seconds for Close, and close only the ad.", "The ad closes and the app remains on Opinion.", "conditional screenshot; wait >=5000; tap close; assert Opinion"),
    ]),
    ("ANON_OPINION_005", "Open Opinion article and Read Full Article", "Critical", [
        ("Open a visible Opinion item and tap Read Full Article when that control is displayed.", "The article detail page opens.", "tap visible Opinion item; conditional tap Read Full Article; assert screen_article_detail"),
        ("Validate the Opinion article header and reading-time label when present.", "Article identity and available reading-time information render correctly.", "assert article detail; optional Min Read; screenshot"),
    ]),
    ("ANON_OPINION_006", "Premium Opinion article paywall restriction", "Blocker", [
        ("Open Opinion articles until a Premium-labelled or paywalled article is found.", "A qualifying restricted article is identified, or the run records that no premium article was found.", "conditional Premium/paywall branch; screenshot"),
        ("For a restricted article, scroll to the access boundary.", "The paywall blocks further reading and shows Subscribe plus Already a subscriber/Login.", "assert Subscribe and Already.*subscriber|Login; screenshot"),
    ]),
    ("ANON_OPINION_007", "Free Opinion article completion", "Critical", [
        ("When an Opinion article has no paywall, scroll through the complete article.", "The anonymous user can reach the article end.", "conditional no-paywall branch; repeat swipe UP; screenshots"),
        ("Verify available end-of-article content.", "Post a Comment and Related/Headlines appear when supported by that article.", "conditional assert Post.*Comment and Related|Headlines"),
    ]),
    ("ANON_OPINION_008", "Opinion AI Summary entitlement", "Critical", [
        ("If AI Summary is available, open it and wait for the sheet to settle.", "AI Summary opens without freezing or covering the app incorrectly.", "conditional tap AI summary; wait; screenshot"),
        ("Validate the entitlement branch.", "Paywalled content shows Subscribe; otherwise the full summary is visible. Absence is recorded when the article is ineligible.", "conditional Subscribe versus summary content; screenshot"),
    ]),
    ("ANON_OPINION_009", "Opinion Listen to Article control", "Major", [
        ("If Listen to Article is available, open it and verify playback controls.", "The audio panel and Play/Pause control respond correctly.", "conditional tap Listen to Article; assert Play|Pause; screenshot"),
        ("If the feature is absent, record it as content-dependent rather than an automation failure.", "The case remains valid for articles that do not support audio.", "conditional optional-feature handling"),
    ]),
    ("ANON_OPINION_010", "Opinion reading options and text size", "Major", [
        ("Open Reading Options/Text size from an Opinion article.", "The reading-options panel is displayed.", "tap Text size|Reading options; screenshot"),
        ("Change the text size, verify a visible article-text change, then close the panel.", "Text sizing works and the article remains open.", "tap size option; visual assertion; close; assert article detail"),
    ]),
    ("ANON_OPINION_011", "Opinion bookmark anonymous sign-in restriction", "Critical", [
        ("Tap Bookmark on an Opinion article.", "The anonymous user is directed to the sign-in page or sign-in prompt.", "tap Bookmark; assert Login|Sign in; screenshot"),
        ("Return from sign-in.", "The app returns safely to the Opinion article.", "back; assert screen_article_detail"),
    ]),
    ("ANON_OPINION_012", "Opinion share sheet", "Major", [
        ("Tap Share on an Opinion article and wait before capturing evidence.", "The native share sheet opens with available targets.", "tap Share; wait; screenshot"),
        ("Dismiss the share sheet.", "The Opinion article remains open and usable.", "back; assert screen_article_detail"),
    ]),
    ("ANON_OPINION_013", "Opinion comment anonymous sign-in restriction", "Critical", [
        ("Tap Comment/Post a Comment when available.", "The comment view displays a sign-in requirement for the anonymous user.", "conditional tap Comment|Post a Comment; assert Login|Sign in; screenshot"),
        ("Return to the article.", "The Opinion article is restored without an app exit.", "back; assert screen_article_detail"),
    ]),
    ("ANON_OPINION_014", "Opinion Subscribe plans and subscriber login", "Blocker", [
        ("Tap the article Subscribe control.", "The subscription offer opens.", "tap Subscribe; assert Yearly and Monthly; screenshot"),
        ("Scroll until Already a subscriber/Login is visible and tap it.", "The Login to your account page opens.", "scrollUntilVisible Already.*subscriber|Login; tap; assert Login to your account; screenshot"),
    ]),
    ("ANON_OPINION_015", "Opinion Cartoon section navigation", "Major", [
        ("Scroll Opinion until Cartoon is visible and open it.", "The Cartoon section or item opens from Opinion.", "scrollUntilVisible Cartoon; tap Cartoon; wait; screenshot"),
        ("Scroll its available content and return.", "Cartoon content remains visually stable and Back returns to Opinion.", "repeat swipe UP; screenshots; back; assert Opinion"),
    ]),
    ("ANON_OPINION_016", "Quick Access swipe away and return to Opinion", "Major", [
        ("From Opinion, swipe horizontally to an adjacent Quick Access section and return to Opinion.", "The adjacent section changes, then Opinion is restored and selected.", "swipe LEFT; wait; swipe RIGHT; assert Opinion; screenshot"),
    ]),
]

wb = Workbook()
ws = wb.active
ws.title = "Anonymous Opinion Quick"
headers = ["test_case_id", "name", "module", "user_state", "precondition",
           "step_number", "step", "expected_result", "priority", "automatable",
           "runtime_assertion"]
ws.append(headers)
for case_id, name, priority, steps in cases:
    for number, (step, expected, assertion) in enumerate(steps, 1):
        ws.append([case_id, name, MODULE, STATE, PRE, number, step, expected,
                   priority, "Yes", assertion])
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="8B0000")
    cell.alignment = Alignment(horizontal="center")
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
widths = {"A": 22, "B": 48, "C": 25, "D": 14, "E": 72, "F": 12,
          "G": 80, "H": 76, "I": 12, "J": 14, "K": 72}
for column, width in widths.items():
    ws.column_dimensions[column].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)

ids, result = create_drafts(OUT, use_ai=True)
print(f"workbook={OUT}")
print(f"cases={result.case_count} steps={result.step_count} drafts={len(ids)} ids={ids}")

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Uploads" / "Ready" / "Subscriber_Podcast_Quick_Access_Approved_Test_Cases.xlsx"
HEADERS = [
    "test_case_id", "name", "module", "user_state", "precondition",
    "step_number", "step", "expected_result", "priority", "automatable",
    "master_sheet_reference", "existing_yaml_reference",
]
PRE = (
    "The app, network and Maestro device are available; valid subscriber credentials come "
    "from protected portal configuration. Wait for loading to settle before every screenshot."
)
MASTER = (
    "TH App Testing Scenarios_AutomationCopy.xlsx: Podcast 2634-2686 and user-state 6338; "
    "az_recorder_20260828_105131.mp4"
)
YAML_REF = (
    "SC_15_Podcast_Section_Subscriber_account.yaml; ProductionBaseline/LOC_podcast.yaml; "
    "OptimizedLocatorSmoke/LOC_podcast.yaml; podcast_page_20260803_143843.json"
)

CASES = [
    ("SUB_PODCAST_001", "Podcast launch and selected tab", "Blocker", [
        ("Log in as a subscriber and tap Podcast in Quick Access.", "The Podcast listing opens and Podcast is visibly selected."),
        ("Wait for the listing to settle and capture its top viewport.", "Podcast series content is displayed without a blank page or request error."),
    ]),
    ("SUB_PODCAST_002", "Podcast listing content and visual layout", "Critical", [
        ("Inspect multiple visible Podcast cards.", "Each available series shows an undistorted cover image, title, readable description and supported action."),
        ("Scroll through at least three viewports with waits and screenshots.", "Cards and text remain aligned without clipping, overlap or duplicate content."),
    ]),
    ("SUB_PODCAST_003", "Podcast refresh and content recovery", "High", [
        ("Pull down to refresh the Podcast listing twice, waiting after each refresh.", "The listing reloads and remains on Podcast without freezing."),
        ("Verify series cards and actions are usable after refresh.", "Refresh does not create blank cards, permanent loaders or lost controls."),
    ]),
    ("SUB_PODCAST_004", "Subscriber Podcast has no monetization block", "Blocker", [
        ("Traverse the Podcast listing, one series page and the mini-player.", "Subscriber Podcast content remains accessible."),
        ("Verify advertisements, Subscribe, Login and paywall prompts are absent.", "No monetization surface blocks Podcast content or playback."),
    ]),
    ("SUB_PODCAST_005", "Selected podcast episode-list loading", "Blocker", [
        ("Tap SEE ALL EPISODES for The Rearview or another current series.", "The selected series page opens with matching title and artwork."),
        ("Wait up to 30 seconds for episode content, taking evidence at 15 and 30 seconds.", "Episode rows load; the page must not remain blank with an endless spinner."),
        ("If the spinner remains with no episodes, record FAIL as a product bug.", "The confirmed blank-loading defect is not classified as NEEDS_REVIEW or an automation error."),
    ]),
    ("SUB_PODCAST_006", "Episode-list identity, ordering and metadata", "Critical", [
        ("Open SEE ALL EPISODES for a series that loads successfully.", "The page title and branding match the selected series."),
        ("Inspect several episode rows.", "Episodes display title, thumbnail, duration and play control in the configured order."),
    ]),
    ("SUB_PODCAST_007", "Play Latest Episode and mini-player identity", "Blocker", [
        ("Tap Play Latest Episode on a current Podcast card.", "Playback starts without navigating away and the mini-player appears."),
        ("Wait up to 30 seconds and compare the mini-player identity with the selected series/episode.", "The correct title appears and the mini-player does not remain on an endless loader or show stale content."),
        ("Capture a mismatch, blank player or endless spinner as a product-bug FAIL.", "Incorrect player identity/loading is reported with screenshot evidence."),
    ]),
    ("SUB_PODCAST_008", "Repeated play and pause stability", "Critical", [
        ("Start a playable episode and wait for playback to settle.", "The Play control changes to Pause and elapsed time begins advancing."),
        ("Tap Pause and Play three times with short waits between each action.", "Every tap changes state exactly once without duplicate audio, freezing or losing the player."),
        ("Leave the episode paused and capture the settled player.", "The elapsed position remains stable while paused."),
    ]),
    ("SUB_PODCAST_009", "Timed playback progress movement", "Blocker", [
        ("Record the elapsed time/progress position, tap Play and wait 30 seconds.", "Playback remains active for the observation period."),
        ("Record the new elapsed time/progress position.", "Elapsed time and the progress bar advance by approximately 30 seconds."),
        ("Pause and wait 10 seconds.", "Elapsed time and progress do not continue moving while paused."),
    ]),
    ("SUB_PODCAST_010", "Forward control changes playback position", "Critical", [
        ("During playback, record the current elapsed position and tap Forward once.", "The elapsed position advances by the control's configured interval, such as 10 or 15 seconds."),
        ("Tap Forward three more times with waits.", "Each tap advances once without freezing, exceeding duration or restarting the episode."),
    ]),
    ("SUB_PODCAST_011", "Rewind control changes playback position", "Critical", [
        ("Move playback safely beyond the rewind interval and record the position.", "A non-zero position is available for rewind validation."),
        ("Tap Rewind once, then three more times with waits.", "Each tap moves backward by the configured interval without underflowing below zero or closing the player."),
    ]),
    ("SUB_PODCAST_012", "Podcast scrubber timeline seeking", "Blocker", [
        ("Start playback and drag the timeline scrubber to a safe later position.", "The thumb and elapsed time move to the selected position and playback resumes from there."),
        ("Drag the scrubber backward to an earlier safe position.", "The displayed time moves backward and playback continues from the selected point."),
        ("Pause, seek once more and resume.", "Seeking remains functional after pause/resume without resetting to the beginning."),
    ]),
    ("SUB_PODCAST_013", "Only one podcast plays at a time", "Blocker", [
        ("Start one Podcast episode and confirm progress is advancing.", "The first episode is the only active audio source."),
        ("Start a different series or episode.", "The first audio stops and the mini-player switches to the newly selected episode."),
        ("Verify no overlapping audio or stale title remains.", "Only the second episode plays and its metadata is displayed."),
    ]),
    ("SUB_PODCAST_014", "Playback continuity while scrolling", "High", [
        ("Start an episode and scroll through the Podcast listing or episode list.", "Playback continues and progress advances while content scrolls."),
        ("Return to the mini-player after scrolling.", "The same episode and valid playback position remain visible."),
    ]),
    ("SUB_PODCAST_015", "Mini-player close stops playback", "Critical", [
        ("Start playback, wait for progress, then tap the mini-player Close control.", "The mini-player closes and audio stops without closing the app."),
        ("Wait 10 seconds and reopen the selected episode.", "No hidden playback continued; the episode can be started again normally."),
    ]),
    ("SUB_PODCAST_016", "Podcast Back navigation and state recovery", "High", [
        ("Open a series episode list and use Android Back.", "The Podcast listing returns with Podcast still selected."),
        ("Repeat with the mini-player visible.", "Back navigation does not crash the app, duplicate the player or unexpectedly stop supported playback."),
    ]),
    ("SUB_PODCAST_017", "Podcast audio-load failure handling", "Critical", [
        ("Attempt playback under a temporarily unavailable or slow audio response.", "The player uses a bounded loading state."),
        ("Wait for the configured timeout.", "A useful error/Retry state appears instead of an endless spinner, and the app remains responsive."),
    ]),
    ("SUB_PODCAST_018", "Podcast rapid interaction resilience", "High", [
        ("Open and return from multiple series pages, then start different episodes in succession.", "The Podcast section remains responsive and shows the latest selected state."),
        ("Exercise Play, Pause, Forward, Rewind and scrubber after rapid navigation.", "Controls remain functional without stale metadata, duplicate audio or permanent loading."),
    ]),
]


def main():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Subscriber Podcast"
    sheet.append(HEADERS)
    for case_id, name, priority, steps in CASES:
        for number, (step, expected) in enumerate(steps, 1):
            sheet.append([
                case_id, name, "Podcast", "SUBSCRIBER", PRE, number, step,
                expected, priority, "Yes", MASTER, YAML_REF,
            ])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="273C75")
        cell.font = Font(color="FFFFFF", bold=True)
    for index, width in enumerate([22, 54, 18, 16, 88, 13, 96, 96, 14, 14, 76, 105], 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Created {OUTPUT} with {len(CASES)} cases and {sum(len(case[3]) for case in CASES)} steps")


if __name__ == "__main__":
    main()

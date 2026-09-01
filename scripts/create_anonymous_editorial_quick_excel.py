from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT = Path("Uploads/Ready/Anonymous_Editorial_Quick_Access_Approved_Test_Cases.xlsx")
HEADERS = ["test_case_id", "name", "module", "user_state", "precondition", "step_number", "step", "expected_result", "priority", "automatable", "runtime_assertion"]
PRE = "The Hindu app is installed; network is available; the user is signed out; Editorial article entitlement may vary dynamically."

cases = [
    ("ANON_EDITORIAL_001", "Editorial landing and exclusive elephant branding", "Critical", [
        ("Launch anonymously, skip the welcome popup when present, and tap Editorial in Quick Access.", "The Editorial landing page opens.", "assert Editorial landing and screen_home"),
        ("Verify Editorial is selected and the Editorial-only elephant symbol is displayed in the header.", "Editorial branding and its unique elephant symbol are visible.", "assert Editorial; visual screenshot of selected tab and elephant"),
    ]),
    ("ANON_EDITORIAL_002", "Editorial listing excludes thumbnail images", "Critical", [
        ("Inspect Editorial cards across at least three page viewports.", "Editorial headlines, summaries and publication details render as text-led cards.", "repeat swipe UP 3; assert article cards/text"),
        ("Verify Editorial listing cards do not display article thumbnail images.", "No thumbnail image is shown beside Editorial listing articles.", "assert no listing thumbnail; screenshots"),
    ]),
    ("ANON_EDITORIAL_003", "Editorial refresh and branding retention", "High", [
        ("Pull down once on the Editorial landing page.", "Editorial refresh completes without leaving the section.", "swipe DOWN; wait"),
        ("Verify the elephant branding, Editorial label and article list remain correctly rendered.", "Editorial reloads without UI distortion.", "assert Editorial/article list; screenshot"),
    ]),
    ("ANON_EDITORIAL_004", "Editorial page sections and pagination", "High", [
        ("Scroll through the complete Editorial listing.", "Featured Editorial, Recent Editorials and More Writes from Editorial appear when supplied.", "scroll and conditional section assertions"),
        ("Tap Show More when it is available.", "Additional Editorial articles load without duplicated or broken cards.", "conditional Show More; assert additional cards"),
        ("When no more content exists, verify Show More is absent or no longer actionable.", "Pagination ends gracefully.", "conditional end-state screenshot"),
    ]),
    ("ANON_EDITORIAL_005", "Editorial long-title and text layout", "Medium", [
        ("Inspect Editorial cards containing long headlines or summaries.", "Text wraps without overlapping labels, dates or adjacent cards.", "multi-viewport screenshots and visible text assertions"),
        ("Verify spacing and readability in the Editorial text-led layout.", "Headlines and summaries remain legible without clipping.", "visual screenshot evidence"),
    ]),
    ("ANON_EDITORIAL_006", "Editorial article opening and metadata", "Critical", [
        ("Open a current Editorial article using its card rather than a fixed headline.", "The corresponding article detail page opens.", "tap dynamic Editorial article; assert article screen"),
        ("Verify Editorial section name, title, author when supplied, and published/updated details.", "Editorial article metadata is readable and correctly associated.", "assert EDITORIAL/title/date and conditional author; screenshot"),
    ]),
    ("ANON_EDITORIAL_007", "Free Editorial full-content behaviour", "Blocker", [
        ("Check a bounded maximum of three current Editorial articles until a non-paywalled article is found.", "A free Editorial article is identified, or dynamic unavailability is recorded as SKIPPED with evidence.", "bounded article selection; free or SKIPPED screenshot"),
        ("Scroll through the free article to the end.", "The complete Editorial content remains readable without a paywall.", "assert article screen and paywall absent through scroll"),
        ("Verify Post a Comment, Related Topics, Recommended or Headlines only when supported by that article.", "Available free-article footer components render correctly.", "conditional footer assertions and screenshot"),
    ]),
    ("ANON_EDITORIAL_008", "Premium Editorial paywall restriction", "Blocker", [
        ("Check a bounded maximum of three current Editorial articles until a premium/paywalled article is found.", "A premium Editorial is identified without relying on a fixed headline.", "bounded article selection; detect Premium/paywall"),
        ("Verify the paywall blocks complete reading and displays a Subscribe/offer message.", "Anonymous access is restricted at the paywall.", "assert Keep reading/offer/Subscribe; screenshot"),
        ("Verify Already a subscriber? Login is available, scrolling within the offer when necessary.", "Existing subscribers have a Login route.", "scrollUntilVisible Already a subscriber/Login"),
    ]),
    ("ANON_EDITORIAL_009", "Editorial Reading Options functionality", "High", [
        ("Open an Editorial article and tap Reading Options or Text Size.", "The reading-options panel opens.", "tap Reading Options/Text Size; assert panel"),
        ("Change an available text-size setting and close the panel.", "The control responds and the Editorial article remains open.", "tap size; close; assert article"),
    ]),
    ("ANON_EDITORIAL_010", "Editorial Bookmark anonymous login gate", "Critical", [
        ("Open an Editorial article and tap Bookmark.", "Anonymous bookmarking opens authentication.", "tap Bookmark"),
        ("Verify Login to your account/Sign in, capture it, and return.", "Bookmark is protected and Back restores the article.", "assert Login/Sign in; screenshot; back"),
    ]),
    ("ANON_EDITORIAL_011", "Editorial Share functionality", "High", [
        ("Open an Editorial article and tap Share.", "The system share sheet opens.", "tap Share; assert share sheet"),
        ("Capture the share sheet and close it using Back.", "The same Editorial article is restored.", "screenshot; back; assert article"),
    ]),
    ("ANON_EDITORIAL_012", "Editorial Comment anonymous login gate", "Critical", [
        ("On a free Editorial article, tap Comment or Post a Comment when available.", "The comment surface opens.", "conditional tap Comment/Post a Comment"),
        ("Verify Sign in/Login is required for an anonymous user, then close it.", "Anonymous commenting is correctly protected.", "assert Sign in/Login; screenshot; close/back"),
    ]),
    ("ANON_EDITORIAL_013", "Editorial Subscribe and subscriber-login navigation", "Critical", [
        ("From an Editorial paywall or floating bar, tap Subscribe.", "The subscription offer/plan page opens.", "tap Subscribe"),
        ("Verify monthly/yearly plan or offer information, then return.", "Subscription navigation is correct.", "assert Monthly/Yearly/offer; back"),
        ("Tap Already a subscriber? Login and verify Login to your account.", "The existing-subscriber link opens the correct login page.", "tap Login; assert Login to your account; screenshot"),
    ]),
    ("ANON_EDITORIAL_014", "Editorial back navigation and retained position", "High", [
        ("Open an Editorial article from a scrolled listing position.", "The article opens from the selected card.", "scroll; tap dynamic card; assert article"),
        ("Use Back and verify Editorial remains selected near the previous listing position.", "Editorial context and reasonable scroll position are restored.", "back; assert Editorial; screenshot"),
    ]),
    ("ANON_EDITORIAL_015", "Editorial interstitial and article advertisement handling", "High", [
        ("Open Editorial articles and detect an optional interstitial advertisement.", "The flow branches safely whether an interstitial appears or not.", "conditional interstitial branch"),
        ("If shown, capture it, wait at least 5 seconds and wait until Close/Skip is available.", "Advertisement evidence is saved without prematurely closing the app.", "screenshot; wait 5 seconds; wait Close/Skip"),
        ("Close only the advertisement and verify Editorial content is restored.", "The article remains open after ad recovery.", "tap ad Close/Skip; assert article"),
    ]),
    ("ANON_EDITORIAL_016", "Editorial loading error and retry recovery", "Medium", [
        ("Refresh Editorial and observe the controlled loading state.", "Editorial either loads content or displays a clear Retry/No Internet state without crashing.", "assert article list or Retry/No Internet"),
        ("If Retry appears, tap it after network availability is restored.", "Editorial content and branding recover successfully.", "conditional tap Retry; assert Editorial/article list; screenshot"),
    ]),
]

wb = Workbook(); ws = wb.active; ws.title = "Anonymous Editorial Quick"; ws.append(HEADERS)
for cid, name, priority, steps in cases:
    for number, (step, expected, assertion) in enumerate(steps, 1):
        ws.append([cid, name, "Editorial Quick Access", "ANONYMOUS", PRE, number, step, expected, priority, "Yes", assertion])
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor="9C0006"); cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center", vertical="center")
for index, width in enumerate([22, 48, 24, 16, 74, 14, 88, 88, 12, 14, 65], 1): ws.column_dimensions[chr(64 + index)].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions; ws.row_dimensions[1].height = 28
OUTPUT.parent.mkdir(parents=True, exist_ok=True); wb.save(OUTPUT)
print(f"Created {OUTPUT} with {len(cases)} cases and {ws.max_row - 1} executable steps")

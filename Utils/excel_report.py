from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from pathlib import Path


def generate_excel_report(results, suite_name, execution_time, report_folder):

    wb = Workbook()

    ws = wb.active
    ws.title = "Dashboard"

    green = PatternFill(fill_type="solid", fgColor="92D050")
    red = PatternFill(fill_type="solid", fgColor="FF6666")
    review = PatternFill(fill_type="solid", fgColor="D9CCFF")
    blue = PatternFill(fill_type="solid", fgColor="4F81BD")

    bold = Font(bold=True, color="FFFFFF")

    passed = len([r for r in results if r["status"] == "PASS"])
    failed = len([r for r in results if r["status"] == "FAIL"])
    needs_review = len([r for r in results if r["status"] == "NEEDS_REVIEW"])
    missing = len([r for r in results if r["status"] == "NOT FOUND"])

    ws["A1"] = "Automation Dashboard"

    ws["A3"] = "Suite"
    ws["B3"] = suite_name

    ws["A4"] = "Execution Date"
    ws["B4"] = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    ws["A5"] = "Total Tests"
    ws["B5"] = len(results)

    ws["A6"] = "Passed"
    ws["B6"] = passed

    ws["A7"] = "Failed"
    ws["B7"] = failed

    ws["A8"] = "Needs Review"
    ws["B8"] = needs_review

    ws["A9"] = "Not Found"
    ws["B9"] = missing

    ws["A10"] = "Execution Time (sec)"
    ws["B10"] = execution_time

    # ======================================
    # Results Sheet
    # ======================================

    result_sheet = wb.create_sheet("Execution Results")

    headers = [
        "Scenario ID",
        "Module",
        "Scenario",
        "Status",
        "Duration (sec)"
    ]

    for col, header in enumerate(headers, start=1):

        cell = result_sheet.cell(row=1, column=col)

        cell.value = header
        cell.fill = blue
        cell.font = bold

    row = 2

    for r in results:

        result_sheet.cell(row=row, column=1).value = r["id"]
        result_sheet.cell(row=row, column=2).value = r["module"]
        result_sheet.cell(row=row, column=3).value = r["name"]
        result_sheet.cell(row=row, column=4).value = r["status"]
        result_sheet.cell(row=row, column=5).value = r["duration"]

        if r["status"] == "PASS":
            result_sheet.cell(row=row, column=4).fill = green

        elif r["status"] == "FAIL":
            result_sheet.cell(row=row, column=4).fill = red

        elif r["status"] == "NEEDS_REVIEW":
            result_sheet.cell(row=row, column=4).fill = review

        row += 1

    output = Path(report_folder) / "Dashboard.xlsx"

    wb.save(output)

    print(f"\nExcel Report Generated : {output}")
